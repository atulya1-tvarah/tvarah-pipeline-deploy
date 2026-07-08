"""
Expanded Scoring Sheet — replicates Scoring.xlsx format
but every parameter shows its sub-data-points explicitly.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Scoring - Expanded"

# ── helpers ──────────────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, color="000000", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)
def aln(h="center", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

thin  = Side(style="thin",   color="BDBDBD")
med   = Side(style="medium", color="7F7F7F")
thick = Side(style="medium", color="404040")
B  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
BM = Border(left=med,   right=med,   top=med,   bottom=med)
BT = Border(left=thick, right=thick, top=thick, bottom=thick)

def cell(r, c, v, bg="FFFFFF", bold=False, color="000000",
         size=9, ha="center", wrap=True, italic=False, border=None):
    cl = ws.cell(row=r, column=c, value=v)
    cl.fill = fill(bg)
    cl.font = fnt(bold=bold, color=color, size=size, italic=italic)
    cl.alignment = aln(ha, wrap)
    cl.border = border or B
    return cl

def merge(r, c1, c2, v, bg, fg="FFFFFF", sz=10, bold=True, ha="center"):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cl = ws.cell(row=r, column=c1, value=v)
    cl.fill = fill(bg)
    cl.font = fnt(bold=bold, color=fg, size=sz)
    cl.alignment = aln(ha, True)
    cl.border = BM

def widths(cols):
    for i, w in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── colour palette ────────────────────────────────────────────────────────
EXP_H  = "1B5E20"; EXP_R  = "E8F5E9"; EXP_A  = "C8E6C9"
EDU_H  = "4E342E"; EDU_R  = "FBE9E7"; EDU_A  = "FFCCBC"
SKL_H  = "311B92"; SKL_R  = "EDE7F6"; SKL_A  = "D1C4E9"
TOT_H  = "263238"; TOT_R  = "ECEFF1"
REC_BG = "E3F2FD"; PNL_BG = "FCE4EC"; RSM_BG = "F1F8E9"
HDR_BG = "1F3864"

# ─────────────────────────────────────────────────────────────────────────
# ROW DATA
# Each record:
# (section, stage_label, parameter_name, sub_data_points,
#  module, method, max_pts, rec_input, panel_input, notes)
#
# stage_label: "Resume" | "Recruiter" | "Panel" | "Resume+Recruiter" | etc.
# rec_input:   what recruiter fills / validates
# panel_input: what panel fills / validates
# ─────────────────────────────────────────────────────────────────────────

ROWS = []

def sec(label, bg):
    ROWS.append(("SEC", label, bg))

def row(section, stage, param, sub_dps, module, method, pts,
        rec_input="—", panel_input="—", notes=""):
    ROWS.append(("ROW", section, stage, param, sub_dps, module,
                 method, pts, rec_input, panel_input, notes))

# ═══════════════════════════════════════════════════════════════════════════
sec("EXPERIENCE  —  40 pts", EXP_H)
# ═══════════════════════════════════════════════════════════════════════════

row("EXP", "Resume\n+\nRecruiter Rescore",
    "Companies Worked With",
    "1. Company Name (official full name)\n"
    "2. Company Tier  —  1=FAANG / 2=Unicorn / 3=Mid-size / 4=IT Services / 5=Unknown\n"
    "3. Industry / Domain  —  BFSI | FINTECH | ECOMMERCE | HEALTHCARE | EDTECH | SAAS | LOGISTICS | MANUFACTURING | etc.\n"
    "4. Company Type  —  Product / Services / Consulting / Startup / MNC / Research\n"
    "5. Funding Stage  —  Bootstrapped / Seed / Series A / B / C / D+ / Pre-IPO / Listed / Govt\n"
    "6. Headcount Band  —  <50 / 50–200 / 200–1K / 1K–5K / 5K–50K / >50K\n"
    "7. Work Type  —  Product Build / Consulting Delivery / Outsourcing / Captive Centre\n"
    "8. Operating Model  —  Consulting / Product / Platform-Infra / Domain Specialist",
    "Experience", "Python + LLM Judge", "5 pts",
    "Confirm company type (Product vs Services), industry, funding stage, headcount if unknown",
    "—",
    "Best tier across career used. Tier 1=5pts, 2=4pts, 3=3pts, 4=2pts, 5=1pt")

row("EXP", "Resume",
    "Overall Experience / Relevant Experience",
    "1. Start Date per role  —  MM/YYYY (month + year both needed)\n"
    "2. End Date per role  —  MM/YYYY or 'Present'\n"
    "3. Total Years of Experience  —  sum of all non-overlapping roles\n"
    "4. Relevant Years of Experience  —  years in domain/role matching JD\n"
    "5. Domain Continuity  —  Continuous / Switcher / Serial Switcher\n"
    "6. Experience Band  —  Fresher / Junior / Mid / Senior / Seasoned / Veteran\n"
    "7. YoE Ratio vs JD  —  relevant_yoe / jd_yoe_min (must be >=0.70 to pass)",
    "Experience", "Python", "3 pts",
    "Confirm exact month/year if only year on resume. Confirm total if roles are missing.",
    "Confirm which roles are most relevant to this JD.",
    "REJECT if relevant YoE < 70% of JD minimum range")

row("EXP", "Resume\n+\nLLM Rescore",
    "Career Progression",
    "1. Job Title per role  —  exact title as on resume\n"
    "2. Seniority Level per title  —  1=IC / 2=Sr IC / 3=Lead / 4=Manager / 5=Director / 6=VP+\n"
    "3. Seniority Sequence  —  [1,2,3,4] growing / [3,3,3] flat / [4,3,2] declining\n"
    "4. Promotion at Same Company  —  Y/N  (same employer, higher title)\n"
    "5. Internal Title Changes Count  —  how many title changes at same employer\n"
    "6. BERT Class  —  FAST_TRACK / GROWING / LATERAL / DECLINING\n"
    "7. Title Velocity  —  seniority levels gained per year\n"
    "8. Same-Company Growth Signal  —  Y/N",
    "Experience", "Python + BERT + LLM Judge", "3 pts",
    "Confirm if title changes at same company were promotions. Clarify vague titles.",
    "How did scope and team size change from [Role A] to [Role B]?",
    "BERT >=0.60 conf: 60% BERT + 40% heuristic")

row("EXP", "Resume\n+\nLLM Judge",
    "Stability",
    "1. Duration per company  —  months in each role\n"
    "2. Average Tenure  —  mean months across all roles\n"
    "3. Short Stints Count  —  roles <12 months\n"
    "4. Very Short Stints Count  —  roles <6 months\n"
    "5. Job-Hopping Rate  —  companies changed per year\n"
    "6. Loyalty Signal  —  LOW / MEDIUM / HIGH\n"
    "7. Startup Exit Context  —  Shutdown / Acquired / Layoff / Voluntary / Unknown\n"
    "8. Contract / Freelance Pattern  —  Y/N  (Archetype A10 if 3+ short stints, same skill cluster)",
    "Experience", "Python + LLM Judge", "3 pts",
    "For every role <12 months: why did you leave so quickly? Layoff, shutdown, or voluntary?",
    "—",
    "36+m avg=5.0 | 24–36=4.0 | 18–24=3.5 | 12–18=3.0 | <12=penalty. >2 roles/yr = -0.5")

row("EXP", "Resume\n+\nLLM Judge",
    "Awards & Recognitions",
    "1. Named Awards  —  company award / industry award / academic prize\n"
    "2. Award Count  —  total number (0/1/2/3+)\n"
    "3. Award Type  —  COMPANY_AWARD / INDUSTRY_AWARD / HACKATHON_WIN / ACADEMIC_PRIZE\n"
    "4. Promotion Count  —  formal promotions with title + salary change\n"
    "5. Patent Count  —  filed or granted patents\n"
    "6. Publication Count  —  journal / conference papers\n"
    "7. Conference Talk Count  —  invited speaker slots\n"
    "8. LLM Verified  —  Genuine Competitive / Generic Participation / Unverifiable",
    "Experience", "Python + LLM Judge", "3 pts",
    "For each award: was it competitive? How many people were considered? What did you do to earn it?",
    "Tell me about [award] — what was the selection process and what specifically did you do?",
    "LLM filters generic participation badges. Only competitive/specific awards counted.")

row("EXP", "Resume\n+\nRecruiter\n(Recruiter fills)",
    "Mentorship / Code Reviews / Interviews",
    "1. Mentored Engineers Count  —  how many juniors mentored\n"
    "2. Mentorship Type  —  Formal (assigned) / Informal\n"
    "3. Number of Roles with Mentorship  —  1 role vs 2+ roles\n"
    "4. Code Review Signal  —  Y/N  (reviewed PRs / code of others)\n"
    "5. Interview Panel Participation  —  Y/N  (conducted technical interviews)\n"
    "6. BERT Class  —  LEAD / FORMAL / IMPLIED / NONE\n"
    "7. Team Leadership  —  led a team of X people",
    "Experience", "Python + BERT + LLM Judge", "3 pts",
    "Have you mentored junior engineers? How many, at which companies, formal or informal?\nDid you do code reviews? Were you on the interview panel?",
    "Tell me about a specific person you mentored — what was the outcome for them?",
    "LEAD (2+ roles)=3pts | FORMAL (1 role)=2pts | IMPLIED=1pt | NONE=0pts")

row("EXP", "Resume\n+\nRecruiter Confirms",
    "International Exposure",
    "1. Onsite / Overseas Work  —  physically worked in another country  (strongest signal)\n"
    "2. Countries Worked In  —  list of countries\n"
    "3. Duration Overseas  —  months worked abroad\n"
    "4. Work Type Overseas  —  Client onsite / Relocation / Transfer / Conference\n"
    "5. Visa / Work Permit  —  H1B / L1 / UK visa / Work permit mentioned\n"
    "6. Global Team Signal  —  Y/N  (multi-timezone collaboration without physical travel)\n"
    "7. International Stakeholders  —  Y/N  (presented to / worked with overseas business stakeholders)",
    "Experience", "LLM Judge", "2 pts",
    "I see [country/onsite] on your resume — did you physically travel and work there? How long, and for what purpose?",
    "Tell me about your international work experience — country, duration, and type of engagement.",
    "Explicit onsite/relocation=2pts | Implied global team=1pt | None=0pts")

row("EXP", "Resume\n+\nRecruiter Validates",
    "Stakeholder Management",
    "1. Client-Facing Role  —  Y/N  (direct work with external clients)\n"
    "2. C-Level / Board Exposure  —  Y/N  (presented to or advised CTO / CEO / Board)\n"
    "3. Cross-Functional Collaboration  —  Y/N  (worked across product, business, ops, finance)\n"
    "4. External Stakeholder Count  —  number of client organisations managed\n"
    "5. Stakeholder Seniority  —  Junior / Mid / Senior / Executive\n"
    "6. BERT Class  —  NONE / INTERNAL / CLIENT_FACING / C_LEVEL",
    "Experience", "Python + BERT + LLM Judge", "2 pts",
    "Were you directly working with external clients? Did you present to or advise C-level leaders?",
    "Give me an example of a difficult stakeholder situation — who were they, what was the conflict, and how did you handle it?",
    "C_LEVEL=2pts | CLIENT_FACING=1.5pts | INTERNAL=1pt | NONE=0pts")

row("EXP", "Resume",
    "Career Breaks",
    "1. Break Count  —  number of gaps >3 months between consecutive roles\n"
    "2. Break Duration per Gap  —  months each gap lasted\n"
    "3. Break Start Date / End Date  —  when each gap occurred\n"
    "4. Break Reason  —  MBA / Maternity-Paternity / Illness / Layoff / Recession / Startup Failure / Travel / Voluntary / Unknown\n"
    "5. Education Overlap  —  Y/N  (gap covered by MBA / full-time degree = exempt)\n"
    "6. COVID Flag  —  Y/N  (gap in 2020 = context note, not penalised)",
    "Experience", "Python", "2 pts",
    "For every gap > 3 months: what were you doing? Was it education, parental leave, health, layoff, or your choice to take a break?",
    "—",
    "0 breaks=2pts | 1=1pt | 2=0pts | 3+=REJECT FLAG. MBA/maternity/layoff/COVID gaps = exempt from penalty.")

row("EXP", "Resume\n+\nRecruiter\n+\nPanel",
    "Project 1  —  Latest Project",
    "1. Project Title  —  name / headline of the project\n"
    "2. Project Type  —  Development / Migration / Analytics / Infrastructure / Research / Consulting / Maintenance\n"
    "3. About the Project  —  problem statement, what was built, why it was needed  (>20 chars minimum)\n"
    "4. Project Duration  —  how long the project ran  (>=3 months)\n"
    "5. Skills Used  —  specific technologies, frameworks, tools\n"
    "6. Skill Depth in Project  —  evidence of how deeply each skill was used (APPLIED / DEEP / EXPERT)\n"
    "7. Domain Depth  —  which business domain, industry context, end-user impact\n"
    "8. Role Played / Ownership  —  what YOU specifically built/designed/led  (ownership verb + >50 chars)\n"
    "9. Quantified Business Impact  —  number / % / money / scale  (e.g. 'reduced latency by 40%')\n"
    "10. LLM Complexity Score  —  0 to 5  (LLM PROJECT_JUDGE rates technical complexity)\n"
    "11. Candidate Signal  —  EXCELLENT / STRONG / AVERAGE / WEAK  (LLM holistic assessment)",
    "Experience", "Python + LLM PROJECT_JUDGE + Panel Feedback", "8 pts",
    "What is the project title? What were you building and why? How long did it run? What was YOUR specific role — not the team's? What was the measurable outcome?",
    "Walk me through the technical architecture of this project. Which components did you personally design? What were the hardest trade-offs?",
    "8 criteria x 1pt each. Most commonly missing: quantified impact (Criterion 9) and ownership clarity (Criterion 8).")

row("EXP", "Resume\n+\nRecruiter\n+\nPanel",
    "Project 2  —  2nd Latest Project",
    "1. Project Title  —  name / headline\n"
    "2. Project Type  —  Development / Migration / Analytics / Infrastructure / Research / Consulting / Maintenance\n"
    "3. About the Project  —  problem statement, what was built  (>20 chars)\n"
    "4. Project Duration  —  how long it ran  (>=3 months)\n"
    "5. Skills Used  —  specific technologies and tools\n"
    "6. Skill Depth in Project  —  evidence of depth per skill\n"
    "7. Domain Depth  —  which business domain, industry context",
    "Experience", "Python + LLM PROJECT_JUDGE + Panel Feedback", "6 pts",
    "Tell me about your second most significant project. Title, domain, duration, your role, and skills used.",
    "Walk me through the technical decisions you made in your second project.",
    "6 criteria x 1pt (same as Project 1 criteria 1–7 only — no ownership verb or quantified impact criteria).")

# ═══════════════════════════════════════════════════════════════════════════
sec("EDUCATION  —  15 pts  (10 core + 5 bonus)", EDU_H)
# ═══════════════════════════════════════════════════════════════════════════

row("EDU", "Resume",
    "Institutes  —  Tier, GPA, Stream",
    "1. Institution Name  —  full official name  (e.g. 'Indian Institute of Technology Bombay')\n"
    "2. Institute Tier  —  TIER_1 (IIT/IIM/ISI/ISB/global top-200) / TIER_2 (NIT/VIT/BITS/Manipal) / TIER_3 / TIER_4\n"
    "3. GPA / CGPA / Percentage  —  raw value as on resume  (e.g. 8.5/10, 75%, 3.9/4.0)\n"
    "4. GPA Scale  —  10-point / 4-point / 100-point\n"
    "5. GPA Band  —  EXCELLENT (>=8.5/10) / GOOD (7.5–8.4) / ACCEPTABLE (6.5–7.4) / LOW (<6.5)\n"
    "6. Degree Stream  —  CS / CSE / IT / Electronics / Mechanical / Statistics / Analytics / Management / Arts\n"
    "7. IT Stream Flag  —  Y/N  (CS, Data Science, AI/ML, MCA, Statistics, Analytics, Quant Econ, OR = YES)",
    "Education", "Python", "5 pts",
    "Confirm full institution name if abbreviation used. Ask for GPA if not listed.",
    "—",
    "TIER_1 base=4pts. +1 if GPA EXCELLENT or GOOD. TIER_2 base=3pts. +0.5 if EXCELLENT. Capped at 5.")

row("EDU", "Resume",
    "Highest Education  —  Degree Level + Stream",
    "1. Degree Level  —  PhD / Master / Bachelor / Diploma / Unknown\n"
    "2. Degree Name  —  B.Tech / M.Tech / M.Sc / MBA / MCA / Ph.D etc.\n"
    "3. Field of Study (Canonical)  —  Engineering / Computer Science-IT / Data Analytics / Management / Science / Arts\n"
    "4. Tech Fit  —  TECH / SEMI_TECH / NON_TECH\n"
    "5. Multiple Degrees  —  Y/N  (if yes: list all, use best for scoring)\n"
    "6. Dual-Degree Combination  —  e.g. B.Tech Non-IT + M.Tech IT = 1.5pts",
    "Education", "Python", "2 pts",
    "What is the exact degree name and level — Bachelor's, Master's, or PhD?",
    "—",
    "PhD/Master=2pts | Bachelor=1.5pts | Diploma=1pt | Unknown=0.5pts")

row("EDU", "Resume\n+\nRecruiter",
    "Education Gaps",
    "1. Education End Date  —  month + year degree was completed\n"
    "2. First Job Start Date  —  month + year first professional role began\n"
    "3. Gap Duration  —  months between education end and first job start\n"
    "4. Gap Reason  —  Competitive exam prep / Further studies / Job search / Personal / Unknown\n"
    "5. Education-During-Gap  —  Y/N  (pursuing another degree during the gap = exempt)",
    "Education", "Python + LLM Judge", "1 pt",
    "When exactly did you finish your degree? When did you start your first job? What were you doing in the gap?",
    "—",
    "<=6m=1pt | 6–12m=0.5pt | >12m=REJECT FLAG. Gap for MBA/GATE prep = context note.")

row("EDU", "Resume",
    "Education to Job Relevance",
    "1. Course Relevance Signal  —  HIGH / MEDIUM / FOUNDATIONAL / UNKNOWN\n"
    "   HIGH: CS / Engineering / MCA / Data Science / Statistics / Analytics / Quant Econ / Econometrics / PhD\n"
    "   MEDIUM: B.Sc / M.Sc / MBA / Management\n"
    "   FOUNDATIONAL: Arts / Commerce / Law\n"
    "2. Stream Relevance Rank  —  1 (ECE) to 5 (Civil) for non-CS streams\n"
    "3. Domain Switch  —  Y/N  (degree domain ≠ job domain across 2+ roles = Archetype A8)",
    "Education", "Python + LLM Judge", "2 pts",
    "Your degree is in [Field] — how did your academic training prepare you for a career in [Target Role]?",
    "—",
    "HIGH=2pts | MEDIUM=1.5pts | FOUNDATIONAL=0.5pts | UNKNOWN=1pt")

row("EDU", "Resume",
    "Executive Education / Distant Learning",
    "1. Programme Name  —  e.g. 'IIM Executive Programme in Data Science'\n"
    "2. Issuing Institution  —  recognised institution Y/N\n"
    "3. Mode  —  Executive / Distance / Online / MOOC\n"
    "4. Job Relevance  —  relevant to current target role Y/N",
    "Education", "Python or LLM Judge", "1 pt",
    "Have you done any executive education programmes, PG diplomas, or major online certifications after your degree?",
    "—",
    "1pt if recognised institution + relevant to role. Random Udemy courses do not qualify.")

row("EDU", "Resume",
    "Patents / Publications",
    "1. Patent Count  —  number of filed or granted patents\n"
    "2. Patent Title  —  name of the patent\n"
    "3. Patent Status  —  Filed / Granted\n"
    "4. Publication Count  —  journal or conference papers\n"
    "5. Publication Venue  —  IEEE / NeurIPS / ICML / KDD / arxiv / etc.\n"
    "6. Conference Talk Count  —  invited speaker at known tech conferences\n"
    "7. TIER_1 Bonus  —  TIER_1 institution gets 0.5pt base credit even without patents",
    "Education", "Python", "2 pts",
    "Do you have any patents filed/granted or papers published? Can you share the title and year?",
    "Tell me about your patent/paper — what was the core contribution and has it been cited?",
    "Any patent OR publication = 2pts. TIER_1 with no patent = 0.5pt base credit.")

row("EDU", "Recruiter\n(Recruiter fills)",
    "LinkedIn / Professional Social Media",
    "1. LinkedIn Profile URL  —  linkedin.com/in/username\n"
    "2. Profile Activity Level  —  Not Present (0pt) / Less Active (0.25pt) / More Active (0.5pt) / Regular Updates (1pt)\n"
    "3. Last Post Recency  —  within last 3 months / 3–12 months / >1 year / none\n"
    "4. Connections Approx  —  <100 / 100–500 / 500+\n"
    "5. GitHub / Kaggle / Stack Overflow Links  —  any additional professional profiles",
    "Education", "Recruiter Checks", "1 pt",
    "Can you share your LinkedIn profile URL? I'll check it during our call.",
    "—",
    "Education scoring Updated.xlsx specifies: Not present=0 | Less active=0.25 | More active=0.5 | Regular=1.0")

row("EDU", "Resume\n+\nRecruiter",
    "Extra-Curricular Activities",
    "1. Activities Listed  —  sports / volunteering / hackathon / club / committee\n"
    "2. Leadership Role  —  Y/N  (coordinator / captain / organiser = stronger signal)\n"
    "3. Tech Relevance  —  Y/N  (coding club / robotics / data competition = more relevant)\n"
    "4. Examples: Placement Coordinator, Event Organiser, NSS/NCC, Sports Captain, Cultural Secretary",
    "Education", "Python / Recruiter", "1 pt",
    "Were you involved in any clubs, societies, sports, or campus activities during college — especially any leadership roles?",
    "—",
    "Boolean: any confirmed activity = 1pt. No activity = 0pt.")

# ═══════════════════════════════════════════════════════════════════════════
sec("SKILLS  —  45 pts total", SKL_H)
# ═══════════════════════════════════════════════════════════════════════════

row("SKL", "Resume\n+\nRecruiter Validates",
    "Skill List  —  Years of Experience  —  Timeline",
    "1. Skill Name  —  canonical name after normalisation (e.g. 'Py' → 'Python')\n"
    "2. Years of Active Professional Use  —  first used to last used\n"
    "3. Weighted Evidence Tenure  —  credible use years (weighted by evidence quality)\n"
    "4. First Used Year  —  when skill first appeared professionally\n"
    "5. Last Used Year  —  when skill was last used\n"
    "6. Currently Active  —  Y/N  (using in current role)\n"
    "7. Evidence Level  —  NONE / MENTION / WEAK / APPLIED / DEEP / EXPERT\n"
    "8. Detection Zone  —  skills section / role description / project description / certification",
    "Skills", "Python", "6 pts",
    "For each key skill: how many years have you been using it professionally? When did you last use it? Are you using it in your current role?",
    "—",
    "Only APPLIED+ skills with confirmed years = 1pt each. Max 6pts. MENTION-only skills do not count.")

row("SKL", "Resume",
    "Mandatory Skills  —  As per JD",
    "1. Skill Name  —  from JD mandatory list\n"
    "2. Found on Resume  —  Y / N\n"
    "3. Evidence Level  —  how strongly evidenced on resume\n"
    "4. Years of Use  —  if found\n"
    "5. Gap Flag  —  skill required by JD but missing or only MENTION-level on resume",
    "Skills", "Python (JD match)", "Flag only — no score",
    "I see [Mandatory Skill] is required for this role. Can you confirm your experience with it and when you last used it?",
    "Walk me through a project where you used [Mandatory Skill] — what specifically did you build?",
    "No pts scored — shown as checkmarks to recruiter. Red flag if critical JD skill is missing entirely.")

row("SKL", "Resume",
    "Good-to-Have Skills  —  As per JD",
    "1. Skill Name  —  from JD nice-to-have list\n"
    "2. Found on Resume  —  Y / N\n"
    "3. Evidence Level  —  if found\n"
    "4. Years of Use  —  if found",
    "Skills", "Python (JD match)", "Flag only — no score",
    "Do you have any experience with [Good-to-Have Skill]?",
    "—",
    "No pts scored — informational flag only.")

row("SKL", "Resume\n+\nRecruiter\n+\nPanel",
    "Certifications  —  Validity + Type + Relevance",
    "1. Certification Name  —  full official name\n"
    "2. Issuing Organisation  —  AWS / Google / Microsoft / Databricks / PMI / CFA Institute etc.\n"
    "3. Year Obtained\n"
    "4. Expiry Year  —  (AWS/Azure/GCP certs expire after 3 years)\n"
    "5. Currently Valid  —  Y/N\n"
    "6. Skill Mapped  —  which skill this certification validates\n"
    "7. Relevance to Target Role  —  HIGH / MEDIUM / LOW / IRRELEVANT\n"
    "8. Certification Type  —  Cloud / Data / ML / Security / Management / Other",
    "Skills", "Python + LLM Judge", "3 pts",
    "Tell me about your certifications — which ones are currently valid and which are most relevant to this role?",
    "—",
    "1 valid relevant cert=1pt. Max 3pts. RED FLAG: >=5 certs AND skill depth <=FOUNDATIONAL = cert farming.")

row("SKL", "Panel\n+\nBERT",
    "Skill Depth",
    "1. Evidence Level per skill  —  NONE / MENTION / WEAK / APPLIED / DEEP / EXPERT\n"
    "   MENTION: listed in skills section, no usage evidence\n"
    "   WEAK: mentioned in role but no project context\n"
    "   APPLIED: used in specific projects with context\n"
    "   DEEP: used in complex/multiple projects, architecture decisions\n"
    "   EXPERT: OSS contribution / patent / publication using this skill\n"
    "2. Depth Label per skill  —  AWARENESS / FOUNDATIONAL / HANDS_ON / ADVANCED / ARCHITECT_LEVEL\n"
    "3. BERT Depth Prediction  —  BERT classifier output for depth\n"
    "4. BERT Confidence Score  —  0.0 to 1.0\n"
    "5. Coding Signal  —  Y/N  (built / developed / implemented / scripted — actually wrote code)\n"
    "6. Architecture Signal  —  Y/N  (designed / architected / system design using this skill)\n"
    "7. Open Source Signal  —  Y/N  (GitHub / OSS contribution / notable repo)\n"
    "8. Project Context  —  Development / Maintenance / Unknown\n"
    "9. Upskill Signal  —  Y/N  (skill growing in recency + repeated across roles)",
    "Skills", "Panel + BERT + LLM Judge", "8 pts",
    "For each key skill: describe a specific project where you used it. What exactly did you build?",
    "Walk me through the most complex use of [Skill] — what design decisions did you make? What trade-offs? What would you do differently?",
    "Top 5 skills scored. avg_blended_score/5 x 8. BERT blend: >=0.65 conf = 65% BERT, <0.45 = evidence only.")

row("SKL", "Panel\n+\nPython",
    "Skill Recency",
    "1. Recency Classification per skill  —  RECENT (<2 yrs) / MID (2–5 yrs) / OLD (>5 yrs) / UNKNOWN\n"
    "2. RECENT Skill Count  —  number of skills used in last 2 years\n"
    "3. Total Skill Count  —  total skills on resume\n"
    "4. RECENT Percentage  —  RECENT_count / total_count x 100\n"
    "5. Currently Active Skills  —  skills being used in current role right now",
    "Skills", "Panel + Python", "6 pts",
    "When did you last actively use [Skill]? Is it something you use currently in your day-to-day work?",
    "In your current role, which of your listed skills are you using regularly?",
    "(RECENT_count / total_skills) x 6. MENTION-only skills = OLD by default.")

row("SKL", "Resume\n+\nRecruiter",
    "Skills Learning Acumen",
    "1. New Skills Per Year  —  list of new skills picked up each year  {2021: [Kafka, dbt], 2022: [LangChain]}\n"
    "2. Fast Learner Flag  —  Y/N  (>=2 new skills/year for >=2 consecutive years)\n"
    "3. Years with New Skill Uptake  —  how many years had at least 1 new skill\n"
    "4. Learning Rate  —  skills per year\n"
    "5. Recent Learning  —  have they picked up skills in the last 12 months",
    "Skills", "LLM Judge", "3 pts",
    "In the last 2–3 years, what new technologies have you actually applied in your work — not just done a tutorial on?",
    "How do you stay current? Give me an example of something you learned in the last 6 months and used in a real project.",
    "Fast learner (>=2 new/yr for >=2 yrs)=3pts | new skills >=3 yrs=2pts | 1–2 yrs=1pt | none=0pts")

row("SKL", "Resume",
    "Coding Platforms / Community Contributions",
    "1. GitHub  —  Y/N  + URL  (has public repositories / contributions)\n"
    "2. Stack Overflow  —  Y/N  (answers questions, shows depth + willingness to help)\n"
    "3. LeetCode  —  Y/N  (solves algorithm problems)\n"
    "4. HackerRank  —  Y/N\n"
    "5. Kaggle  —  Y/N  (data science competitions — very relevant for ML/DS roles)\n"
    "6. Personal Tech Blog / Articles  —  Y/N  (dev.to, Medium, personal blog)\n"
    "7. Platform Count  —  total platforms  (>=3=3pts | 2=2pts | 1=1pt | 0=0pts)",
    "Education", "Python + LLM Judge", "3 pts",
    "Do you have a GitHub profile? Can you share the URL? Are you active on Kaggle, LeetCode, or any other coding platforms?",
    "Walk me through your most significant GitHub repository — what problem does it solve and how many people use it?",
    "")

row("SKL", "Recruiter\n+\nPanel",
    "Communication and Presentation Skills",
    "1. Verbal Clarity  —  can express technical concepts clearly  (1–5)\n"
    "2. Logical Structure  —  organises answers with beginning-middle-end  (1–5)\n"
    "3. Confidence Level  —  not nervous, not overconfident  (1–5)\n"
    "4. Audience Adaptability  —  adjusts jargon and depth to audience  (1–5)\n"
    "5. English Proficiency  —  fluent / functional / limited\n"
    "6. Presentation Skills  —  can walk through slides / demos clearly",
    "Skills", "Panel Feedback + LLM Judge", "5 pts",
    "Pre-screen impression from recruiter call (proxy only)",
    "Explain [technical concept from their work] as if I'm a non-technical business executive. Then explain it differently to a junior engineer.",
    "5pts — Panel scored. 1=Very Poor | 2=Poor | 3=Acceptable | 4=Good | 5=Excellent")

row("SKL", "Panel",
    "Domain Skills",
    "1. Domain Area Tested  —  BFSI / ECOMMERCE / HEALTHCARE / LOGISTICS etc.\n"
    "2. Business Context Understanding  —  knows why the problem exists, not just how to solve it technically\n"
    "3. Domain-Specific Concepts  —  credit risk / supply chain planning / patient cohort analysis etc.\n"
    "4. Regulatory / Compliance Awareness  —  knows relevant regulations for the domain\n"
    "5. Industry Metrics  —  knows the KPIs and success metrics used in the domain\n"
    "6. Scenario Question Score  —  0 to 5  (panel assigns based on depth shown in domain scenario)",
    "Skills", "Panel Feedback + LLM Judge", "5 pts",
    "N/A",
    "BFSI: Walk me through how you'd build a credit risk model — data, approach, regulatory constraints.\nHEALTHCARE: How would you approach patient readmission prediction — features, validation, deployment?\n(Tailor to candidate's claimed domain)",
    "5pts — Panel scored via scenario questions in the candidate's own domain.")

row("SKL", "Recruiter\n+\nPanel",
    "Project Explanation Skills",
    "1. Problem Statement Clarity  —  can articulate WHY the project existed (1–3)\n"
    "2. Solution Design Quality  —  explains the technical approach clearly (1–3)\n"
    "3. Role Played Clarity  —  distinguishes what THEY did vs the team (1–3)\n"
    "4. Outcome Articulation  —  states the result, ideally with a number (1–3)\n"
    "5. Trade-off Awareness  —  Y/N  (considered alternatives, explains why this approach was chosen)\n"
    "6. Walk-through Score  —  0=can't explain | 1=disjointed | 2=good structure | 3=clear P→Design→Outcome",
    "Skills", "Panel Feedback + LLM Judge", "3 pts",
    "Walk me through your most recent project — what was the problem, your approach, your role, and the outcome?",
    "Walk me through the technical architecture of [Project]. Which parts did you personally design? What alternatives did you consider?",
    "3pts: 0=cannot explain | 1=disjointed | 2=good structure | 3=clear problem-to-outcome narrative")

row("SKL", "Panel",
    "Coding Skills",
    "1. Language Tested  —  Python / SQL / Java etc.\n"
    "2. Problem Level  —  EASY / MEDIUM / HARD\n"
    "3. Approach Quality  —  brute force vs optimal  (1–5)\n"
    "4. Clean Code  —  Y/N  (readable variable names, modular functions)\n"
    "5. Edge Cases Handled  —  Y/N  (null inputs, empty arrays, overflow)\n"
    "6. Time/Space Complexity Awareness  —  Y/N  (knows Big-O, discusses trade-offs)\n"
    "7. Testing Awareness  —  Y/N  (considers how to test the solution)\n"
    "8. Panel Narrative  —  free text assessment of coding ability",
    "Skills", "Panel Feedback + LLM Judge", "No direct score\n(Qualitative narrative)",
    "N/A",
    "Write a [role-appropriate] function/query. Walk me through your approach. Then: What is the time complexity? What edge cases should we handle?",
    "Qualitative narrative only — no numeric score. Feeds into holistic panel recommendation.")

row("SKL", "Panel",
    "Conceptual Skills",
    "1. Concept Area Tested  —  e.g. Transformer architecture / Spark internals / CAP theorem\n"
    "2. Depth of Understanding  —  surface vs deep  (1–5)\n"
    "3. Theory vs Practice Balance  —  can apply concepts, not just recite definitions  (1–5)\n"
    "4. First-Principles Thinking  —  Y/N  (derives understanding from basics, not just memorised)\n"
    "5. Misconceptions Detected  —  any factual errors in their understanding\n"
    "6. Panel Narrative  —  free text",
    "Skills", "Panel Feedback + LLM Judge", "No direct score\n(Qualitative narrative)",
    "N/A",
    "Explain how [relevant concept] works from first principles — don't look it up, just walk me through your understanding.",
    "Qualitative narrative only — no numeric score.")

row("SKL", "Panel",
    "Problem Solving Skills",
    "1. Problem Structuring  —  does the candidate clarify requirements before solving?  (1–5)\n"
    "2. Approach Taken  —  structured decomposition vs random attempts  (1–5)\n"
    "3. Creativity  —  Y/N  (considered a non-obvious approach)\n"
    "4. Time to Solution  —  minutes taken\n"
    "5. Hints Needed  —  count of hints required  (0 hints=best | 3+=struggling)\n"
    "6. Problem Solving Score  —  0 to 3  (panel assigns)",
    "Skills", "Panel Feedback + LLM Judge", "3 pts",
    "N/A",
    "Here's a scenario: [Open-ended domain-relevant problem]. Walk me through how you'd approach this. Take your time.",
    "3pts — Panel scored. Fewer hints = higher score. Structure and creativity weighted.")

# ═══════════════════════════════════════════════════════════════════════════
sec("TOTALS", TOT_H)
# ═══════════════════════════════════════════════════════════════════════════
row("TOT","","Experience Subtotal","","Experience","","40 pts","~29 auto","~40 max","Best tier drives E1. Projects drive bulk of auto-score.")
row("TOT","","Education Subtotal","","Education","","15 pts","~13 auto","~15 max","TIER_1 + EXCELLENT GPA = 5pts ceiling on ED1.")
row("TOT","","Skills Subtotal","","Skills","","45 pts","~23 auto","~45 max","Panel fills 16pts. Recruiter validates 6pts.")
row("TOT","","GRAND TOTAL","","All","","100 pts","Stage R: normalise /76","Stage B: /100","3-stage model. Resume→Recruiter→Panel.")

# ─────────────────────────────────────────────────────────────────────────
# WRITE TO SHEET
# ─────────────────────────────────────────────────────────────────────────

# Row 1-2: Title
merge(1, 1, 11, "RESUME INTELLIGENCE — SCORING SHEET WITH EXPANDED DATA POINTS", HDR_BG, sz=13)
merge(2, 1, 11,
      "Every parameter shows its complete list of sub-data-points.  "
      "If a data point is blank on the resume — Recruiter asks.  If still blank — Panel asks.",
      "2E75B6", sz=9)

# Row 3: Stage legend
merge(3, 1, 11,
      "STAGE:  Resume = auto-extracted by system  |  "
      "Recruiter = recruiter fills / validates during phone screen  |  "
      "Panel = panel fills during technical interview",
      "37474F", fg="ECEFF1", sz=9)

# Row 4: Column headers
HDR_COLS = [
    "Stage\n(Who Fills)",
    "Parameter\n(with Max Pts)",
    "SUB DATA POINTS\n(what we look for / capture)",
    "Module",
    "Method",
    "Max\nPts",
    "Benchmark\nScore",
    "Score\n(Resume)",
    "Score\n(Recruiter)",
    "Score\n(Panel)",
    "Notes / Rules"
]
for ci, h in enumerate(HDR_COLS, 1):
    cell(4, ci, h, HDR_BG, bold=True, color="FFFFFF", size=8, ha="center", wrap=True, border=BM)
ws.row_dimensions[4].height = 32

SEC_CFG = {
    "EXP": (EXP_H, EXP_R, EXP_A),
    "EDU": (EDU_H, EDU_R, EDU_A),
    "SKL": (SKL_H, SKL_R, SKL_A),
    "TOT": (TOT_H, TOT_R, TOT_R),
}

row_n = 5
alt = False
for record in ROWS:
    if record[0] == "SEC":
        _, label, bg = record
        merge(row_n, 1, 11, f"  {label}", bg, sz=10)
        ws.row_dimensions[row_n].height = 16
        row_n += 1
        alt = False
        continue

    (_, section, stage, param, sub_dps, module, method, pts,
     rec_in, panel_in, notes) = record

    cfg = SEC_CFG.get(section, (TOT_H, TOT_R, TOT_R))
    side_bg = cfg[0]
    row_bg  = cfg[2] if alt else cfg[1]
    alt = not alt

    # row height — proportional to sub_dps length
    lines = sub_dps.count("\n") + 1
    ws.row_dimensions[row_n].height = max(60, lines * 13)

    cell(row_n, 1, stage, row_bg, bold=False, size=8, ha="center")
    cell(row_n, 2, param, row_bg, bold=True, size=9, ha="left")
    cell(row_n, 3, sub_dps, "FFFFFF", bold=False, size=8, ha="left")
    cell(row_n, 4, module, row_bg, size=8, ha="center")
    cell(row_n, 5, method, row_bg, size=8, ha="center")
    cell(row_n, 6, pts, row_bg, bold=True, size=9, ha="center")

    # benchmark — blank input cell
    bm = ws.cell(row=row_n, column=7, value="")
    bm.fill = fill("FFFDE7"); bm.font = fnt(size=9)
    bm.alignment = aln("center", False); bm.border = B

    # score resume
    sr = ws.cell(row=row_n, column=8, value="")
    sr.fill = fill(RSM_BG); sr.font = fnt(size=9)
    sr.alignment = aln("center", False); sr.border = BM

    # score recruiter
    srec = ws.cell(row=row_n, column=9, value=rec_in if section == "TOT" else "")
    srec.fill = fill(REC_BG); srec.font = fnt(size=8, italic=(section!="TOT"))
    srec.alignment = aln("left" if section=="TOT" else "center", True); srec.border = BM

    # score panel
    spnl = ws.cell(row=row_n, column=10, value=panel_in if section == "TOT" else "")
    spnl.fill = fill(PNL_BG); spnl.font = fnt(size=8, italic=(section!="TOT"))
    spnl.alignment = aln("left" if section=="TOT" else "center", True); spnl.border = BM

    cell(row_n, 11, notes, "FAFAFA", italic=True, size=7, ha="left")

    row_n += 1

# ── column widths ─────────────────────────────────────────────────────────
widths([12, 26, 60, 11, 20, 7, 9, 9, 9, 9, 36])

for i in range(1, 5):
    ws.row_dimensions[i].height = 20

out = "E:/Dev/resume_intelligence/Scoring_Expanded_DataPoints.xlsx"
wb.save(out)
print(f"Saved: {out}")
