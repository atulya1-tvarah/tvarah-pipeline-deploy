"""
Resume Intelligence — Granular Data Dictionary Excel
Every atomic field for every parameter with:
- Plain English definition
- Every possible value
- Auto-detection method + reliability
- Score impact
- EXACT recruiter question if missing
- EXACT panel question if still missing
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Colours ──────────────────────────────────────────────────────────────
C = {
    "exp_hdr":   "1E4620", "exp_row":   "E8F5E9", "exp_alt":   "C8E6C9",
    "skill_hdr": "4A148C", "skill_row": "F3E5F5", "skill_alt": "E1BEE7",
    "edu_hdr":   "7B3B00", "edu_row":   "FFF3E0", "edu_alt":   "FFE0B2",
    "rec_hdr":   "0D47A1", "rec_row":   "E3F2FD", "rec_alt":   "BBDEFB",
    "panel_hdr": "880E4F", "panel_row": "FCE4EC", "panel_alt": "F8BBD0",
    "meta_hdr":  "263238", "meta_row":  "ECEFF1", "meta_alt":  "CFD8DC",
    "crit":      "FFCDD2", "imp":       "FFF9C4", "nice":      "F1F8E9",
    "auto_hi":   "C8E6C9", "auto_med":  "FFF9C4", "auto_low":  "FFCCBC",
    "white":     "FFFFFF", "gray":      "F5F5F5",
    "score_hi":  "1B5E20", "score_med": "E65100", "score_low": "B71C1C",
}

def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False, color="000000", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)
def align(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

thin = Side(style="thin", color="BDBDBD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell(ws, r, c, val, bg="FFFFFF", bold=False, color="000000",
         size=9, halign="left", wrap=True, italic=False):
    cl = ws.cell(row=r, column=c, value=val)
    cl.fill = fill(bg)
    cl.font = font(bold=bold, color=color, size=size, italic=italic)
    cl.alignment = align(halign, wrap)
    cl.border = BORDER
    return cl

def merge(ws, r, c1, c2, val, bg, fg="FFFFFF", size=11, bold=True):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cl = ws.cell(row=r, column=c1, value=val)
    cl.fill = fill(bg)
    cl.font = font(bold=bold, color=fg, size=size)
    cl.alignment = align("center", False)
    cl.border = Border(
        left=Side(style="medium"), right=Side(style="medium"),
        top=Side(style="medium"), bottom=Side(style="medium"))

def widths(ws, cols):
    for i, w in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

COLS = [
    "#", "Parameter", "Section",
    "Atomic Field Name",
    "Plain English — What This Means",
    "All Possible Values (complete list)",
    "Auto-Detected From Resume?",
    "Detection Reliability",
    "Score Impact",
    "Priority",
    "RECRUITER: Question to Ask When Missing",
    "PANEL: Question to Ask When Still Missing",
    "Notes / Edge Cases"
]
COL_W = [4, 20, 10, 22, 36, 38, 28, 10, 18, 10, 40, 40, 32]

# ─────────────────────────────────────────────────────────────────────────
# ALL DATA — each tuple = one atomic field row
# Columns: (param_id, param_name, section, field, definition,
#            values, auto_detection, reliability, score_impact,
#            priority, recruiter_q, panel_q, notes)
# ─────────────────────────────────────────────────────────────────────────
ALL_ROWS = []

def add_section(label):
    ALL_ROWS.append(("SECTION", label, "", "", "", "", "", "", "", "", "", "", ""))

def add(param_id, param_name, section, field, defn, values, auto, rel,
        score, prio, rec_q, panel_q, notes=""):
    ALL_ROWS.append((param_id, param_name, section, field, defn, values,
                     auto, rel, score, prio, rec_q, panel_q, notes))

# ══════════════════════════════════════════════════════════════════════════
# E1 — COMPANIES WORKED WITH (5 pts)
# ══════════════════════════════════════════════════════════════════════════
add_section("E1 — COMPANIES WORKED WITH  |  5 pts  |  Method: Python + LLM Judge")

add("E1","Companies Worked With","Experience",
    "Company Name (Raw)",
    "The company name exactly as written on the candidate's resume",
    "Any text — e.g. 'Google India Pvt Ltd', 'Accenture', 'TCS'",
    "Python NLP extracts from experience section headers", "HIGH",
    "Feeds tier lookup → E1 score (1–5 pts)",
    "CRITICAL",
    "Please confirm the full legal/official name of the company you worked at — e.g. was it 'Amazon' or 'Amazon Web Services India'?",
    "Can you confirm which exact entity/division of [Company] you were employed by?",
    "Many candidates write abbreviations. Full name needed for correct tier match.")

add("E1","Companies Worked With","Experience",
    "Canonical Company Name",
    "Normalised, deduplicated company name after fuzzy matching against our database of 500+ companies",
    "From our company database — e.g. 'Google' → 'Alphabet / Google'",
    "Python fuzzy match (SequenceMatcher ≥0.75 ratio) + alias lookup", "MEDIUM",
    "Determines which tier is looked up",
    "CRITICAL",
    "Is [Company Name] the same company as [Canonical Name]? For example, is 'Goog India' the same as Google?",
    "—",
    "If no match found at ≥0.75 threshold, falls back to LLM classification")

add("E1","Companies Worked With","Experience",
    "Company Tier",
    "Quality rating of the employer on a 1–5 scale. This is the single biggest driver of the E1 score. "
    "Tier 1 = globally recognised product/tech giants. Tier 5 = completely unknown company.",
    "1 = FAANG / Hyper-scale (Google, Amazon, Meta, Microsoft, Apple, Netflix, OpenAI, Stripe, etc.)\n"
    "2 = Unicorn / Top Regional (Flipkart, Razorpay, CRED, Freshworks, Zerodha, Swiggy, Grab, etc.)\n"
    "3 = Mid-size funded / MNC subsidiary (Thoughtworks, Persistent, JP Morgan India, Samsung R&D India, etc.)\n"
    "4 = IT Services / Large SME (TCS, Infosys, Wipro, HCL, Accenture, Capgemini, Cognizant, etc.)\n"
    "5 = Unknown / Very small (not in our database)",
    "Python lookup against 500+ company DB + fuzzy match. LLM fallback for unknowns when enabled.", "MEDIUM",
    "Tier 1 → 5 pts | Tier 2 → 4 pts | Tier 3 → 3 pts | Tier 4 → 2 pts | Tier 5 → 1 pt",
    "CRITICAL",
    "I see [Company] on your resume — is it a well-known product company, an IT services company, or a startup? "
    "Is it listed on a stock exchange or funded by VCs? Roughly how many employees does it have?",
    "Can you describe what [Company] does and its scale — revenue, users, team size?",
    "Best tier across ALL companies in the candidate's career is used, not average. "
    "One FAANG stint = Tier 1 even if other roles are Tier 4.")

add("E1","Companies Worked With","Experience",
    "Signal Strength",
    "Derived quality indicator for the employer — HIGH means strong evidence of good work environment and talent bar",
    "HIGH (Tier 1 or Tier 2 well-known company)\n"
    "MEDIUM (Tier 3 established regional/MNC)\n"
    "LOW (Tier 4 services/outsourcing)\n"
    "UNKNOWN (Tier 5 or unrecognised)",
    "Python: derived from tier + company_type combo", "HIGH",
    "Qualitative — informs recruiter narrative but does not directly add pts",
    "IMPORTANT",
    "—",
    "—",
    "Shown to recruiter as context, not a scored field")

add("E1","Companies Worked With","Experience",
    "Primary Industry / Domain",
    "The main industry sector the company operates in. This tells us what DOMAIN the candidate has experience in — "
    "a BFSI company means the candidate has financial services context; EDTECH means education product context.",
    "ECOMMERCE | BFSI | FINTECH | PAYMENTS | HEALTHCARE | PHARMA | EDTECH | SAAS | MEDIA | GAMING | "
    "TELECOM | MANUFACTURING | LOGISTICS | SUPPLY_CHAIN | REAL_ESTATE | ENERGY | AEROSPACE | IOT | "
    "RETAIL | IT_SERVICES | CONSULTING | GOVERNMENT | DEFENCE | AUTOMOTIVE | AGRITECH | PROPTECH | LEGALTECH",
    "Python keyword match on company name + description. DB lookup for known companies.", "MEDIUM",
    "No direct score — feeds domain_tags on experience, which feeds project domain scoring",
    "IMPORTANT",
    "What industry does [Company] operate in? For example: banking, insurance, e-commerce, healthcare, SaaS?",
    "Which specific domain were most of your projects in at [Company] — e.g. credit risk, supply chain planning, ad-tech?",
    "One company can have multiple domains — take the primary one the candidate worked in")

add("E1","Companies Worked With","Experience",
    "Sub-Domain",
    "More specific sector within the primary domain — e.g. within BFSI it matters whether it was retail banking vs capital markets",
    "BFSI sub: RETAIL_BANKING | INVESTMENT_BANKING | CAPITAL_MARKETS | INSURANCE | PAYMENTS | LENDING | BROKERAGE | WEALTH_MANAGEMENT\n"
    "HEALTHCARE sub: HOSPITAL_IT | PHARMA | MED_TECH | DIAGNOSTICS | HEALTHTECH\n"
    "ECOMMERCE sub: MARKETPLACE | D2C | GROCERY | FASHION | B2B_COMMERCE\n"
    "SAAS sub: CRM | HRM | ERP | ANALYTICS | DEVTOOLS | MARTECH\n"
    "FINTECH sub: NEOBANK | INSURANCE_TECH | WEALTHTECH | REGTECH | LENDING_TECH",
    "LLM infers from company description + role context", "LOW",
    "No direct score — enriches domain context for JD matching",
    "NICE_TO_HAVE",
    "Within [Primary Domain] — what specific area was [Company] focused on? e.g. retail loans or investment banking?",
    "What were the main business problems your team solved — was it risk modelling, fraud detection, customer analytics?",
    "Used for JD matching precision, not for base resume scoring")

add("E1","Companies Worked With","Experience",
    "Company Type",
    "What business model does the company follow — does it build its own product, or deliver services to clients?",
    "PRODUCT (builds its own product/platform for end users)\n"
    "SERVICES (delivers IT services/solutions to client companies)\n"
    "CONSULTING (advisory + delivery to client businesses)\n"
    "STARTUP (early-stage, <Series B, <200 people)\n"
    "MNC (multinational, likely captive/shared services centre in India)\n"
    "RESEARCH (academic, govt lab, R&D institute)\n"
    "HYBRID (does both product and services)",
    "Python DB lookup. Heuristic: TCS/Infosys/Wipro = SERVICES, Flipkart/Zomato = PRODUCT", "HIGH",
    "Feeds DNA score — PRODUCT companies → PRODUCT DNA, SERVICES companies → CONSULTING DNA",
    "CRITICAL",
    "Was [Company] a product company that sells its own software/platform, "
    "or did it deliver projects/services to client companies?",
    "Was your work building [Company]'s own product, or delivering solutions for external clients?",
    "Key distinction: a Data Scientist at a product company has very different work than one at a services firm")

add("E1","Companies Worked With","Experience",
    "Business Model (B2B / B2C)",
    "Who does the company ultimately sell to — businesses or individual consumers?",
    "B2B (sells to other businesses — e.g. Salesforce, Freshworks)\n"
    "B2C (sells to individual consumers — e.g. Zomato, Swiggy, Netflix)\n"
    "B2B2C (sells via businesses to consumers — e.g. payment gateways)\n"
    "INTERNAL (captive/shared services — serves internal business units only)\n"
    "GOVT (sells to government)",
    "LLM infers from company context", "LOW",
    "Enriches DNA and domain context",
    "NICE_TO_HAVE",
    "Does [Company] sell to businesses (B2B) or to end consumers (B2C)?",
    "—",
    "")

add("E1","Companies Worked With","Experience",
    "Funding Stage",
    "Stage of external funding — indicates company maturity and risk profile. A candidate from a Series D unicorn "
    "has a very different risk tolerance and pace than one from a BOOTSTRAPPED or LISTED company.",
    "BOOTSTRAPPED (no external funding)\n"
    "SEED (first external funding, typically <$5M)\n"
    "SERIES_A ($5M–$15M, proving product-market fit)\n"
    "SERIES_B ($15M–$50M, scaling)\n"
    "SERIES_C ($50M–$150M, accelerating growth)\n"
    "SERIES_D_PLUS ($150M+, late-stage)\n"
    "PRE_IPO (preparing for public listing)\n"
    "LISTED (publicly traded company)\n"
    "GOVT_FUNDED (government or grant-funded)\n"
    "N/A (for services/consulting firms)",
    "Python DB lookup for known companies. Unknown → recruiter fills.", "MEDIUM",
    "Proxy for company quality + risk profile. Informs archetype detection (A9 Founder, A10 Contractor)",
    "IMPORTANT",
    "Is [Company] a funded startup? If yes — what stage? e.g. Series A, Series C, listed on a stock exchange?",
    "—",
    "Used as a quality proxy for unknown companies — Series C+ funded = Tier 2 equivalent signal")

add("E1","Companies Worked With","Experience",
    "Headcount / Company Size",
    "Approximate number of employees at the time the candidate worked there. "
    "Size indicates the scale of problems the candidate was exposed to.",
    "<50 (very early stage startup)\n"
    "50–200 (small startup or scale-up)\n"
    "200–1,000 (growth stage)\n"
    "1,000–5,000 (mid-size)\n"
    "5,000–50,000 (large enterprise)\n"
    ">50,000 (mega-corp)",
    "Python DB lookup", "MEDIUM",
    "No direct score — feeds scale signal. Large company → higher complexity expectation",
    "IMPORTANT",
    "When you were at [Company], roughly how many employees did it have? And how big was your immediate team?",
    "How large was the engineering/data organisation you were part of at [Company]?",
    "Team size within the company is more important than total headcount for judging individual impact")

add("E1","Companies Worked With","Experience",
    "Work Type",
    "The nature of work done at this company — were they building new products, maintaining existing ones, "
    "doing client delivery, or providing outsourced support?",
    "PRODUCT_BUILD (greenfield development of new products/features)\n"
    "CONSULTING_DELIVERY (delivering solutions to client companies)\n"
    "OUTSOURCING (managing/operating existing systems for clients)\n"
    "CAPTIVE_CENTER (in-house shared services for parent MNC)\n"
    "RESEARCH (academic or applied R&D)\n"
    "HYBRID",
    "Python heuristic: services companies = OUTSOURCING/CONSULTING, product companies = PRODUCT_BUILD", "MEDIUM",
    "Informs project type scoring and DNA",
    "IMPORTANT",
    "At [Company], was most of your work building new products from scratch, maintaining existing systems, "
    "or delivering solutions for external clients?",
    "What was the split between greenfield development vs maintenance/BAU at [Company]?",
    "A candidate doing PRODUCT_BUILD gets higher project complexity signals")

add("E1","Companies Worked With","Experience",
    "Investor Signal / Investor Quality",
    "Quality of investors backing the company — top-tier VCs signal that the company went through rigorous "
    "due diligence and has a high talent bar",
    "TIER_1_VC (Sequoia, a16z, Tiger Global, Accel, SoftBank, Lightspeed)\n"
    "TIER_2_VC (Blume, Nexus, Matrix, Kalaari, etc.)\n"
    "STRATEGIC_INVESTOR (e.g. Google Ventures, Microsoft M12)\n"
    "CROWDFUNDED\n"
    "UNKNOWN_VC\n"
    "NONE\n"
    "N/A",
    "Python DB lookup for known funded companies", "LOW",
    "Proxy for unknown company quality — VC-backed = partial Tier 2 signal",
    "NICE_TO_HAVE",
    "Do you know who the major investors in [Company] are? e.g. Sequoia, SoftBank?",
    "—",
    "Only relevant for unknown/small companies. Not needed for known Tier 1–3 companies.")

add("E1","Companies Worked With","Experience",
    "Operating Model",
    "How is work organised at this company — are they consulting-style (project by project for clients), "
    "product-style (continuous product iteration), or platform/infra (keeping systems running)?",
    "CONSULTING (client-facing delivery, multiple clients, project-based billing)\n"
    "PRODUCT (own product roadmap, user-facing features, iterative releases)\n"
    "PLATFORM_INFRA (internal platform, SRE, DevOps, infra reliability)\n"
    "DOMAIN_SPECIALIST (deep niche — quant finance, healthcare AI, etc.)",
    "Python DNA engine: title zone (3x) + skills (1.5x) + description (1x) keyword scoring", "MEDIUM",
    "Feeds DNA classification — CONSULTING/PRODUCT/PLATFORM_INFRA/DOMAIN_SPECIALIST",
    "IMPORTANT",
    "Was [Company] more of a consulting shop (client projects) or a product company (own roadmap)?",
    "Describe your team's work model — were you working on a client's problem or [Company]'s own product?",
    "")

# ══════════════════════════════════════════════════════════════════════════
# ROLE LEVEL — still under E1 context
# ══════════════════════════════════════════════════════════════════════════
add("E1 / E3","Role at Company (per tenure)","Experience",
    "Job Title (Raw)",
    "The exact job title the candidate held — as written on the resume",
    "Free text — e.g. 'Senior Data Scientist', 'Staff ML Engineer', 'Associate Vice President'",
    "Python NLP extracts from experience section", "HIGH",
    "Feeds seniority level → career_progression (E3)",
    "CRITICAL",
    "What was your exact job title at [Company]? Was it a designation given by the company or your working title?",
    "How did your role evolve from when you joined [Company] to when you left?",
    "Some titles are inflated — panel should probe scope even if title sounds senior")

add("E1 / E3","Role at Company (per tenure)","Experience",
    "Seniority Level",
    "Numeric mapping of the title to a seniority ladder for career progression tracking. "
    "This is how we determine if someone has been growing or stagnating.",
    "1 = Individual Contributor (Analyst, Associate, Junior, Entry-level)\n"
    "2 = Senior Individual Contributor (Senior, Senior II, Staff, Specialist)\n"
    "3 = Lead / Principal (Lead, Principal, Tech Lead, Subject Matter Expert)\n"
    "4 = Manager / Head (Manager, Head of, Engineering Manager)\n"
    "5 = Director / VP (Director, VP, Group Head)\n"
    "6 = C-Suite / Partner (CTO, Chief Scientist, Partner, Founder)",
    "Python title keyword heuristic", "MEDIUM",
    "Seniority sequence across roles → career_progression score (E3, 3 pts)",
    "CRITICAL",
    "—",
    "In your role as [Title] at [Company] — how many people reported to you directly? "
    "Did you manage a team or were you an individual contributor?",
    "Titles vary wildly. A 'VP' at a startup of 20 people ≠ a 'VP' at Goldman Sachs. "
    "Panel must probe scope, team size, reporting structure.")

add("E2","Overall & Relevant Experience","Experience",
    "Start Date (Role)",
    "The month and year the candidate started working in this role. Used to calculate duration and detect gaps.",
    "MM/YYYY format — e.g. Jan 2019, March 2021\nSometimes only year is available: 2019",
    "Python date parser from resume text", "HIGH",
    "Feeds total_yoe, relevant_yoe, gap detection, career_progression",
    "CRITICAL",
    "Can you confirm the month and year you started at [Company]? The resume just shows the year.",
    "—",
    "If only year given, assume January start to avoid over-counting tenure")

add("E2","Overall & Relevant Experience","Experience",
    "End Date (Role)",
    "The month and year the candidate left this role. 'Present' means they are currently in this role.",
    "MM/YYYY or 'Present'/'Current'",
    "Python date parser", "HIGH",
    "Feeds duration, gap detection, recency of skills",
    "CRITICAL",
    "Can you confirm when you left [Company]? The resume shows [year] — was that mid-year or end of year?",
    "—",
    "If 'Present' → check against application date to compute current tenure")

add("E2","Overall & Relevant Experience","Experience",
    "Duration at Company (months)",
    "How long the candidate worked at this company in this role — in months. "
    "Short durations (<12 months) trigger stability penalties.",
    "Integer — calculated as (End Date - Start Date)",
    "Python calculation from start/end dates", "HIGH",
    "Feeds stability score (E4). <6m = -0.8 penalty weight. <12m = -0.4 penalty",
    "CRITICAL",
    "Just to confirm — you were at [Company] from [Start] to [End], that's approximately [X] months, correct?",
    "—",
    "Overlapping dates suggest consulting/part-time work — probe for clarity")

add("E2","Overall & Relevant Experience","Experience",
    "Total Years of Experience (YoE)",
    "Sum of all non-overlapping employment durations. This is the headline number used for band scoring.",
    "Float — e.g. 6.5 years",
    "Python: sum of all role durations, deducting overlaps", "HIGH",
    "E2 score. No JD: 10+=3pts | 6–10=2.5 | 4–6=2.0 | 2–4=1.5 | 1–2=1.0 | <1=0.5",
    "CRITICAL",
    "Your total experience comes to [X] years. Is that right? Are there any roles not on the resume?",
    "—",
    "")

add("E2","Overall & Relevant Experience","Experience",
    "Relevant Years of Experience",
    "Years of experience directly relevant to the target role — calculated by filtering only roles "
    "where the candidate's skills, domain, or role family matches the JD requirements.",
    "Float — subset of total YoE\nCan be 0 if candidate is switching domain completely",
    "Python: filters roles by JD domain + skill overlap", "MEDIUM",
    "REJECT GATE: if relevant_yoe < 70% of JD min range → automatic reject flag",
    "CRITICAL",
    "Of your [X] years of experience, how many years were you directly working in [Target Domain/Role]?",
    "Can you walk me through which of your roles are most relevant to [Target JD]?",
    "Relevant YoE can differ significantly from total YoE for career switchers")

add("E2","Overall & Relevant Experience","Experience",
    "Experience Band",
    "Bucket the candidate falls into based on total YoE — used for quick human reference",
    "FRESHER (<1 year)\nJUNIOR (1–2 years)\nMID (2–4 years)\nSENIOR (4–6 years)\nSEASONED (6–10 years)\nVETERAN (10+ years)",
    "Python: calculated from total_yoe", "HIGH",
    "Determines weight split (Archetypes A5=FRESHER, A6=VETERAN)",
    "IMPORTANT",
    "—",
    "—",
    "")

add("E2","Overall & Relevant Experience","Experience",
    "Domain Continuity",
    "Has the candidate stayed within the same broad domain across all roles, or have they switched domains? "
    "Domain continuity = all relevant YoE counts. Domain switch = only post-switch years count for new domain.",
    "CONTINUOUS (same domain throughout — full YoE counts)\n"
    "SWITCHER (changed domain 1–2 times — only relevant years count)\n"
    "SERIAL_SWITCHER (3+ domain changes)",
    "Python: compare domain_tags across all roles", "MEDIUM",
    "Triggers Archetype A8 (Domain Switcher). Relevant YoE recalculated.",
    "CRITICAL",
    "I see you've moved from [Domain A] to [Domain B] — how much of your work in [Company] was relevant to [Target Domain]?",
    "Walk me through how your experience in [Previous Domain] helps you in [Target Domain].",
    "E.g. 6 yrs DE + 2 yrs DS → for DS role, relevant = 2 yrs unless DE skills directly apply")

# ══════════════════════════════════════════════════════════════════════════
# E3 — CAREER PROGRESSION
# ══════════════════════════════════════════════════════════════════════════
add_section("E3 — CAREER PROGRESSION  |  3 pts  |  Method: Python + BERT + LLM Judge")

add("E3","Career Progression","Experience",
    "Title Seniority Sequence",
    "The sequence of seniority levels (1–6) across all roles in chronological order. "
    "An increasing sequence = progression. A flat or decreasing sequence = stagnation or decline.",
    "List of integers — e.g. [1, 2, 2, 3, 4] = growing\n"
    "[3, 3, 3, 3] = stagnant\n"
    "[4, 3, 2] = declining",
    "Python: maps each title to seniority level, sequences chronologically", "MEDIUM",
    "Primary input to career_progression score (0–5 → scaled to 3 pts)",
    "CRITICAL",
    "Looking at your titles — [Title 1] → [Title 2] → [Title 3] — were these promotions, "
    "lateral moves, or downgrades in responsibility?",
    "How did your scope and responsibilities change from [Earlier Role] to [Later Role]?",
    "Title can stay the same but scope can grow — panel must probe responsibilities not just titles")

add("E3","Career Progression","Experience",
    "Promotion (at same company)",
    "Did the candidate receive a promotion — a title change and increase in responsibility — while staying at the same employer?",
    "Y (promotion detected — same employer, different title at higher level)\n"
    "N (no promotion at this employer)",
    "Python: detects same employer + seniority level increase between consecutive roles", "HIGH",
    "+ve signal for career_progression. Exempt from stability penalty (tenure counted per company, not per title)",
    "IMPORTANT",
    "At [Company] you moved from [Title A] to [Title B] — was that a formal promotion with a salary increase, "
    "or just a title change?",
    "—",
    "")

add("E3","Career Progression","Experience",
    "BERT Career Progression Class",
    "What the BERT classifier predicts about this candidate's overall career trajectory based on the full resume text",
    "FAST_TRACK (rapid promotions, top-tier company moves, growing scope every 1–2 yrs)\n"
    "GROWING (consistent upward movement, maybe slower)\n"
    "LATERAL (same level moves across companies, no real progression)\n"
    "DECLINING (moved to lower-seniority roles or smaller companies over time)",
    "BERT model trained on labeled resumes", "MEDIUM",
    "Blended with Python score. ≥0.60 conf → 60% BERT weight. <0.40 → Python heuristic only",
    "CRITICAL",
    "—",
    "How would you describe your career trajectory over the past [X] years? "
    "What was driving each of your job changes?",
    "BERT confidence <0.40 = model uncertain. LLM judge then makes final call.")

add("E3","Career Progression","Experience",
    "Title Velocity",
    "How fast the candidate is getting promoted — measured as seniority levels gained per year. "
    "High velocity = fast tracker.",
    "Float — e.g. 0.3 levels/year = standard | 0.6+ = fast track | 0.0 = no movement",
    "Python: (highest_seniority - lowest_seniority) / total_yoe", "MEDIUM",
    "Informs FAST_TRACK vs LATERAL classification",
    "IMPORTANT",
    "—",
    "—",
    "Velocity of 0 for 5+ years triggers red flag R13 (stagnation)")

add("E3","Career Progression","Experience",
    "Same Company Growth Signal",
    "Did the candidate grow within a single company rather than job-hopping? "
    "Internal growth = strong signal (shows the company valued them enough to promote).",
    "Y (2+ different titles / levels at same employer)\n"
    "N",
    "Python: detects multiple records with same company name + different titles", "HIGH",
    "+0.5 bonus to career_progression when detected",
    "IMPORTANT",
    "You were at [Company] for [X] years — did you receive any promotions or role changes during that time?",
    "—",
    "")

# ══════════════════════════════════════════════════════════════════════════
# E4 — STABILITY
# ══════════════════════════════════════════════════════════════════════════
add_section("E4 — STABILITY  |  3 pts  |  Method: Python + LLM Judge")

add("E4","Stability","Experience",
    "Average Tenure (months)",
    "The average length of time spent at each employer — the single biggest input to the stability score. "
    "Low average tenure = job-hopper signal.",
    "Float — months per company\n"
    "36+ months = stable | 24–36 = acceptable | 18–24 = borderline | 12–18 = concern | <12 = red flag",
    "Python: mean of all role durations", "HIGH",
    "Primary stability score: 36+m→5.0 | 24–36→4.0 | 18–24→3.5 | 12–18→3.0 | 8–12→2.0 | <8→1.5",
    "CRITICAL",
    "Your average tenure across companies is [X] months. Were any of those short stints due to a company shutting down, layoffs, or contract roles?",
    "—",
    "")

add("E4","Stability","Experience",
    "Loyalty Signal",
    "Classification of the candidate's overall loyalty pattern",
    "HIGH (avg tenure >36 months, no job-hopping)\n"
    "MEDIUM (18–36 months average, occasional short stints)\n"
    "LOW (<18 months average or frequent short stints)",
    "Python loyalty_bucket() function", "HIGH",
    "Feeds narrative — not direct score but part of LLM judge input",
    "IMPORTANT",
    "—",
    "—",
    "")

add("E4","Stability","Experience",
    "Short Stints Count",
    "Number of roles where the candidate stayed less than 12 months. Each short stint adds a penalty.",
    "Integer — count of roles <12 months duration",
    "Python: count of durations < 12 months", "HIGH",
    "<6m stints → -0.8 penalty per stint. 6–12m stints → -0.4 penalty per stint",
    "CRITICAL",
    "You were at [Company] for only [X] months — can you explain why you left so quickly? "
    "Was it a layoff, contract role, company closure, or your choice?",
    "—",
    "Context matters: startup shutdown / layoff → no penalty. Voluntary quick exits → full penalty")

add("E4","Stability","Experience",
    "Job-Hopping Rate",
    "Number of companies changed per year. More than 1.5 companies per year = job-hopper signal.",
    "Float — e.g. 0.4 = 1 company per 2.5 years (stable) | 1.5 = 1.5 companies/year (borderline) | 2.0+ = red flag",
    "Python: total_companies / total_yoe", "HIGH",
    ">2 roles/year → -0.5 stability penalty",
    "CRITICAL",
    "Looking at your career, you've changed [X] companies in [Y] years — "
    "were any of these intentional moves for growth, or were some forced (layoffs, company shut down)?",
    "—",
    "Soft flag raised if rate > 1.5. Recruiter note added.")

add("E4","Stability","Experience",
    "Consultant Pattern",
    "Is the candidate actually a consultant/contractor (same skill cluster across short stints) "
    "rather than a serial job-hopper?",
    "Y (Archetype A10: 3+ stints <18m with consistent skill cluster → consultant pattern)\n"
    "N",
    "Python: cross-reference stints + skill_cluster consistency", "MEDIUM",
    "If Y → Archetype A10 applies. Stability algorithm adjusted (short stints not penalised as heavily)",
    "IMPORTANT",
    "Have any of your roles been on a contract / consulting basis rather than permanent employment?",
    "—",
    "Consultant pattern should NOT be penalised the same as genuine job-hopping")

add("E4","Stability","Experience",
    "Startup Exit Context",
    "Was a short tenure due to the startup shutting down or getting acquired? If yes, short tenure should not be penalised.",
    "SHUTDOWN (company ceased operations)\n"
    "ACQUIRED (company was bought — candidate may have left post-acquisition)\n"
    "LAYOFF (mass layoff, not performance-related)\n"
    "VOLUNTARY (candidate chose to leave)\n"
    "UNKNOWN",
    "Python keyword scan for 'shut down', 'acquired', 'laid off', 'redundancy' — LLM assists", "LOW",
    "SHUTDOWN/LAYOFF → stability penalty waived for that stint",
    "CRITICAL",
    "I see you were at [Company] for only [X] months — did the company shut down, were you laid off, "
    "or did you choose to leave?",
    "—",
    "This is the most important context question for short stints")

# ══════════════════════════════════════════════════════════════════════════
# E5 — AWARDS & RECOGNITIONS
# ══════════════════════════════════════════════════════════════════════════
add_section("E5 — AWARDS & RECOGNITIONS  |  3 pts  |  Method: Python + LLM Judge")

add("E5","Awards & Recognitions","Experience",
    "Award / Recognition Type",
    "What kind of recognition was it — company award, industry honour, academic prize, etc.",
    "COMPANY_AWARD (e.g. 'Employee of the Quarter', 'Star Performer')\n"
    "INDUSTRY_AWARD (e.g. Forbes 30 Under 30, Nasscom award)\n"
    "ACADEMIC_PRIZE (e.g. Gold Medal, Institute Rank 1)\n"
    "HACKATHON_WIN (e.g. won internal or external hackathon)\n"
    "PROMOTION (formal promotion = recognition signal)\n"
    "PATENT (filed or granted patent)\n"
    "PUBLICATION (peer-reviewed paper or conference paper)\n"
    "CONFERENCE_TALK (invited speaker at a tech conference)\n"
    "OPEN_SOURCE_RECOGNITION (e.g. GitHub stars, notable OSS contribution)",
    "Python regex: named awards, 'won', 'recipient', 'received', 'promoted' keywords", "MEDIUM",
    "Count: 0→0pts | 1→1pt | 2→2pts | 3+→3pts. LLM validates genuine vs generic",
    "IMPORTANT",
    "I see [Achievement] on your resume — can you tell me more about it? "
    "Was it a competitive award, and what were you recognised for?",
    "Tell me about [Award/Achievement] — what was the selection process? "
    "How many people were considered, and what specifically did you do to earn it?",
    "LLM filters out generic 'Best Team Player' participation certificates. "
    "Only counted if genuinely competitive or rare.")

add("E5","Awards & Recognitions","Experience",
    "Promotion Count",
    "Number of formal promotions across the entire career — promotions are strong recognition signals",
    "Integer — e.g. 3 promotions in 8 years",
    "Python: detects title level increase at same employer", "HIGH",
    "Each promotion = +1 to award_count for scoring",
    "IMPORTANT",
    "Across your career, have you received any formal promotions with a change in title and compensation?",
    "—",
    "")

add("E5","Awards & Recognitions","Experience",
    "LLM Verified (Genuine vs Generic)",
    "Whether the LLM judge confirmed this as a genuine, competitive recognition vs a generic participation trophy",
    "VERIFIED_GENUINE (competitive, rare, specific)\n"
    "GENERIC (participation certificate, attendance award)\n"
    "UNVERIFIABLE (not enough context)",
    "LLM judges each achievement in context", "MEDIUM",
    "Only VERIFIED_GENUINE items count toward the 3-pt score",
    "IMPORTANT",
    "—",
    "You mentioned [Award] — was this given to the top X% of employees, or was it more of a participation recognition?",
    "")

# ══════════════════════════════════════════════════════════════════════════
# E6 — MENTORSHIP / CODE REVIEWS / INTERVIEWS
# ══════════════════════════════════════════════════════════════════════════
add_section("E6 — MENTORSHIP / CODE REVIEWS / INTERVIEWS  |  3 pts  |  Method: Python + BERT + LLM — RECRUITER FILLS")

add("E6","Mentorship Signal","Experience",
    "Mentored Engineers Count",
    "How many junior engineers / team members has the candidate directly mentored?",
    "0 (no mentorship evidence)\n"
    "1–2 (informal mentorship)\n"
    "3–5 (structured mentorship of small group)\n"
    "6+ (team-scale mentorship)",
    "Python keyword: 'mentored', 'guided', 'coached', 'onboarded' + number if mentioned", "LOW",
    "Feeds mentorship_signal BERT + LLM scoring",
    "CRITICAL",
    "Have you mentored junior engineers or team members? If yes — how many, and for how long? "
    "Was it formal (assigned by the company) or informal?",
    "Tell me about a time you mentored someone — what was the outcome for them?",
    "Most candidates mention this briefly. Recruiter must probe for specifics.")

add("E6","Mentorship Signal","Experience",
    "Code Review Signal",
    "Has the candidate been doing code reviews — a standard indicator of technical seniority and mentorship",
    "Y (code review mentioned or implied by title/role)\n"
    "N",
    "Python keyword: 'code review', 'PR review', 'pull request review'", "LOW",
    "Contributes to BERT mentorship prediction",
    "IMPORTANT",
    "In your current/recent role, were you responsible for reviewing code written by other team members?",
    "How did you approach code reviews — what did you look for beyond just correctness?",
    "")

add("E6","Mentorship Signal","Experience",
    "Interview Panel Participation",
    "Has the candidate been part of the hiring panel — conducting technical interviews? "
    "This is a strong seniority and mentorship signal.",
    "Y (conducted interviews for the company)\n"
    "N",
    "Python keyword: 'conducted interviews', 'technical interviewer', 'hiring panel'", "LOW",
    "Contributes to BERT mentorship prediction",
    "IMPORTANT",
    "Have you been involved in interviewing candidates for your team or company?",
    "What kind of interviews did you conduct — coding, system design, behavioural?",
    "")

add("E6","Mentorship Signal","Experience",
    "BERT Mentorship Class",
    "BERT's classification of the overall mentorship evidence level in the resume",
    "LEAD (led engineers in 2+ roles, clear evidence of team leadership)\n"
    "FORMAL (one instance of formal mentorship or team lead)\n"
    "IMPLIED (soft signals — 'worked with junior devs', 'trained new joiners')\n"
    "NONE (no mentorship signals found)",
    "BERT classifier on full resume text", "MEDIUM",
    "BERT ≥0.60 conf → primary. <0.60 → Python heuristic. Score: LEAD=3 | FORMAL=2 | IMPLIED=1 | NONE=0",
    "CRITICAL",
    "—",
    "—",
    "")

add("E6","Mentorship Signal","Experience",
    "Roles with Mentorship",
    "At which companies / roles was the candidate doing mentorship? Helps verify if it was consistent or one-off.",
    "List of companies/roles — e.g. ['Google (2021–2023)', 'Flipkart (2019–2021)']",
    "Python: cross-reference mentorship keywords against role dates", "LOW",
    "Mentorship in 2+ roles = LEAD classification = 3 pts",
    "IMPORTANT",
    "At which companies were you mentoring people? Was it something you did consistently or just at one place?",
    "—",
    "")

# ══════════════════════════════════════════════════════════════════════════
# E7 — INTERNATIONAL EXPOSURE
# ══════════════════════════════════════════════════════════════════════════
add_section("E7 — INTERNATIONAL EXPOSURE  |  2 pts  |  Method: LLM Judge — RECRUITER CONFIRMS")

add("E7","International Exposure","Experience",
    "Onsite / Overseas Work Signal",
    "Did the candidate physically work in another country for a client, project, or relocation? "
    "This is the strongest international exposure signal.",
    "Y (explicit: worked onsite in another country)\n"
    "N",
    "Python keyword: 'onsite', 'on-site', 'London', 'Singapore', 'Dubai', 'US', 'UK', 'Europe', 'relocation'", "MEDIUM",
    "2 pts if explicit onsite/relocation confirmed",
    "CRITICAL",
    "I see [Keyword] on your resume — did you actually travel to/work from [Country] for a project or relocation? "
    "For how long, and in what capacity?",
    "Tell me about your international work experience — what was the context and duration?",
    "Many candidates list global clients but never travel — that's not the same as onsite. Recruiter must distinguish.")

add("E7","International Exposure","Experience",
    "Global Team Signal",
    "Did the candidate work with globally distributed teams — multi-timezone collaboration, global stakeholders?",
    "Y (global team / multi-country collaboration without physical travel)\n"
    "N",
    "Python keyword: 'global team', 'multi-timezone', 'cross-border', 'international stakeholders'", "LOW",
    "1 pt (implied exposure — not as strong as physical onsite)",
    "IMPORTANT",
    "Did you regularly collaborate with teams or stakeholders in other countries? "
    "What timezones were you working across?",
    "—",
    "")

add("E7","International Exposure","Experience",
    "Countries Mentioned",
    "The specific countries where the candidate worked or collaborated — used to assess geographic breadth",
    "List — e.g. ['UK', 'Singapore', 'US']\nNone if only India-based",
    "Python: named entity recognition + country keywords", "MEDIUM",
    "More countries = stronger exposure signal",
    "NICE_TO_HAVE",
    "Which countries did you work in or with?",
    "—",
    "")

add("E7","International Exposure","Experience",
    "Visa / Work Permit Mention",
    "Is there a mention of a work visa — strong evidence of legal right to work abroad, i.e. physical work",
    "Y (visa mentioned — H1B, L1, work permit, etc.)\n"
    "N",
    "Python keyword: 'H1B', 'L1', 'work permit', 'visa', 'PR status'", "LOW",
    "Confirms onsite signal → 2 pts",
    "IMPORTANT",
    "Do you currently hold or have you held a work visa for any country? e.g. H1B, L1, UK visa?",
    "—",
    "")

# ══════════════════════════════════════════════════════════════════════════
# E8 — STAKEHOLDER MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════
add_section("E8 — STAKEHOLDER MANAGEMENT  |  2 pts  |  Method: Python + BERT + LLM — RECRUITER VALIDATES")

add("E8","Stakeholder Management","Experience",
    "Client-Facing Role",
    "Did the candidate work directly with external clients — presenting, gathering requirements, delivering to them?",
    "Y (direct client interaction confirmed)\n"
    "N",
    "Python keyword: 'client', 'customer', 'external stakeholder', 'client-facing'", "MEDIUM",
    "Client-facing → 1.5 pts base",
    "CRITICAL",
    "In your role at [Company], were you directly working with external clients — "
    "like presenting to them, taking requirements from them, or managing their expectations?",
    "Give me an example of a difficult stakeholder situation you managed.",
    "")

add("E8","Stakeholder Management","Experience",
    "C-Suite / Senior Leadership Exposure",
    "Did the candidate present to or work directly with C-level executives (CTO, CEO, CFO, Board)?",
    "Y (explicit C-level interaction)\n"
    "N",
    "Python keyword: 'C-suite', 'CTO', 'CEO', 'Board', 'Executive committee', 'MD', 'VP presentations'", "LOW",
    "C-level → 2 pts (max score)",
    "CRITICAL",
    "Did your work ever involve presenting to or advising C-level leadership or board members? "
    "At which company, and what was the context?",
    "Tell me about a time you presented to senior leadership — what was the audience level and what were you presenting?",
    "")

add("E8","Stakeholder Management","Experience",
    "Cross-Functional Collaboration",
    "Did the candidate work across multiple teams / functions — engineering, product, business, ops — "
    "rather than siloed within a single team?",
    "Y (cross-functional work mentioned)\n"
    "N",
    "Python keyword: 'cross-functional', 'worked with product', 'collaborated with business', 'multiple teams'", "MEDIUM",
    "Internal cross-functional → 1 pt",
    "IMPORTANT",
    "Did your role require you to work closely with other teams like product, business, or operations? "
    "Give an example.",
    "—",
    "")

add("E8","Stakeholder Management","Experience",
    "BERT Stakeholder Class",
    "BERT's classification of the candidate's stakeholder management level",
    "C_LEVEL (presented to or advised C-suite / board)\n"
    "CLIENT_FACING (external client work)\n"
    "INTERNAL (cross-functional but internal only)\n"
    "NONE (pure individual contributor, no stakeholder management evidence)",
    "BERT classifier on full resume text", "MEDIUM",
    "BERT ≥0.60 conf → primary. C_LEVEL=2 | CLIENT_FACING=1.5 | INTERNAL=1 | NONE=0",
    "CRITICAL",
    "—",
    "—",
    "")

# ══════════════════════════════════════════════════════════════════════════
# E9 — CAREER BREAKS
# ══════════════════════════════════════════════════════════════════════════
add_section("E9 — CAREER BREAKS  |  2 pts  |  Method: Pure Python — AUTO DETECT")

add("E9","Career Breaks","Experience",
    "Career Break Count",
    "Number of gaps of more than 3 months between consecutive employment periods. "
    "The total count determines the score and whether a REJECT flag is triggered.",
    "0 (no breaks → 2 pts)\n"
    "1 (one break → 1 pt)\n"
    "2 (two breaks → 0 pts)\n"
    "3+ (three or more breaks → 0 pts + REJECT FLAG)",
    "Python: sorts all roles by date, finds gaps >3 months between end_date and next start_date", "HIGH",
    "0 breaks=2pts | 1=1pt | 2=0pts | 3+=REJECT FLAG",
    "CRITICAL",
    "I can see a gap from [Date] to [Date] on your resume — about [X] months. "
    "Can you tell me what you were doing during that period?",
    "—",
    "REJECT if >2 breaks. Exceptions: MBA during gap (no penalty), maternity ≤18m (soft flag only), COVID 2020 (context note)")

add("E9","Career Breaks","Experience",
    "Break Duration (per break, months)",
    "How long each individual break lasted",
    "Integer — months per gap",
    "Python: end_date_of_role_n to start_date_of_role_(n+1)", "HIGH",
    "Any break >3 months counted. Break >12 months triggers education gap check.",
    "CRITICAL",
    "The gap from [Date A] to [Date B] is [X] months — what were you doing?",
    "—",
    "")

add("E9","Career Breaks","Experience",
    "Break Reason Classification",
    "Why the candidate took the break — context determines whether it's penalised or excused",
    "MBA (pursuing full-time education → EXEMPT if edu record matches the dates)\n"
    "MATERNITY_PATERNITY (parental leave ≤18m → SOFT_FLAG, not penalised)\n"
    "ILLNESS (medical → not penalised with documentation)\n"
    "RECESSION_LAYOFF (2020 COVID wave or mass layoff → context note, not penalised)\n"
    "TRAVEL (sabbatical / travel → borderline, note added)\n"
    "PERSONAL (family reasons → borderline)\n"
    "STARTUP_FAILURE (own startup closed → context note)\n"
    "VOLUNTARY (chose to take a break → mild penalty)\n"
    "UNKNOWN (no explanation given → full penalty)",
    "Python keyword scan + LLM classification. Cross-reference education dates for MBA gap.", "LOW",
    "EXEMPT/SOFT_FLAG reasons → penalty waived. UNKNOWN/VOLUNTARY → full penalty",
    "CRITICAL",
    "Was this break [reason]? e.g. Were you studying, on parental leave, dealing with health issues, or between opportunities?",
    "—",
    "The reason classification directly determines whether the penalty is applied or waived")

add("E9","Career Breaks","Experience",
    "Break Dates (Start and End)",
    "The exact start and end date of each gap period",
    "MM/YYYY to MM/YYYY — e.g. Apr 2020 to Oct 2020",
    "Python: derived from role end and next role start dates", "HIGH",
    "Needed to cross-reference with education dates (MBA check) and economic events (COVID 2020 check)",
    "IMPORTANT",
    "The gap appears to be from approximately [Month Year] to [Month Year]. Does that sound right?",
    "—",
    "")

# ══════════════════════════════════════════════════════════════════════════
# E10/E11 — PROJECTS
# ══════════════════════════════════════════════════════════════════════════
add_section("E10 — PROJECT 1 (Latest)  |  8 pts  |  Method: Python + LLM PROJECT_JUDGE + Panel Feedback")

add("E10","Project 1 — Latest","Experience",
    "Project Title",
    "The name or headline of the project — used to identify what the project was about at a glance",
    "Free text — e.g. 'Real-time Fraud Detection Pipeline', 'Customer Churn Prediction Model'",
    "Python extracts from project section or largest description block under most recent role", "MEDIUM",
    "Criterion (2): Title present → +1 pt",
    "CRITICAL",
    "What is the name or title you'd give to your most significant recent project?",
    "Walk me through your most recent project — what were you building and why?",
    "Many candidates don't list project titles. Recruiter must elicit this.")

add("E10","Project 1 — Latest","Experience",
    "Project Type",
    "What category of technical work the project involved — this tells us the nature of the challenge",
    "DEVELOPMENT (greenfield, building from scratch)\n"
    "MIGRATION (moving from old to new — database, cloud, platform migration)\n"
    "ANALYTICS (data analysis, dashboards, insights)\n"
    "INFRASTRUCTURE (setting up cloud infra, DevOps, pipelines)\n"
    "RESEARCH (experiments, papers, novel algorithms)\n"
    "CONSULTING (advisory, solution design for client)\n"
    "MAINTENANCE (bug fixes, BAU, support)\n"
    "TRANSFORMATION (large-scale business or tech transformation)\n"
    "HYBRID (multiple types combined)",
    "Python keyword heuristics: DEVELOPMENT verbs, MIGRATION keywords, etc. LLM PROJECT_JUDGE confirms.", "MEDIUM",
    "Criterion (1): Type known → +1 pt. LLM confirms type and flags if mismatch with description.",
    "CRITICAL",
    "Was this project building something new from scratch, migrating existing systems, doing data analysis, "
    "or maintaining something that already existed?",
    "What was the core technical challenge of this project — was it a new build, a migration, or optimising something existing?",
    "MAINTENANCE projects score lower on complexity. DEVELOPMENT + MIGRATION score highest.")

add("E10","Project 1 — Latest","Experience",
    "Project Description (length and richness)",
    "How detailed and informative the project description is — a one-liner means we can't score it properly",
    "RICH (>100 words, explains problem + solution + outcome)\n"
    "ADEQUATE (50–100 words, explains the project)\n"
    "MINIMAL (20–50 words, basic description)\n"
    "TOO_SHORT (<20 words, cannot score)",
    "Python: character count + keyword density", "HIGH",
    "Criterion (3): Description >20 chars → +1 pt. Richer descriptions → higher LLM complexity score",
    "CRITICAL",
    "Your project description is quite brief. Can you tell me more about what you built — "
    "the problem you were solving, your approach, and what the outcome was?",
    "Walk me through this project in detail — what was the problem statement, your technical approach, "
    "challenges you faced, and the final outcome?",
    "This is the most important field to enrich. A thin description caps the project score severely.")

add("E10","Project 1 — Latest","Experience",
    "Project Duration (months)",
    "How long the project ran — longer duration suggests larger scope and more sustained ownership",
    "Integer — months\nTypically 3–24 months for a significant project",
    "Python: extracts duration text ('6-month project', 'Jan 2022 – Jun 2022')", "MEDIUM",
    "Criterion (4): Duration ≥3 months → +1 pt",
    "CRITICAL",
    "How long did this project run, approximately?",
    "Was this a one-time project or ongoing? How long were you personally involved?",
    "Short projects (<3 months) suggest POC or experimental work — flag accordingly")

add("E10","Project 1 — Latest","Experience",
    "Skills Used in Project",
    "The specific technologies, frameworks, and tools the candidate used in this project",
    "List of skill names — e.g. ['Python', 'Spark', 'AWS S3', 'XGBoost', 'Kafka']",
    "Python: extracts skill mentions within project description", "MEDIUM",
    "Criterion (5): ≥1 skill listed → +1 pt. Feeds skill_depth scoring for these specific skills.",
    "CRITICAL",
    "What specific technologies and tools did you use in this project?",
    "Beyond the tools listed, were you using any cloud services, databases, or ML frameworks?",
    "")

add("E10","Project 1 — Latest","Experience",
    "Business Domain Tag",
    "What industry / business domain this project was for — "
    "a fraud detection project is in BFSI, a recommendation engine could be ECOMMERCE",
    "ECOMMERCE | BFSI | FINTECH | HEALTHCARE | EDTECH | LOGISTICS | MANUFACTURING | "
    "RETAIL | SAAS | INTERNAL_TOOL | CROSS_DOMAIN | UNKNOWN",
    "Python: keywords in project description matched to domain_hints dictionary", "MEDIUM",
    "Criterion (6): Domain tag present → +1 pt",
    "IMPORTANT",
    "What was the business context of this project — which industry or domain was it for?",
    "Who were the end users or beneficiaries of this project?",
    "")

add("E10","Project 1 — Latest","Experience",
    "Role Played / Ownership Level",
    "What was the candidate's specific role in the project — did they lead it, contribute to part of it, "
    "or just support? This is the most critical signal for differentiating owners from contributors.",
    "ARCHITECT/LEAD (designed the solution end-to-end, owned all decisions)\n"
    "CORE_CONTRIBUTOR (built a significant component, made key technical decisions)\n"
    "CONTRIBUTOR (built one module/feature among many)\n"
    "SUPPORT (debugging, monitoring, BAU support)\n"
    "UNKNOWN",
    "Python: ownership verbs detection — 'built', 'led', 'owned', 'designed', 'architected', "
    "'implemented', 'deployed', 'launched', 'migrated', 'created' + description >50 chars", "MEDIUM",
    "Criterion (7): Ownership verb + description >50 chars → +1 pt",
    "CRITICAL",
    "In this project, what was your specific role? Did you design the architecture, build specific components, "
    "or were you part of a larger team? How many people were on the team?",
    "You say you 'built' this — can you tell me which specific parts you personally designed and implemented? "
    "Were you the sole developer or one of many?",
    "Red flag: 'We did X' without 'I did Y' is not owned work. Panel must probe individual contribution.")

add("E10","Project 1 — Latest","Experience",
    "Quantified Business Impact",
    "Did the candidate state a measurable outcome for this project — a number, percentage, or business metric? "
    "This is what separates high-impact engineers from those who just ship code.",
    "QUANTIFIED (specific: 'reduced latency by 40%', 'saved ₹2Cr/year', 'processed 10M records/day')\n"
    "VERBAL_IMPACT ('significantly improved', 'doubled throughput' — verbal but not numeric)\n"
    "VAGUE ('improved performance', 'made it faster')\n"
    "NONE (no impact mentioned)",
    "Python regex: % sign or numbers + outcome words (reduced/increased/improved/saved/accelerated/deployed) "
    "OR verbal impact patterns (doubled/tripled/halved)", "MEDIUM",
    "Criterion (8): Quantified → +1 pt. Max project score without this = 7/8.",
    "CRITICAL",
    "What was the measurable outcome of this project? For example, did it reduce costs, improve speed, "
    "increase revenue, or improve a key metric — and by how much?",
    "What was the business impact of your work on this project? How did you know it was successful?",
    "The most commonly missing field. Recruiter MUST get a number or metric before ending the call.")

add("E10","Project 1 — Latest","Experience",
    "LLM Complexity Score",
    "How technically complex was this project, rated by the LLM PROJECT_JUDGE on a 0–5 scale. "
    "Most projects score 2–3. Only truly exceptional projects (global scale, novel algorithms, "
    "distributed systems at 10M+ scale) score 4+.",
    "0 (trivial — tutorial-level work)\n"
    "1 (basic — standard CRUD or simple model)\n"
    "2 (moderate — standard engineering with some complexity)\n"
    "3 (good — real-world complexity, multi-system, production-grade)\n"
    "4 (excellent — rare: complex distributed systems, novel approach, significant scale)\n"
    "5 (exceptional — top 5% of all projects seen)",
    "LLM PROJECT_JUDGE analyses full project description", "MEDIUM",
    "Enriches project score. Complexity ≥4 flags candidate for senior/architect roles.",
    "IMPORTANT",
    "—",
    "Tell me about the most technically challenging aspect of this project. "
    "What made it hard, and how did you solve it?",
    "Most internal projects score 2–3. A score of 4+ requires hard evidence of scale or novelty.")

add("E10","Project 1 — Latest","Experience",
    "Candidate Signal (LLM Assessment)",
    "The LLM's overall assessment of the candidate's potential based on this project",
    "EXCELLENT (top-tier work, clear ownership, quantified impact, technical depth)\n"
    "STRONG (good work, most signals present, minor gaps)\n"
    "AVERAGE (adequate, standard work, limited depth evidence)\n"
    "WEAK (thin description, no ownership, no impact, or support-level work)",
    "LLM PROJECT_JUDGE holistic assessment", "MEDIUM",
    "Used in recruiter narrative — not a direct score but informs recommendation",
    "IMPORTANT",
    "—",
    "—",
    "")

add("E10","Project 1 — Latest","Experience",
    "Implied Skills (LLM)",
    "Skills that the LLM infers the candidate must have used, even if not explicitly mentioned — "
    "e.g. a Spark streaming project implies Kafka knowledge even if not listed",
    "List of inferred skills — e.g. ['Kafka', 'HDFS', 'YARN', 'AWS EMR']",
    "LLM PROJECT_JUDGE analyses project type + mentioned skills", "LOW",
    "Enriches skill_depth scoring — implied skills get WEAK evidence level",
    "NICE_TO_HAVE",
    "Did you also work with [Implied Skill] in this project? It seems like it would be needed.",
    "—",
    "")

add("E10","Project 1 — Latest","Experience",
    "Green Flags (LLM)",
    "Positive signals the LLM detected in the project description",
    "2–4 items — e.g. 'Production-scale deployment mentioned', 'Quantified business impact stated', "
    "'Clear ownership with architectural decisions'",
    "LLM PROJECT_JUDGE", "LOW",
    "Shown to recruiter as supporting evidence",
    "NICE_TO_HAVE",
    "—",
    "—",
    "")

add("E10","Project 1 — Latest","Experience",
    "Red Flags (LLM)",
    "Concerns the LLM detected in the project description",
    "1–3 items — e.g. 'Team contribution unclear — no individual ownership stated', "
    "'Impact is vague — no metrics provided', 'Technology stack seems outdated'",
    "LLM PROJECT_JUDGE", "LOW",
    "Shown to recruiter to guide validation questions",
    "CRITICAL",
    "—",
    "[Red flag becomes interview probe question]",
    "Each red flag should generate a recruiter or panel probe question")

add_section("E11 — PROJECT 2 (2nd Latest)  |  6 pts  |  Same fields as Project 1 except no Ownership Verb or Quantified Impact criteria")
add("E11","Project 2 — 2nd Latest","Experience",
    "All fields same as Project 1",
    "Project 2 captures the same atomic fields as Project 1 — Title, Type, Description, Duration, Skills, Domain, "
    "Role Played, LLM Complexity Score — EXCEPT the Ownership Verb criterion and Quantified Impact criterion are NOT scored "
    "(only 6 criteria instead of 8)",
    "Same as Project 1",
    "Same as Project 1", "MEDIUM",
    "Max 6 pts (criteria 1–6 only)",
    "CRITICAL",
    "Tell me about your second most significant project. What were you building, in what domain, for how long, "
    "and what was your specific role?",
    "Walk me through the technical design decisions you made in your second project.",
    "If resume only shows 1 project, recruiter must ask: 'What was the next most significant project you worked on?'")

# ══════════════════════════════════════════════════════════════════════════
# EDUCATION
# ══════════════════════════════════════════════════════════════════════════
add_section("ED1 — INSTITUTE TIER / GPA / STREAM  |  5 pts  |  Method: Pure Python")

add("ED1","Institute Tier","Education",
    "Institution Name (Raw)",
    "The college/university name exactly as written on the resume",
    "Free text",
    "Python NLP extract from education section", "HIGH",
    "Feeds tier lookup → institute_tier score (up to 5 pts)",
    "CRITICAL",
    "Can you confirm the full name of your college/university? The resume shows '[Name]' — is that correct?",
    "—",
    "Many abbreviations used — 'BITS' vs 'BITS Pilani', 'IIT' without the campus name")

add("ED1","Institute Tier","Education",
    "Institute Tier",
    "Quality tier of the educational institution — the most important education signal",
    "TIER_1 (IIT all campuses, IIM all campuses, ISI Kolkata, ISB Hyderabad, IIIT Hyderabad, NLU Delhi, "
    "global top-200 universities: MIT, Stanford, Oxford, Cambridge, CMU, IVY League, etc.)\n"
    "TIER_2 (NIT all campuses, BITS Pilani, Manipal, VIT, NMIMS, XLRI, IIFT, SP Jain, MICA, "
    "well-ranked state universities, top foreign universities outside top-200)\n"
    "TIER_3 (Average state universities, regional private colleges with decent placements)\n"
    "TIER_4 (Below-average institutions, unknown private colleges)\n"
    "UNKNOWN (not in our database — treated as TIER_3 unless context suggests otherwise)",
    "Python lookup with longest-match algorithm against 500+ institution dictionary. "
    "Full name aliases: 'IIT Bombay' AND 'Indian Institute of Technology Bombay' both map to TIER_1.", "HIGH",
    "TIER_1=4 base pts. TIER_2=3 base pts. TIER_3=2 pts. TIER_4=1 pt. "
    "GPA bonus: EXCELLENT → TIER_1 gets +1 (total 5), TIER_2 gets +0.5. Capped at 5.",
    "CRITICAL",
    "I see you attended [Institution] — is that the correct full name? And was it a full-time programme?",
    "—",
    "TIER_1 + GPA EXCELLENT = max 5 pts. TIER_4 + any GPA = 1 pt max.")

add("ED1","Institute Tier","Education",
    "GPA / CGPA / Percentage (Raw)",
    "The academic grade the candidate achieved — as written on the resume",
    "Various formats: '8.5/10', '75%', '3.9/4.0', '9.2 CGPA', 'First Class', 'Distinction'",
    "Python regex extracts numeric value and scale", "MEDIUM",
    "Feeds GPA Band → bonus pts to institute score",
    "IMPORTANT",
    "What was your GPA or percentage in your [Degree]? The resume doesn't mention it.",
    "—",
    "If GPA not mentioned and it's a TIER_1 institution, assume ACCEPTABLE (no bonus applied)")

add("ED1","Institute Tier","Education",
    "GPA Band",
    "Normalised academic performance category — derived from GPA + scale",
    "EXCELLENT (≥8.5/10, ≥3.7/4.0, ≥85%) → adds bonus pts to institute score\n"
    "GOOD (7.5–8.4/10, 3.3–3.6/4.0, 75–84%) → adds smaller bonus\n"
    "ACCEPTABLE (6.5–7.4/10, 2.7–3.2/4.0, 65–74%) → no bonus, no penalty\n"
    "LOW (<6.5/10, <2.7/4.0, <65%) → small penalty",
    "Python benchmarks applied after scale normalisation", "HIGH",
    "EXCELLENT on TIER_1 → +1 pt. EXCELLENT on TIER_2 → +0.5 pt. LOW on any tier → -0.5 pt.",
    "IMPORTANT",
    "What was your grade in this degree — roughly what percentage or CGPA?",
    "—",
    "")

add("ED1","Institute Tier","Education",
    "IT Stream",
    "Is this an IT/technology-aligned degree stream? Relevant to tech role scoring.",
    "YES (CS, CSE, IT, Software Engineering, Data Science, AI/ML, MCA, BCA, Computer Applications, "
    "Information Systems, Statistics, Analytics, Quantitative Economics, Econometrics, "
    "Operations Research, Actuarial Science, MSQE)\n"
    "NO (Non-IT: Mechanical, Civil, Chemical, Commerce, Arts, Management — without tech specialisation)",
    "Python: stream keywords matched against IT_STREAM_PATTERNS list", "MEDIUM",
    "IT stream → TECH fit → education_job_relevance = HIGH → 2 pts",
    "CRITICAL",
    "Your degree is in [Field] — was your curriculum focused on programming, data, or statistics? "
    "Or was it more theoretical?",
    "—",
    "Stream relevance ranking: ECE > EE/Mech > Chemistry > Math > Civil (for CS/data roles)")

add_section("ED2 — DEGREE LEVEL  |  2 pts  |  Method: Pure Python")

add("ED2","Degree Level","Education",
    "Degree Level",
    "The highest academic qualification completed — determines the base degree_level score",
    "PHD (Doctorate — highest academic qualification, implies deep research expertise) → 2 pts\n"
    "MASTER (M.Tech, M.Sc, MBA, MCA, MS, M.E.) → 2 pts\n"
    "BACHELOR (B.Tech, B.Sc, BCA, B.E., B.Com) → 1.5 pts\n"
    "DIPLOMA (Polytechnic diploma, post-secondary diploma) → 1 pt\n"
    "UNKNOWN (cannot determine from text) → 0.5 pts",
    "Python: keyword match on degree text — 'B.Tech'/'Bachelor' → BACHELOR, 'M.Tech'/'Master' → MASTER, 'Ph.D'/'Doctorate' → PHD", "HIGH",
    "PHD/MASTER=2pts | BACHELOR=1.5pts | DIPLOMA=1pt | UNKNOWN=0.5pts",
    "CRITICAL",
    "What is the highest degree you've completed? e.g. B.Tech, M.Tech, MBA, PhD?",
    "—",
    "")

add("ED2","Degree Level","Education",
    "Field of Study (Canonical)",
    "The normalised field of study after mapping through our course dictionary",
    "ENGINEERING (B.Tech/M.Tech in any engineering stream)\n"
    "COMPUTER_SCIENCE_IT (CS, CSE, IT, MCA)\n"
    "DATA_ANALYTICS (Data Science, Analytics, Statistics, Quant Econ)\n"
    "MANAGEMENT (MBA, PGDM)\n"
    "SCIENCE (B.Sc, M.Sc in Physics/Maths/Chemistry)\n"
    "ARTS_COMMERCE (BBA, B.Com, BA)\n"
    "RESEARCH (PhD, Research Scholar)",
    "Python: longest-match on course dictionary", "MEDIUM",
    "Feeds education_job_relevance score",
    "IMPORTANT",
    "What was your specialisation within [Degree]? e.g. Computer Science, Electronics, Mathematics?",
    "—",
    "")

add_section("ED3 — EDUCATION GAP  |  1 pt  |  Method: Python — REJECT FLAG IF >12m")

add("ED3","Education Gap","Education",
    "Gap Duration (months)",
    "The gap in months between the end of the candidate's last degree and the start of their first job. "
    "Large gaps (>12 months) are flagged as a reject condition.",
    "Integer — months\n0–6 months → normal (1 pt)\n6–12 months → minor gap (0.5 pt)\n>12 months → REJECT FLAG (0 pt)",
    "Python: last_education_end_date → first_job_start_date delta", "HIGH",
    "≤6m=1pt | 6–12m=0.5pt | >12m=0pt + REJECT FLAG",
    "CRITICAL",
    "After completing your [Degree] in [Year], your first job appears to start in [Year+X]. "
    "What were you doing in the [X] months gap?",
    "—",
    "MBA overlapping gap = no penalty. Competitive exam preparation (UPSC/GATE) = context note.")

add_section("ED4–ED6 — EDUCATION RELEVANCE, EXEC EDUCATION, PATENTS  |  5 pts combined")

add("ED4","Education to Job Relevance","Education",
    "Course Relevance Signal",
    "How relevant the candidate's degree is to the target role — "
    "a Computer Science degree for a Data Scientist role is HIGH; a History degree is FOUNDATIONAL",
    "HIGH (2 pts): CS, Engineering, IT, MCA, PhD in tech, Data Science, Statistics, Analytics, "
    "Quantitative Economics, Econometrics, Operations Research, Actuarial\n"
    "MEDIUM (1.5 pts): B.Sc/M.Sc, MBA, Management, Science\n"
    "FOUNDATIONAL (0.5 pts): Arts, Commerce, Law, History\n"
    "UNKNOWN (1 pt): Cannot determine relevance",
    "Python: course_family + field_tech_fit → maps to relevance signal", "MEDIUM",
    "HIGH=2pts | MEDIUM=1.5pts | FOUNDATIONAL=0.5pts | UNKNOWN=1pt",
    "IMPORTANT",
    "Your degree is in [Field] — how did your academic training prepare you for a career in [Target Role]?",
    "—",
    "")

add("ED5","Executive Education","Education",
    "Exec / Distance / Continuing Education",
    "Any additional courses, certifications, or programmes taken after the primary degree — "
    "especially executive education, online learning, or professional development",
    "Y (1 pt): executive programme, continuing education, distance learning, MOOC, "
    "online certification from recognised institute (e.g. Coursera specialisation, IIM executive programme)\n"
    "N (0 pts)",
    "Python keyword: 'executive', 'continuing', 'distance', 'certification', 'online', 'mooc', 'e-learning'", "MEDIUM",
    "Boolean: 1 or 0 pt",
    "NICE_TO_HAVE",
    "Have you done any executive education programmes, online certifications, or continuing education "
    "after your primary degree?",
    "—",
    "Must be from a recognised institution to count. Random Udemy courses do not qualify.")

add("ED6","Patents / Publications","Education",
    "Patent / Publication Count",
    "Has the candidate filed patents, published papers, or presented at conferences? "
    "These are among the strongest signals of deep expertise.",
    "Patents: count of filed (applied or granted) patents\n"
    "Publications: count of papers in journals or conferences\n"
    "Conference Talks: count of invited speaker slots at known tech conferences\n"
    "Combined → 2 pts if any exist\n"
    "Special: TIER_1 institution = 0.5 base credit even without patents (elite calibration)",
    "Python: 'patent', 'filed', 'granted', 'published', 'paper', 'journal', 'conference', 'arxiv'", "MEDIUM",
    "Any confirmed patent or publication → 2 pts. TIER_1 no patent → 0.5 base credit.",
    "IMPORTANT",
    "Do you have any patents filed or granted, or papers published in journals or conferences?",
    "Tell me about your publication/patent — what was the contribution, where was it published, "
    "and what was its impact in the field?",
    "")

add("ED7","LinkedIn / Social Media","Education",
    "LinkedIn Profile Activity Level",
    "How active and complete the candidate's LinkedIn profile is — "
    "an active LinkedIn signals professional engagement and thought leadership",
    "NOT_PRESENT (0 pts): No LinkedIn profile or completely empty\n"
    "LESS_ACTIVE (0.25 pts): Profile exists but rarely updated, no recent posts\n"
    "MORE_ACTIVE (0.5 pts): Regular updates, recommendations, connections >200\n"
    "REGULAR_UPDATE (1 pt): Frequent posts, articles, endorsements, visible thought leadership",
    "RECRUITER checks profile during screening call. Cannot be auto-detected.", "N/A — Recruiter only",
    "0 pts (not present) to 1 pt (regular updates)",
    "IMPORTANT",
    "Can you share your LinkedIn profile URL? I'd like to check it during our call.",
    "—",
    "Education scoring Updated.xlsx specifies these exact 4 bands with scores 0/0.25/0.5/1.0")

add("ED8","Extra-Curricular Activities","Education",
    "Extra-Curricular Activities",
    "Non-academic activities that show leadership, teamwork, and personal drive — "
    "relevant especially for early-career candidates",
    "Y (1 pt): Placement coordinator, event organiser, sports captain, cultural lead, "
    "hackathon participant (not winner — that's an award), student club member in a leadership role, "
    "NSS/NCC, volunteering in technical capacity\n"
    "N (0 pts): No activities mentioned",
    "Python: keywords in education section or resume footer. Recruiter confirms.", "LOW",
    "Boolean: 1 or 0 pt",
    "NICE_TO_HAVE",
    "Were you involved in any clubs, societies, sports, or activities during college — "
    "especially any leadership roles like placement committee or event management?",
    "—",
    "")

# ══════════════════════════════════════════════════════════════════════════
# SKILLS
# ══════════════════════════════════════════════════════════════════════════
add_section("S1 — SKILL LIST / YEARS / TIMELINE  |  6 pts  |  Method: Python — RECRUITER VALIDATES")

add("S1","Skill List — Years & Timeline","Skills",
    "Skill Name (per skill)",
    "The name of each technical skill on the resume, after normalisation to a canonical form",
    "Canonical examples: Python, SQL, Apache Spark, TensorFlow, PyTorch, scikit-learn, "
    "AWS (S3/EC2/Lambda/SageMaker), Azure ML, GCP, Hadoop, Kafka, Airflow, dbt, "
    "Tableau, Power BI, PostgreSQL, MongoDB, Docker, Kubernetes, Git, Jira",
    "Python NLP extracts from skills section + role descriptions. Alias mapping normalises variants "
    "(e.g. 'Py', 'Python3', 'Python 3.x' → 'Python')", "HIGH",
    "Each validated APPLIED+ skill with confirmed years = 1 pt. Max 6 pts.",
    "CRITICAL",
    "You've listed [Skill] — can you confirm approximately how many years you've been using it "
    "and when you last used it?",
    "—",
    "Recruiter must go through each key skill and confirm: years of use + last used date")

add("S1","Skill List — Years & Timeline","Skills",
    "Years of Active Use (per skill)",
    "How many years the candidate has actively used this specific skill in a professional context — "
    "not just listed it on the resume",
    "Float — e.g. 4.5 years\nCalculated from first mention to last mention across all roles",
    "Python: weighted by role durations where skill appears. Weighted by evidence quality.", "MEDIUM",
    "Only APPLIED+ skills with confirmed years count toward S1 (1 pt each, max 6)",
    "CRITICAL",
    "For [Skill] — when did you first start using it professionally, and when did you last use it? "
    "Are you currently using it?",
    "—",
    "Raw years vs weighted years differ — raw = calendar years listed, weighted = credible usage years")

add("S1","Skill List — Years & Timeline","Skills",
    "Last Used Date (per skill)",
    "When the candidate most recently used this skill — determines recency classification",
    "MM/YYYY or 'Present'/'Current'\nInferred from the end date of the last role where skill appears",
    "Python: cross-references skill mentions with role dates", "MEDIUM",
    "Feeds skill_recency (S6) — RECENT if <2 years, MID if 2–5 years, OLD if >5 years",
    "CRITICAL",
    "When did you last actively use [Skill] in a project?",
    "In your current role, are you still using [Skill] regularly?",
    "")

add_section("S4 — CERTIFICATIONS  |  3 pts  |  Method: Python + LLM")

add("S4","Certifications","Skills",
    "Certification Name",
    "The full name of the certification",
    "Examples: AWS Certified Solutions Architect, Google Professional Data Engineer, "
    "Azure DP-100, PMP, CFA, CISSP, Databricks Certified Associate Developer, "
    "TensorFlow Developer Certificate, Coursera Deep Learning Specialisation",
    "Python: extracts from certifications section", "HIGH",
    "Each valid cert = 1 pt. Max 3 pts.",
    "IMPORTANT",
    "Tell me about your certifications — which ones are currently valid?",
    "—",
    "")

add("S4","Certifications","Skills",
    "Certification Validity",
    "Is the certification still currently valid? Some certifications (like AWS) expire after 3 years.",
    "VALID (not expired)\n"
    "EXPIRED (past expiry date)\n"
    "LIFETIME (no expiry — e.g. most Coursera specialisations)\n"
    "UNKNOWN (no expiry date mentioned)",
    "Python: checks year obtained + known expiry windows for popular certs", "LOW",
    "Only VALID/LIFETIME certs count. EXPIRED certs = 0 pts for that cert.",
    "IMPORTANT",
    "Is your [Certification] still valid? When did you obtain it and when does it expire?",
    "—",
    "")

add("S4","Certifications","Skills",
    "Certification Relevance to Role",
    "How relevant is this certification to the target job role?",
    "HIGH (directly relevant to the JD skills — e.g. AWS cert for a Cloud Engineer role)\n"
    "MEDIUM (partially relevant — useful but not core to the role)\n"
    "LOW (not relevant to the target role — e.g. PMP for a coding role)\n"
    "IRRELEVANT (completely unrelated)",
    "LLM: maps cert to JD required skills", "MEDIUM",
    "Only relevant certs count toward score. Irrelevant certs do not add pts.",
    "IMPORTANT",
    "For the role you're applying to, which of your certifications do you feel are most relevant?",
    "—",
    "Cert farming red flag: ≥5 certs AND skill_depth ≤ FOUNDATIONAL across all certified skills")

add_section("S5 — SKILL DEPTH  |  8 pts  |  Method: BERT Primary — Panel Validates")

add("S5","Skill Depth","Skills",
    "Evidence Level (per skill)",
    "How strongly evidenced is this skill from the resume — this is the raw evidence signal before BERT blending",
    "NONE (skill not mentioned at all)\n"
    "MENTION (listed in skills section only — no evidence of actual use) = 0.5 pts\n"
    "WEAK (mentioned briefly in a role but no context or depth) = 1.5 pts\n"
    "APPLIED (used in specific projects/roles with some context, normal professional use) = 3.0 pts\n"
    "DEEP (used repeatedly, in complex contexts, with architecture or design involvement) = 4.0 pts\n"
    "EXPERT (open source contributions, patents, publications, or architected major systems using this skill) = 5.0 pts",
    "Python rule engine: evidence_level based on context richness, role mentions, project mentions, "
    "coding signal, architecture signal, OSS signal", "MEDIUM",
    "(evidence_score / 5) × 8 before BERT blending",
    "CRITICAL",
    "You've listed [Skill] on your resume. Can you describe a specific project where you used it "
    "and what you built with it?",
    "Walk me through a technically complex use of [Skill]. What was the problem, your approach, "
    "and what design decisions did you make?",
    "MENTION with no evidence = flagged. Panel must convert MENTION/WEAK to APPLIED/DEEP or note as inflated.")

add("S5","Skill Depth","Skills",
    "Depth Label (per skill)",
    "The final categorised depth level for this skill after BERT blending",
    "AWARENESS (knows the concept, may have used it once superficially — 0.5 pts)\n"
    "FOUNDATIONAL (can use it for basic tasks, has done so professionally — 1.5 pts)\n"
    "HANDS_ON (comfortable using it for standard professional work — 3.0 pts)\n"
    "ADVANCED (can solve complex problems, make architectural decisions — 4.0 pts)\n"
    "ARCHITECT_LEVEL (can design systems around this skill, has led others in using it — 5.0 pts)",
    "Python + BERT blended. BERT confidence tiers determine blend ratio.", "MEDIUM",
    "(depth_score / 5) × 8 for top-5 skills average",
    "CRITICAL",
    "—",
    "On a scale from beginner to expert in [Skill] — where would you place yourself and why? "
    "Give me an example of the most complex thing you've done with it.",
    "")

add("S5","Skill Depth","Skills",
    "Coding Signal (per skill)",
    "Was this skill used to write actual code — build, implement, script, deploy — vs just configuring or using a GUI?",
    "Y (coding signal detected: built/developed/implemented/scripted/coded/automated/deployed with this skill)\n"
    "N (no coding evidence — only mentioned or used in advisory capacity)",
    "Python keyword: 'built', 'developed', 'implemented', 'script', 'code', 'API', 'pipeline', 'automation'", "MEDIUM",
    "Coding signal → promotes evidence_level. No coding signal → evidence capped at WEAK",
    "IMPORTANT",
    "When you used [Skill] — were you writing code yourself, or more configuring/managing it at a higher level?",
    "Show me code or walk me through the code you wrote using [Skill].",
    "")

add("S5","Skill Depth","Skills",
    "Architecture Signal (per skill)",
    "Was this skill used at a system design / architectural level — not just implementation but design decisions?",
    "Y (architecture signal: 'designed', 'architected', 'system design', 'scalable', 'migration design', 'led the architecture')\n"
    "N",
    "Python keyword: 'designed', 'architected', 'system design', 'scalable', 'migration design'", "LOW",
    "Architecture signal → promotes evidence_level to DEEP or higher",
    "IMPORTANT",
    "Did you design the architecture for the system that used [Skill], or were you implementing "
    "someone else's design?",
    "Walk me through the system architecture — what design decisions did you make around [Skill]?",
    "")

add("S5","Skill Depth","Skills",
    "Open Source Signal (per skill)",
    "Has the candidate contributed to open source projects using this skill?",
    "Y (GitHub contributions, PRs merged, OSS projects, notable repositories)\n"
    "N",
    "Python keyword: 'GitHub', 'contributor', 'pull request', 'open source', 'maintained'", "LOW",
    "OSS signal → promotes to EXPERT level for that skill",
    "NICE_TO_HAVE",
    "Do you have any GitHub repositories or open source contributions that use [Skill]?",
    "Walk me through your most significant open source contribution using [Skill].",
    "")

add_section("S6 — SKILL RECENCY  |  6 pts  |  Method: Python")

add("S6","Skill Recency","Skills",
    "Recency Classification (per skill)",
    "How recently the candidate actively used this skill in professional work",
    "RECENT (used within the last 2 years — full weight in recency score)\n"
    "MID (used 2–5 years ago — partial weight)\n"
    "OLD (last used >5 years ago — low weight)\n"
    "UNKNOWN (cannot determine from dates)",
    "Python: compares skill's last_used_date against current date", "HIGH",
    "(count_RECENT_or_CURRENT / total_skills) × 6",
    "CRITICAL",
    "When did you last use [Skill] actively in a project?",
    "Are you currently using [Skill] in your day-to-day work?",
    "")

add_section("S7 — SKILLS LEARNING ACUMEN  |  3 pts  |  Method: Python + LLM")

add("S7","Skills Learning Acumen","Skills",
    "Yearly New Skill Acquisition",
    "How many NEW skills has the candidate picked up each year — demonstrating continuous learning",
    "List per year — e.g. {2021: ['Kafka', 'dbt'], 2022: ['LangChain', 'RAG'], 2023: ['CrewAI']}\n"
    "fast_learner: ≥2 new skills per year for ≥2 consecutive years\n"
    "moderate: 1–2 new skills per year\n"
    "stagnant: no new skills for 2+ years",
    "Python: yearly_skill_learning() — tracks first-appearance year per skill", "MEDIUM",
    "fast_learner (≥2/yr for ≥2yrs) = 3pts | New skills ≥3 yrs = 2pts | 1–2 yrs = 1pt | none = 0pts",
    "IMPORTANT",
    "In the last 2–3 years, what new technologies have you learned and applied in your work?",
    "How do you typically stay current with new developments in your field? "
    "Can you give an example of a skill you recently picked up and used in a project?",
    "")

add_section("S8 — CODING COMMUNITY / OPEN SOURCE  |  3 pts  |  Method: Python")

add("S8","Coding Community","Skills",
    "Platform Presence (per platform)",
    "Is the candidate active on coding/community platforms — this signals engagement with the tech community beyond their day job",
    "GitHub (has repositories, contributions, stars — strongest signal)\n"
    "Stack Overflow (answers questions — shows willingness to help + deep knowledge)\n"
    "LeetCode (solves algorithm problems — relevant for interview prep + problem-solving)\n"
    "HackerRank (competitive programming)\n"
    "Kaggle (data science competitions — very relevant for ML/DS roles)\n"
    "Other (personal blog, dev.to, medium articles on tech)",
    "Python: profile URLs or platform names in resume", "LOW",
    "≥3 platforms = 3pts | 2 = 2pts | 1 = 1pt | 0 = 0pts",
    "NICE_TO_HAVE",
    "Do you have a GitHub profile? Can you share your GitHub URL? "
    "Are you active on any other coding platforms like Kaggle or LeetCode?",
    "Walk me through your most significant GitHub repository — what problem does it solve?",
    "")

add_section("S9 — COMMUNICATION SKILLS  |  5 pts  |  Method: PANEL ONLY")

add("S9","Communication Skills","Skills",
    "Verbal Clarity",
    "How clearly does the candidate express technical concepts verbally — can they explain complex things simply?",
    "1 = Very Poor (rambling, incoherent, cannot complete sentences)\n"
    "2 = Poor (frequent hesitation, unclear explanations)\n"
    "3 = Acceptable (gets the point across but not fluent)\n"
    "4 = Good (clear, well-structured explanations)\n"
    "5 = Excellent (articulate, precise, adjusts complexity to audience)",
    "PANEL evaluates during live interview", "N/A — Panel only",
    "Part of S9 aggregate (total 5 pts for communication_skills)",
    "CRITICAL",
    "N/A",
    "Explain [technical concept relevant to their work] as if I'm a non-technical business stakeholder.",
    "")

add("S9","Communication Skills","Skills",
    "Logical Structure",
    "Does the candidate organise their answers logically — beginning, middle, end — or do they ramble?",
    "1 = No structure (jumps around, no logical flow)\n"
    "2 = Partially structured\n"
    "3 = Acceptable structure\n"
    "4 = Well-structured (clear problem → approach → outcome)\n"
    "5 = Excellent structure (STAR format, crisp transitions)",
    "PANEL evaluates during live interview", "N/A — Panel only",
    "Part of S9 aggregate",
    "IMPORTANT",
    "N/A",
    "Tell me about a challenging technical problem you solved — walk me through it from start to finish.",
    "")

add("S9","Communication Skills","Skills",
    "Audience Adaptability",
    "Can the candidate adjust their communication style — technical when talking to engineers, "
    "simplified when talking to business stakeholders?",
    "1 = Cannot adapt (one style only)\n"
    "2 = Slightly adaptable\n"
    "3 = Moderate — responds to cues\n"
    "4 = Good adaptability\n"
    "5 = Excellent — naturally adjusts depth and jargon",
    "PANEL evaluates during live interview", "N/A — Panel only",
    "Part of S9 aggregate",
    "IMPORTANT",
    "N/A",
    "Now explain the same concept to me as if I'm a senior business executive with no technical background.",
    "")

add_section("S10 — DOMAIN SKILLS  |  5 pts  |  Method: PANEL ONLY")

add("S10","Domain Skills","Skills",
    "Domain Depth Score",
    "How deeply does the candidate understand the business domain they claim to work in — "
    "not just the tech, but the domain-specific concepts, constraints, and nuances",
    "0–5 scale:\n"
    "0–1 = No domain understanding (just uses the tech without knowing the business context)\n"
    "2 = Basic (knows the terminology, understands the surface level)\n"
    "3 = Functional (can discuss domain problems, suggest solutions, knows key metrics)\n"
    "4 = Deep (understands regulatory context, market dynamics, domain-specific ML challenges)\n"
    "5 = Expert (thought leader level — can educate others on domain, knows edge cases)",
    "PANEL: scenario-based domain questions", "N/A — Panel only",
    "0–5 pts for domain_skills",
    "CRITICAL",
    "N/A",
    "In [Candidate's Domain — e.g. BFSI]: Walk me through how you would build a credit risk model. "
    "What are the key regulatory constraints? What data challenges would you face?",
    "Domain questions should be tailored to the candidate's specific claimed domain")

add_section("S11 — PROJECT EXPLANATION SKILLS  |  3 pts  |  Method: Recruiter + Panel")

add("S11","Project Explanation Skills","Skills",
    "Project Walk-through Quality",
    "Can the candidate explain their project clearly — problem, approach, their role, and outcome? "
    "This tests both communication and depth of ownership.",
    "0 = Cannot explain at all (vague or says 'I don't remember')\n"
    "1 = Disjointed (explains parts but no coherent narrative)\n"
    "2 = Good structure (explains P→Solution but misses impact or role clarity)\n"
    "3 = Excellent (clear Problem → Solution Design → Their Role → Quantified Outcome)",
    "RECRUITER scores during pre-screen. PANEL re-scores during interview.", "N/A — Human only",
    "0–3 pts",
    "CRITICAL",
    "Can you walk me through your most recent project — what was the business problem, "
    "your technical approach, your specific role, and what the outcome was?",
    "You mentioned [Project]. Walk me through the technical architecture — "
    "how did the different components connect, and what was your specific contribution to each?",
    "")

add_section("S12 — CODING SKILLS  |  Qualitative — No Direct Score")

add("S12","Coding Skills","Skills",
    "Coding Problem Level",
    "What difficulty level of coding problem the candidate can solve confidently",
    "EASY (basic syntax, simple loops, array manipulation — junior level)\n"
    "MEDIUM (data structures, recursion, sorting, basic algorithms — mid level)\n"
    "HARD (dynamic programming, graph algorithms, system design constraints — senior level)\n"
    "EXPERT (novel algorithm design, optimisation under extreme constraints)",
    "PANEL: live coding session", "N/A — Panel only",
    "Qualitative — informs skill_depth narrative but no direct score added",
    "CRITICAL",
    "N/A",
    "Write [code problem appropriate to role level] and walk me through your approach.",
    "Problem difficulty should be calibrated to the role level — don't give senior problems to mid-level candidates")

add("S12","Coding Skills","Skills",
    "Code Quality Signals",
    "Assessment of the quality of code written — beyond just correctness",
    "Clean code (Y/N): meaningful variable names, clear structure\n"
    "Edge cases handled (Y/N): null inputs, empty arrays, overflow\n"
    "Time complexity awareness (Y/N): knows Big-O, can discuss trade-offs\n"
    "Modular thinking (Y/N): breaks problem into functions/classes\n"
    "Testing awareness (Y/N): considers how to test the code",
    "PANEL: observes during live coding", "N/A — Panel only",
    "Qualitative narrative",
    "IMPORTANT",
    "N/A",
    "After they write the solution: What is the time and space complexity? How would you test this? "
    "What edge cases should we handle?",
    "")

add_section("S14 — PROBLEM SOLVING SKILLS  |  3 pts  |  Method: PANEL ONLY")

add("S14","Problem Solving Skills","Skills",
    "Problem Structuring",
    "Does the candidate structure unfamiliar problems logically before jumping to a solution?",
    "1 = Jumps to solution without understanding problem\n"
    "2 = Partially structures problem\n"
    "3 = Good — clarifies requirements, breaks down problem\n"
    "4 = Excellent — systematic decomposition, identifies edge cases upfront\n"
    "5 = Exceptional — frames problem clearly, considers multiple approaches, justifies chosen approach",
    "PANEL: presents an open-ended business/technical problem", "N/A — Panel only",
    "Part of S14 aggregate (0–3 pts)",
    "CRITICAL",
    "N/A",
    "[Present scenario]: 'You are hired as the first data scientist at a new fintech startup. "
    "The CEO wants to reduce loan default rates by 20%. Where do you start?'",
    "")

add("S14","Problem Solving Skills","Skills",
    "Hints Needed",
    "Did the candidate need hints to progress — indicates how independently they can work",
    "0 hints (solved independently)\n"
    "1–2 hints (minor nudges needed)\n"
    "3+ hints (needed significant guidance)",
    "PANEL tracks during problem solving", "N/A — Panel only",
    "Fewer hints = higher problem solving score",
    "IMPORTANT",
    "N/A",
    "—",
    "Panel should only give hints if candidate is completely stuck for >3 minutes")

# ─────────────────────────────────────────────────────────────────────────────
# BUILD THE SHEET
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Data Dictionary"
ws.freeze_panes = "D5"

# Banner rows
merge(ws, 1, 1, 13,
      "RESUME INTELLIGENCE — COMPLETE GRANULAR DATA DICTIONARY",
      "1F3864", size=14)
merge(ws, 2, 1, 13,
      "Every atomic field for every parameter — definition, all possible values, "
      "how detected, score impact, and EXACT question for recruiter / panel when missing",
      "2E75B6", size=10)
merge(ws, 3, 1, 13,
      "COLOUR KEY:  Green rows = Experience  |  Yellow = Skills  |  Orange = Education  |  "
      "Blue = Recruiter captures  |  Pink = Panel captures  |  "
      "CRITICAL priority = must have  |  IMPORTANT = strong signal  |  NICE_TO_HAVE = enrichment",
      "37474F", fg="ECEFF1", size=9)

# Column headers
for ci, col in enumerate(COLS, 1):
    cell(ws, 4, ci, col, bg="1F3864", bold=True, color="FFFFFF",
         size=9, halign="center", wrap=True)
ws.row_dimensions[4].height = 32

SECTION_COLOURS = {
    "E": ("1E4620", "E8F5E9", "C8E6C9"),
    "S": ("4A148C", "F3E5F5", "E1BEE7"),
    "ED": ("7B3B00", "FFF3E0", "FFE0B2"),
}
PRIO_COLOURS = {"CRITICAL": "FFCDD2", "IMPORTANT": "FFF9C4", "NICE_TO_HAVE": "F1F8E9"}
REL_COLOURS = {"HIGH": "C8E6C9", "MEDIUM": "FFF9C4", "LOW": "FFCCBC",
               "N/A — Panel only": "E1BEE7", "N/A — Recruiter only": "BBDEFB"}

row_n = 5
alt = False
for r in ALL_ROWS:
    if r[0] == "SECTION":
        ws.row_dimensions[row_n].height = 20
        merge(ws, row_n, 1, 13, f"  {r[1]}",
              "263238" if "pts" not in r[1] else
              ("1E4620" if "E" in r[1][:4] else
               "4A148C" if "S" in r[1][:4] else "7B3B00"),
              size=10)
        row_n += 1
        alt = False
        continue

    (pid, pname, section, field, defn, vals, auto, rel,
     score, prio, rec_q, panel_q, notes) = r

    # row background
    s = pid[:2].rstrip("0123456789")
    col_set = SECTION_COLOURS.get(s, ("263238", "ECEFF1", "CFD8DC"))
    bg = col_set[2] if alt else col_set[1]
    alt = not alt

    ws.row_dimensions[row_n].height = 80

    cell(ws, row_n, 1, pid, bg=bg, bold=True, halign="center", wrap=False)
    cell(ws, row_n, 2, pname, bg=bg, bold=True)
    cell(ws, row_n, 3, section, bg=bg, halign="center", wrap=False)
    cell(ws, row_n, 4, field, bg=bg, bold=True)
    cell(ws, row_n, 5, defn, bg="FFFFFF")
    cell(ws, row_n, 6, vals, bg="F9F9F9")
    cell(ws, row_n, 7, auto, bg="F5F5F5")
    cell(ws, row_n, 8, rel, bg=REL_COLOURS.get(rel, "FFFFFF"), halign="center", wrap=False)
    cell(ws, row_n, 9, score, bg="FFF8E1")
    cell(ws, row_n, 10, prio, bg=PRIO_COLOURS.get(prio, "FFFFFF"),
         halign="center", bold=(prio=="CRITICAL"), wrap=False)
    cell(ws, row_n, 11, rec_q, bg="E3F2FD" if rec_q != "N/A" else "F5F5F5")
    cell(ws, row_n, 12, panel_q, bg="FCE4EC" if panel_q not in ("N/A","—","") else "F5F5F5")
    cell(ws, row_n, 13, notes, bg="FAFAFA", italic=True)

    row_n += 1

widths(ws, COL_W)
ws.row_dimensions[1].height = 24
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 16

out = "E:/Dev/resume_intelligence/Resume_Intelligence_Granular_DataDictionary.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total rows: {row_n}")
