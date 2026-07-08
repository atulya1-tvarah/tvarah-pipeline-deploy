"""
Resume Intelligence — Resume Data Checklist
For each data point:
  Column A: Section
  Column B: Parameter
  Column C: Data Point Name
  Column D: What exactly to look for in the raw resume
  Column E: Score impact (brief)
  Column F: [BLANK — human fills] Found on Resume? Y/N
  Column G: If NOT on resume — RECRUITER question
  Column H: If STILL missing after recruiter — PANEL question
  Column I: Priority
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Resume Data Checklist"
ws.freeze_panes = "A6"

def fill(h): return PatternFill("solid", fgColor=h)
def f(bold=False, color="000000", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)
def al(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="top", wrap_text=wrap)
thin = Side(style="thin", color="BDBDBD")
med  = Side(style="medium", color="9E9E9E")
B = Border(left=thin, right=thin, top=thin, bottom=thin)
BM = Border(left=med, right=med, top=med, bottom=med)

def c(row, col, val, bg="FFFFFF", bold=False, color="000000",
      size=9, ha="left", wrap=True, italic=False):
    cl = ws.cell(row=row, column=col, value=val)
    cl.fill = fill(bg)
    cl.font = f(bold=bold, color=color, size=size, italic=italic)
    cl.alignment = al(ha, wrap)
    cl.border = B
    return cl

def m(r, c1, c2, val, bg, fg="FFFFFF", sz=11, bold=True):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cl = ws.cell(row=r, column=c1, value=val)
    cl.fill = fill(bg)
    cl.font = f(bold=bold, color=fg, size=sz)
    cl.alignment = al("center", False)
    cl.border = BM

# ── Colour map ────────────────────────────────────────────────────────────
SEC = {   # section_key: (header_bg, odd_bg, even_bg, text_hdr)
    "EXP":  ("1B5E20", "E8F5E9", "DCEDC8", "FFFFFF"),
    "EDU":  ("4E342E", "FBE9E7", "FFCCBC", "FFFFFF"),
    "SKL":  ("4A148C", "F3E5F5", "E1BEE7", "FFFFFF"),
}
PRIO_BG = {"CRITICAL": "FFCDD2", "IMPORTANT": "FFF9C4", "NICE": "F1F8E9"}

# ── Header rows ───────────────────────────────────────────────────────────
m(1, 1, 9, "RESUME INTELLIGENCE — RAW RESUME DATA CHECKLIST", "1F3864", sz=14)
m(2, 1, 9,
  "This sheet lists every data point the system extracts from the resume.  "
  "If a field is BLANK / MISSING on the resume, the recruiter must ask the question in Column G.  "
  "If STILL MISSING after recruiter call, the panel asks the question in Column H.",
  "2E75B6", sz=10)
m(3, 1, 9,
  "HOW TO USE:  For each candidate, go row by row.  Mark Column F as Y (found) or N (missing).  "
  "For every N — use the recruiter question in Column G during the screening call.  "
  "Anything still N after recruiter call — panel probes using Column H.",
  "37474F", fg="ECEFF1", sz=9)
m(4, 1, 9,
  "COLOUR KEY:  Green = Experience  |  Orange = Education  |  Purple = Skills  |  "
  "Red row = CRITICAL (missing this = problem)  |  Yellow = IMPORTANT  |  White = NICE TO HAVE",
  "546E7A", fg="ECEFF1", sz=9)

COLS = [
    "Section",
    "Parameter",
    "Data Point",
    "What to look for in the RAW RESUME\n(exact text/field/signal)",
    "Score Impact",
    "Found on\nResume?\n(Y / N)",
    "If NOT found on resume —\nRECRUITER asks this question",
    "If STILL MISSING after recruiter —\nPANEL asks this question",
    "Priority"
]
WIDTHS = [10, 20, 24, 40, 22, 8, 44, 44, 10]

for ci, col in enumerate(COLS, 1):
    cl = ws.cell(row=5, column=ci, value=col)
    cl.fill = fill("1F3864")
    cl.font = f(bold=True, color="FFFFFF", size=9)
    cl.alignment = al("center", True)
    cl.border = BM
ws.row_dimensions[5].height = 36

# ── DATA ──────────────────────────────────────────────────────────────────
# Each tuple: (sec_key, param, data_point, resume_look_for, score_impact,
#              recruiter_q, panel_q, priority)
ROWS = []

def section(key, label):
    ROWS.append(("SEC", key, label, "", "", "", "", "", ""))

def row(sec, param, dp, resume, score, rq, pq, prio="CRITICAL"):
    ROWS.append((sec, param, dp, resume, score, rq, pq, prio))

# ═══════════════════════════════════════════════════════════════════════════
# EXPERIENCE
# ═══════════════════════════════════════════════════════════════════════════
section("EXP", "EXPERIENCE SECTION  |  40 pts total")

# ── E1: Companies ──────────────────────────────────────────────────────────
row("EXP","E1 — Companies\n(5 pts max)","Company Name",
    "Look for: Company name written as a header or sub-header in the experience section.\n"
    "e.g.  'Google India Pvt Ltd', 'Tata Consultancy Services', 'XYZ Analytics Pvt Ltd'",
    "Used to look up tier (1–5).\nTier determines score: Tier 1=5pts, Tier 2=4pts, Tier 3=3pts, Tier 4=2pts, Tier 5=1pt",
    "What is the full official name of the company you worked at — is it [Name on Resume] or a parent/subsidiary?",
    "Which entity exactly did you work for — the holding company or a specific division/subsidiary?",
    "CRITICAL")

row("EXP","E1 — Companies\n(5 pts max)","Company Type\n(Product / Services / Consulting)",
    "Look for:\n- 'Product' in company description or candidate's own words in the role\n"
    "- Job title or description containing 'client delivery', 'consulting', 'advisory'\n"
    "- If it's TCS/Infosys/Wipro/Accenture/HCL = SERVICES automatically\n"
    "- If it's Flipkart/Zomato/CRED/Swiggy = PRODUCT automatically",
    "Determines DNA score and operating model.\nProduct companies → PRODUCT DNA.\nServices companies → CONSULTING DNA.",
    "Was [Company] a product company that built its own product for end users, or did it deliver projects and services to client companies?",
    "In your day-to-day work at [Company] — were you building the company's own product/platform, or building solutions for an external client?",
    "CRITICAL")

row("EXP","E1 — Companies\n(5 pts max)","Industry / Domain",
    "Look for: Industry keywords in the company name or role description.\n"
    "e.g.  'fintech' / 'banking' / 'insurance' / 'healthcare' / 'e-commerce' / 'logistics' / 'edtech'\n"
    "MISSING if: company name gives no domain hint (e.g. 'Infosys' — need to know which client domain)",
    "No direct score but feeds project domain tags.\nDomain continuity across roles → relevant YoE calculation.",
    "What industry does [Company] operate in — for example, banking, healthcare, e-commerce, logistics, SaaS?",
    "Which specific business domain were most of your projects at [Company] focused on — e.g. credit risk, supply chain, ad-tech?",
    "IMPORTANT")

row("EXP","E1 — Companies\n(5 pts max)","Company Size\n(Headcount)",
    "Look for: Rarely on resume.\n"
    "Sometimes in summary: 'startup of 30 people' / 'Series B company' / 'Fortune 500'\n"
    "MISSING for most candidates — must ask",
    "Proxy for problem scale. Large company = higher complexity expectation.\nFunded startup = quality signal for unknown companies.",
    "Roughly how many employees did [Company] have when you were there? And how big was your immediate team?",
    "How large was the engineering / data team you were part of at [Company]?",
    "IMPORTANT")

row("EXP","E1 — Companies\n(5 pts max)","Funding Stage\n(for startups)",
    "Look for: 'Series A / B / C', 'funded startup', 'pre-IPO', 'listed company'\n"
    "MISSING for most — especially for lesser-known startups",
    "Proxy for company quality. Series C+ funded = Tier 2 equivalent signal even if not in our database.",
    "Is [Company] a funded startup? If yes — what stage of funding? e.g. Series A, Series B, or is it listed?",
    "—",
    "IMPORTANT")

row("EXP","E1 — Companies\n(5 pts max)","Work Type\n(Build vs Maintain vs Consult)",
    "Look for: Verbs in role description:\n"
    "BUILD: 'developed', 'built', 'designed', 'launched', 'deployed'\n"
    "MAINTAIN: 'maintained', 'supported', 'monitored', 'bug fixes', 'BAU', 'L2/L3'\n"
    "CONSULT: 'delivered to client', 'client solutions', 'advisory'\n"
    "MISSING if description only lists tools with no action verbs",
    "Build work → higher project complexity score.\nMaintain work → lower complexity, maintenance project type.",
    "At [Company], was most of your work building new systems from scratch, maintaining existing ones, or delivering to external clients?",
    "What was the split between greenfield development and BAU / maintenance work at [Company]?",
    "IMPORTANT")

# ── E2: Overall Experience ─────────────────────────────────────────────────
row("EXP","E2 — Overall Experience\n(3 pts max)","Start Date (per role)",
    "Look for: Month + Year at the start of each role\n"
    "e.g.  'Jan 2019', 'March 2021', 'Jan '20'\n"
    "MISSING if: Only year given (e.g. '2019') — year alone is not enough for gap calculation",
    "Feeds total YoE and gap detection.\n10+ yrs=3pts | 6–10=2.5 | 4–6=2.0 | 2–4=1.5 | <2=1.0",
    "For [Company], can you confirm the month and year you started? The resume only shows the year.",
    "—",
    "CRITICAL")

row("EXP","E2 — Overall Experience\n(3 pts max)","End Date (per role)",
    "Look for: Month + Year at the end of each role, or 'Present' / 'Current'\n"
    "e.g.  'Jun 2023', 'Dec 2022 – Present'\n"
    "MISSING if: Only year given, or 'till date' without year",
    "Feeds duration, gap detection, and skill recency.",
    "For [Company], when did you leave? The resume shows [year] — was that early or late in the year?",
    "—",
    "CRITICAL")

row("EXP","E2 — Overall Experience\n(3 pts max)","Total Experience\n(calculated)",
    "Look for: Stated total in summary — 'X years of experience'\n"
    "If not stated: System calculates from start/end dates of all roles\n"
    "VERIFY: Candidate's stated total vs system-calculated total",
    "Band scoring: 10+=3pts | 6–10=2.5 | 4–6=2.0 | 2–4=1.5 | 1–2=1.0 | <1=0.5",
    "Your resume shows approximately [X] years of experience — does that include all your roles? Are there any roles not listed?",
    "—",
    "CRITICAL")

row("EXP","E2 — Overall Experience\n(3 pts max)","Relevant Experience\n(domain / skill match to JD)",
    "Look for: Roles where the domain OR skill cluster matches the target JD.\n"
    "MISSING if: Candidate has worked in multiple domains — need to isolate relevant years.\n"
    "e.g. 6 yrs in Data Engineering + 2 yrs in Data Science → for DS role, relevant = 2 yrs unless DE skills apply",
    "REJECT GATE: relevant YoE < 70% of JD minimum → auto reject flag.\nThis is the single most important gate.",
    "Of your [X] years of experience, how many years were you working directly in [Target Role / Domain]?",
    "Which of your roles do you consider most relevant to [Target JD]? Walk me through why.",
    "CRITICAL")

# ── E3: Career Progression ────────────────────────────────────────────────
row("EXP","E3 — Career Progression\n(3 pts max)","Job Title (per role)",
    "Look for: Official job title for each role\n"
    "e.g.  'Data Scientist', 'Senior ML Engineer', 'Associate Vice President'\n"
    "MISSING if: No title given — just company name and description",
    "Feeds seniority level mapping (1=IC to 6=VP).\nSeniority sequence across roles → progression score.",
    "What was your exact job title at [Company]?",
    "Was your title [X] a designation from the company's grading system, or a working title?",
    "CRITICAL")

row("EXP","E3 — Career Progression\n(3 pts max)","Promotion at Same Company\n(title change within same employer)",
    "Look for: Same company listed twice with different, higher-level titles\n"
    "e.g.  'Analyst → Senior Analyst at TCS' across two consecutive entries with same company\n"
    "MISSING if: Only one entry per company even if candidate was promoted",
    "+0.5 bonus to progression. Internal promotions show company valued the candidate.",
    "At [Company], did you receive any promotions or title changes during your time there? If yes, from what title to what?",
    "—",
    "IMPORTANT")

row("EXP","E3 — Career Progression\n(3 pts max)","Seniority Trajectory\n(IC to Lead to Manager etc.)",
    "Look for: Sequence of titles across all companies and time.\n"
    "FAST_TRACK: Analyst → Senior → Lead → Manager in 4–5 years\n"
    "GROWING: Gradual upward movement\n"
    "LATERAL: Same title / same level across multiple companies\n"
    "DECLINING: Senior title → lower title (e.g. VP → Manager at next company)\n"
    "MISSING if: Vague titles that don't map to a level (e.g. 'Data Person', 'Tech Guy')",
    "BERT classifies: FAST_TRACK=3pts | GROWING=2.5 | LATERAL=1.5 | DECLINING=0.5",
    "Looking at your career — [Title A] at [Company 1] then [Title B] at [Company 2] — were these moves upward, lateral, or in a different direction?",
    "How did your scope and team size change from [Role A] to [Role B]? Did you take on more responsibility?",
    "CRITICAL")

# ── E4: Stability ─────────────────────────────────────────────────────────
row("EXP","E4 — Stability\n(3 pts max)","Duration per Company\n(months in each role)",
    "Look for: Start date to end date for each company.\n"
    "CONCERN: Any role <12 months\n"
    "RED FLAG: 2+ roles <12 months or hop_rate >1.5 companies/year\n"
    "MISSING if: Dates are only years (2020–2021 could be 1 month or 24 months)",
    "Avg tenure: 36+m=5.0 | 24–36=4.0 | 18–24=3.5 | 12–18=3.0 | <12m = penalty.\n<6m stints = -0.8 penalty each.",
    "You were at [Company] from [Start] to [End] — that's about [X] months. Can you confirm? And why did you leave after that time?",
    "—",
    "CRITICAL")

row("EXP","E4 — Stability\n(3 pts max)","Short Stint Context\n(reason for <12 month roles)",
    "Look for: Any explanation in the resume for a short tenure.\n"
    "e.g.  'contract role', 'startup shut down', 'laid off', 'company acquired'\n"
    "MISSING for most — candidates rarely explain short stints on the resume",
    "Context determines if penalty is applied.\nLayoff/shutdown/contract = penalty waived.\nVoluntary quick exit = full penalty.",
    "I see you were at [Company] for only [X] months — can you tell me why you left so quickly? Was it a layoff, contract role, company closure, or your own decision?",
    "—",
    "CRITICAL")

row("EXP","E4 — Stability\n(3 pts max)","Contract vs Permanent\n(employment type)",
    "Look for: Keywords in resume — 'contractor', 'freelance', 'consulting', 'fixed-term'\n"
    "MISSING for most — rarely stated explicitly",
    "Contract/consultant pattern detected → Archetype A10 → stability algorithm adjusted.\nShort contract stints not penalised the same as voluntary exits.",
    "Were any of your roles on a contract or freelance basis rather than as a permanent employee?",
    "—",
    "IMPORTANT")

# ── E5: Awards ────────────────────────────────────────────────────────────
row("EXP","E5 — Awards & Recognitions\n(3 pts max)","Named Awards / Prizes",
    "Look for: Explicit award names in achievements section or role descriptions\n"
    "e.g.  'Employee of the Quarter', 'Star Performer Award', 'Gold Medal', 'Forbes 30U30'\n"
    "NOT COUNTED: Generic 'best team player', participation certificates, training completion badges",
    "Count: 0=0pts | 1=1pt | 2=2pts | 3+=3pts.\nLLM verifies if genuine and competitive.",
    "The award '[Name]' on your resume — can you tell me more about it? Was it a competitive award, and what were you recognised for specifically?",
    "You mentioned [Award] — how many people were typically considered for it, and what did you specifically do to earn it?",
    "IMPORTANT")

row("EXP","E5 — Awards & Recognitions\n(3 pts max)","Promotions (count)",
    "Look for: 'Promoted to', 'Elevated to', title progression at same company\n"
    "MISSING if: Promotions happened but candidate only listed one title per company",
    "Each promotion = +1 to award count.",
    "Have you received any formal promotions across your career — a title change with a salary increase?",
    "—",
    "IMPORTANT")

row("EXP","E5 — Awards & Recognitions\n(3 pts max)","Patents Filed / Granted",
    "Look for: 'Patent', 'Filed', 'Patent No.', 'USPTO', 'Indian Patent Office'\n"
    "e.g.  'Filed patent for real-time anomaly detection system (USPTO 2022)'\n"
    "MISSING if: Patent exists but candidate didn't list it",
    "Patent detected → 2 bonus pts in education section (patents_publications).\nAlso counts as 1 recognition in E5.",
    "Do you have any patents filed or granted? If yes, can you share the patent title and year?",
    "Tell me about your patent — what was the core invention and what problem does it solve?",
    "IMPORTANT")

row("EXP","E5 — Awards & Recognitions\n(3 pts max)","Publications / Conference Talks",
    "Look for: 'Published', 'Paper', 'Conference', 'Journal', 'arxiv', 'IEEE', 'NeurIPS', 'ICML'\n"
    "MISSING if: Research done but not listed",
    "Publication → 2 bonus pts in education.\nCounts as 1 recognition in E5.",
    "Do you have any research papers published or conference talks given? If yes, on what topic?",
    "Walk me through your paper — what was the core contribution and how was it received by the community?",
    "IMPORTANT")

# ── E6: Mentorship ────────────────────────────────────────────────────────
row("EXP","E6 — Mentorship\n(3 pts max — Recruiter fills)","Mentored Engineers\n(count and context)",
    "Look for: 'Mentored X engineers', 'guided junior team members', 'led a team of X'\n"
    "MISSING for most — candidates rarely quantify mentorship on resume",
    "LEAD (2+ roles with mentorship)=3pts | FORMAL (1 instance)=2pts | IMPLIED=1pt | NONE=0pts",
    "Have you mentored junior engineers or team members? If yes — how many people, at which companies, and was it formal or informal?",
    "Tell me about a specific time you mentored someone — what did you teach them, how often did you meet, and what was the outcome for them?",
    "CRITICAL")

row("EXP","E6 — Mentorship\n(3 pts max — Recruiter fills)","Code Reviews\n(signal on resume)",
    "Look for: 'Code review', 'PR review', 'reviewed code written by'\n"
    "MISSING for most — not commonly written on resumes",
    "Contributes to mentorship BERT classification.",
    "In your current or recent role, were you responsible for reviewing code written by other team members?",
    "How did you approach code reviews — what specifically did you look for beyond just correctness?",
    "IMPORTANT")

row("EXP","E6 — Mentorship\n(3 pts max — Recruiter fills)","Interview Panel Participation",
    "Look for: 'Conducted technical interviews', 'part of hiring panel', 'interviewed candidates'\n"
    "MISSING for most",
    "Seniority signal. Contributes to mentorship classification.",
    "Have you been involved in interviewing candidates for your company — as a technical interviewer?",
    "What kind of interviews did you conduct — coding, system design, or behavioural?",
    "IMPORTANT")

# ── E7: International ─────────────────────────────────────────────────────
row("EXP","E7 — International Exposure\n(2 pts — Recruiter confirms)","Onsite / Overseas Work",
    "Look for: Country names (UK, US, Singapore, Dubai, Germany etc.), 'onsite', 'relocated to', 'visa'\n"
    "e.g.  'Worked onsite in London for 6 months', 'Travelled to client site in Singapore'\n"
    "NOT COUNTED: 'Global team' alone without physical travel",
    "Explicit onsite/relocation=2pts | Implied global team=1pt | None=0pts",
    "I see [London/Singapore/US] mentioned on your resume — did you actually travel to and work from there? For how long and for what purpose?",
    "Tell me about your international work experience — what country, for how long, and was it a client visit, relocation, or transfer?",
    "CRITICAL")

row("EXP","E7 — International Exposure\n(2 pts — Recruiter confirms)","Global Team / Multi-Timezone",
    "Look for: 'Global team', 'multi-timezone', 'cross-border', 'international stakeholders'\n"
    "MISSING if: Only listed domestic work but collaborated with overseas teams",
    "1 pt for implied global collaboration (without physical travel).",
    "Did you regularly work with teams or stakeholders in other countries, even if you were based in India?",
    "—",
    "IMPORTANT")

# ── E8: Stakeholder ───────────────────────────────────────────────────────
row("EXP","E8 — Stakeholder Management\n(2 pts — Recruiter validates)","Client-Facing Work",
    "Look for: 'client', 'customer', 'stakeholder', 'presented to', 'client requirements', 'client delivery'\n"
    "MISSING if: Description uses only internal terminology",
    "Client-facing=1.5pts | C-level=2pts | Internal only=1pt | None=0pts",
    "At [Company], were you directly working with external clients — taking requirements, presenting solutions, or managing their expectations?",
    "Give me an example of a difficult external stakeholder situation you had to navigate.",
    "CRITICAL")

row("EXP","E8 — Stakeholder Management\n(2 pts — Recruiter validates)","C-Level / Senior Leadership Exposure",
    "Look for: 'CTO', 'CEO', 'Board', 'presented to senior leadership', 'executive stakeholders', 'C-suite'\n"
    "MISSING for most — only senior candidates typically have this",
    "C-level exposure = 2pts (max score for this parameter).",
    "Did your work ever involve presenting to or directly advising C-level executives or board members?",
    "Tell me about a time you presented to C-level leadership — what was the audience, the topic, and the outcome?",
    "IMPORTANT")

row("EXP","E8 — Stakeholder Management\n(2 pts — Recruiter validates)","Cross-Functional Collaboration",
    "Look for: 'cross-functional', 'worked with product team', 'collaborated with business', 'multiple teams'\n"
    "MISSING if: Description is purely technical with no mention of other teams",
    "Internal cross-functional = 1pt.",
    "Did your role require collaboration with non-technical teams — product, business, operations, or finance?",
    "—",
    "IMPORTANT")

# ── E9: Career Breaks ─────────────────────────────────────────────────────
row("EXP","E9 — Career Breaks\n(2 pts max)","Gap Between Roles\n(dates don't connect)",
    "Look for: Gaps in the timeline — end date of one role to start date of next role > 3 months\n"
    "e.g.  'TCS: Jan 2019 – Dec 2020' then 'Flipkart: Apr 2021' → 3-month gap\n"
    "MISSING if: Only years given — gap cannot be precisely calculated",
    "0 breaks=2pts | 1=1pt | 2=0pts | 3+=REJECT FLAG",
    "I notice a gap from approximately [Date] to [Date] — about [X] months. What were you doing during that period?",
    "—",
    "CRITICAL")

row("EXP","E9 — Career Breaks\n(2 pts max)","Reason for Each Gap",
    "Look for: Explanation of gap in resume — 'pursuing MBA', 'parental leave', 'health reasons'\n"
    "MISSING for most — gaps are almost never explained on resumes",
    "Reason determines if penalty applies:\nMBA/education = EXEMPT | Maternity <=18m = SOFT FLAG | COVID 2020 = note | Layoff = EXEMPT | Voluntary = penalty",
    "The gap from [Date A] to [Date B] — were you studying, on parental leave, dealing with health issues, travelling, or between opportunities?",
    "—",
    "CRITICAL")

# ── E10/E11: Projects ─────────────────────────────────────────────────────
row("EXP","E10 — Project 1\n(8 pts max)","Project Title",
    "Look for: A named project heading in a projects section, or the largest described initiative in the most recent role\n"
    "e.g.  'Real-time Fraud Detection System', 'Customer Lifetime Value Model'\n"
    "MISSING if: No project section, or role description is just a bullet list of tools",
    "Criterion 2: Title present = +1pt.",
    "What is the name or headline you'd give to your most significant recent project?",
    "Walk me through your most recent project from the beginning — what were you building and why?",
    "CRITICAL")

row("EXP","E10 — Project 1\n(8 pts max)","Project Type\n(what kind of work)",
    "Look for: Keywords in project description:\n"
    "DEVELOPMENT: 'built', 'developed', 'designed', 'created', 'launched'\n"
    "MIGRATION: 'migrated', 'cloud migration', 'modernisation', 'lift and shift'\n"
    "ANALYTICS: 'analysis', 'dashboard', 'reporting', 'insights', 'forecasting'\n"
    "MAINTENANCE: 'support', 'bug fix', 'monitoring', 'BAU', 'L2/L3'\n"
    "MISSING if: Description only lists tech stack with no action words",
    "Criterion 1: Type known = +1pt.\nDEVELOPMENT/MIGRATION = highest complexity. MAINTENANCE = lowest.",
    "Was this project building something new, migrating from old systems, doing data analysis, or maintaining something existing?",
    "What was the core technical challenge — was it scale, accuracy, real-time processing, or something else?",
    "CRITICAL")

row("EXP","E10 — Project 1\n(8 pts max)","Project Description\n(richness / length)",
    "Look for: Any text describing WHAT the project did and HOW it was done\n"
    "RICH: Problem statement + technical approach + outcome (>100 words)\n"
    "ADEQUATE: Describes the project (50–100 words)\n"
    "MINIMAL: Very brief, just names tools (<50 words)\n"
    "MISSING: No description at all — just a project name",
    "Criterion 3: Description >20 chars = +1pt.\nRicher description → higher LLM complexity score (0–5).",
    "Your project description is quite brief. Can you tell me what problem you were solving, your technical approach, and what the outcome was?",
    "Walk me through this project in detail — problem statement, your architecture choices, the challenges, and the final result.",
    "CRITICAL")

row("EXP","E10 — Project 1\n(8 pts max)","Project Duration\n(how long it ran)",
    "Look for: Duration stated — '6-month project', 'Jan 2022 – Jun 2022', 'over 1 year'\n"
    "MISSING for most — candidates rarely state duration",
    "Criterion 4: Duration >=3 months = +1pt.",
    "How long did this project run, approximately? And how long were you personally involved in it?",
    "—",
    "CRITICAL")

row("EXP","E10 — Project 1\n(8 pts max)","Skills / Tech Stack Used in Project",
    "Look for: Technology names in the project description\n"
    "e.g.  'Built using Python, Apache Spark, AWS S3, and scikit-learn'\n"
    "MISSING if: Description talks about the project without naming specific tools",
    "Criterion 5: >=1 skill listed = +1pt.\nAlso feeds skill depth — skills used in projects get APPLIED evidence level.",
    "What specific technologies, tools, and frameworks did you use in this project?",
    "Beyond what's listed, were you using any cloud services, databases, or ML libraries?",
    "CRITICAL")

row("EXP","E10 — Project 1\n(8 pts max)","Business Domain\n(which industry this was for)",
    "Look for: Industry keywords in project description\n"
    "e.g.  'fraud detection for an NBFC', 'demand forecasting for e-commerce', 'patient risk scoring'\n"
    "MISSING if: Description is purely technical — 'built a Spark pipeline' without saying what it was for",
    "Criterion 6: Domain tag present = +1pt.",
    "What was the business context of this project — which industry or domain was it for, and who were the end users?",
    "Who were the stakeholders for this project and what business decision did your work support?",
    "IMPORTANT")

row("EXP","E10 — Project 1\n(8 pts max)","Role Played / Ownership\n(what YOU specifically did)",
    "Look for: OWNERSHIP VERBS: 'I built', 'I designed', 'I led', 'I architected', 'I deployed'\n"
    "NOT ENOUGH: 'We built', 'The team developed', 'Was part of a team that...'\n"
    "Description must be >50 characters AND contain at least one ownership verb\n"
    "MISSING if: Description uses 'we' throughout or is vague about individual role",
    "Criterion 7: Ownership verb + description >50 chars = +1pt.",
    "In this project, what was YOUR specific role — did you design the architecture, build specific components, lead the team, or were you one of many developers?",
    "Tell me exactly which parts of this system you personally built or designed. How many developers were on the team and what did the others do?",
    "CRITICAL")

row("EXP","E10 — Project 1\n(8 pts max)","Quantified Business Impact\n(numbers / metrics)",
    "Look for: NUMBERS + OUTCOME:\n"
    "- Percentage: 'reduced latency by 40%', 'improved accuracy to 92%'\n"
    "- Money: 'saved $2M/year', 'generated ₹5Cr revenue'\n"
    "- Scale: 'processed 10M records/day', 'served 2M users'\n"
    "- Comparison: 'reduced processing time from 4 hours to 45 minutes'\n"
    "MISSING: 'improved performance' (no number), 'helped the business' (no metric)",
    "Criterion 8: Quantified impact = +1pt. WITHOUT this, max project score = 7/8.\nThis is the MOST COMMONLY MISSING criterion.",
    "What was the measurable outcome of this project? Can you give me a specific number — did it reduce costs by X%, increase speed by Y, or handle Z users?",
    "How did you know this project was successful? What metric moved, and by how much?",
    "CRITICAL")

row("EXP","E11 — Project 2\n(6 pts max)","All Project 2 fields\n(same as Project 1 minus last 2 criteria)",
    "Look for: Same as Project 1 but for the second-most-recent significant project.\n"
    "Only 6 criteria scored (no ownership verb or quantified impact criteria).\n"
    "MISSING if: Candidate only describes 1 project on resume",
    "Max 6pts. Criteria: title + type + description + duration + skills + domain.",
    "What was your second most significant project? Tell me the title, what you were building, for which domain, how long, and your role.",
    "Walk me through the technical decisions you made on your second project — what problem did you solve and how?",
    "CRITICAL")

# ═══════════════════════════════════════════════════════════════════════════
# EDUCATION
# ═══════════════════════════════════════════════════════════════════════════
section("EDU", "EDUCATION SECTION  |  15 pts total")

row("EDU","ED1 — Institute\n(up to 5 pts)","College / University Name",
    "Look for: Institution name in education section\n"
    "e.g.  'IIT Bombay', 'University of Delhi', 'VIT Vellore'\n"
    "PROBLEM: Abbreviations — 'IIT' without campus, 'BITS' without city\n"
    "PROBLEM: Typos or non-standard spelling",
    "TIER_1 (IIT/IIM/ISI/ISB/global top-200) = 4 base pts.\nTIER_2 (NIT/VIT/BITS etc.) = 3 base pts.\nTIER_3 = 2pts. TIER_4 = 1pt.",
    "Can you confirm the full name of your college — for example, is it IIT Bombay or IIT Madras? Which campus?",
    "—",
    "CRITICAL")

row("EDU","ED1 — Institute\n(up to 5 pts)","GPA / CGPA / Percentage",
    "Look for: Numeric grade in education section\n"
    "e.g.  '8.5/10', '75%', '3.9/4.0', '9.2 CGPA', 'First Class with Distinction'\n"
    "MISSING: Many candidates deliberately omit GPA if it's average",
    "GPA EXCELLENT (>=8.5/10 or >=85%) adds +1pt on TIER_1, +0.5pt on TIER_2.\nLOW GPA = -0.5pt penalty.",
    "What was your GPA or percentage in your degree? The resume doesn't mention it.",
    "—",
    "IMPORTANT")

row("EDU","ED1 — Institute\n(up to 5 pts)","Degree Stream / Specialisation",
    "Look for: What the degree was in\n"
    "e.g.  'B.Tech in Computer Science and Engineering', 'M.Sc Statistics', 'MBA Marketing'\n"
    "IT STREAMS (full credit): CS, CSE, IT, Software Engineering, Data Science, AI/ML, MCA, Statistics, Analytics, Quant Econ, Econometrics, Operations Research\n"
    "NON-IT (partial credit): Mechanical, Civil, Chemistry, Commerce, Arts",
    "IT stream = TECH fit = education_job_relevance = HIGH = 2pts.\nNon-IT = MEDIUM or FOUNDATIONAL.",
    "What was your specialisation within your degree — for example, Computer Science, Electronics, Statistics, or something else?",
    "—",
    "CRITICAL")

row("EDU","ED2 — Degree Level\n(up to 2 pts)","Degree Type",
    "Look for: Degree abbreviation or full name\n"
    "e.g.  'B.Tech', 'M.Tech', 'Ph.D', 'MBA', 'MCA', 'B.Sc', 'Diploma'\n"
    "MISSING if: Candidate just writes 'Engineering' without specifying level",
    "PhD/Master=2pts | Bachelor=1.5pts | Diploma=1pt | Unknown=0.5pts",
    "What level of degree is this — is it a bachelor's, master's, or PhD?",
    "—",
    "CRITICAL")

row("EDU","ED2 — Degree Level\n(up to 2 pts)","Multiple Degrees\n(if more than one)",
    "Look for: Two separate education entries at different levels\n"
    "e.g.  B.Tech + M.Tech listed separately\n"
    "IMPORTANT: Use BEST entry for scoring. Note if degrees are from different tier institutions.",
    "Best entry used for score. PhD at TIER_2 scores better than B.Tech at TIER_1.",
    "Do you have multiple degrees? Which do you consider your primary qualification for this role?",
    "—",
    "IMPORTANT")

row("EDU","ED3 — Education Gap\n(1 pt)","Education End Date to First Job Start Date",
    "Look for: Year education ended + year first job started\n"
    "e.g.  'B.Tech 2015–2019' and 'First job: Jan 2020' = 7-month gap (acceptable)\n"
    "MISSING if: Only years given — cannot calculate gap precisely",
    "<=6m gap=1pt | 6–12m=0.5pt | >12m=REJECT FLAG",
    "When exactly did you finish your degree — which month and year? And when did you start your first job?",
    "What were you doing in the gap between finishing your degree and starting your first job?",
    "CRITICAL")

row("EDU","ED4 — Relevance\n(2 pts)","Education-to-Job Domain Match",
    "Look for: Whether the degree subject matches the target role.\n"
    "HIGH match: CS degree → Software/Data role\n"
    "MEDIUM match: Science/Maths degree → Data role\n"
    "LOW match: Arts/Commerce degree → Tech role\n"
    "Domain switch: Engineering degree but working in a completely different field",
    "HIGH=2pts | MEDIUM=1.5pts | FOUNDATIONAL=0.5pts | UNKNOWN=1pt",
    "Your degree is in [Field] — how did your academic training prepare you for a career in [Target Role]?",
    "—",
    "IMPORTANT")

row("EDU","ED5 — Exec Education\n(1 pt)","Executive / Online Certification Courses\n(post-degree)",
    "Look for: Any additional learning after the primary degree\n"
    "e.g.  'IIM Executive Programme in Data Science', 'Coursera Deep Learning Specialisation', 'PG Diploma in AI/ML'\n"
    "MISSING if: Candidate did courses but didn't list them in education section",
    "Recognised exec programme or online cert = 1pt bonus.",
    "Have you done any executive education programmes, PG diplomas, or significant online certification programmes after your degree?",
    "—",
    "NICE")

row("EDU","ED6 — Patents / Publications\n(2 pts)","Patent or Publication Listed",
    "Look for: Patent details or paper citations in achievements or education section\n"
    "e.g.  'Filed patent for X (USPTO 2022)', 'Published paper in IEEE 2023'\n"
    "MISSING if: Patent/publication exists but candidate didn't list it",
    "Any patent or publication = 2pts.\nTIER_1 institution even without patent = 0.5pt base credit.",
    "Do you have any patents filed or granted, or research papers published in journals or conferences?",
    "Tell me about your patent/paper — what was the core idea and was it filed/published successfully?",
    "IMPORTANT")

row("EDU","ED7 — LinkedIn\n(1 pt — Recruiter fills)","LinkedIn Profile",
    "Look for: LinkedIn URL in resume header or contact section\n"
    "e.g.  'linkedin.com/in/johndoe'\n"
    "MISSING if: No LinkedIn URL on resume",
    "Active profile=1pt | Less active=0.5pt | Not present=0pt",
    "Can you share your LinkedIn profile URL? I'll check it during our call.",
    "—",
    "IMPORTANT")

row("EDU","ED8 — Extra-Curriculars\n(1 pt)","Activities / Clubs / Leadership",
    "Look for: Activities section at bottom of resume\n"
    "e.g.  'Placement Committee Coordinator', 'Hackathon winner', 'NSS Volunteer', 'Cricket team captain'\n"
    "MISSING if: No activities section, or candidate left it blank",
    "Boolean: Any activity = 1pt | None = 0pt.",
    "Were you involved in any clubs, societies, sports, or campus activities during college — especially any leadership roles?",
    "—",
    "NICE")

# ═══════════════════════════════════════════════════════════════════════════
# SKILLS
# ═══════════════════════════════════════════════════════════════════════════
section("SKL", "SKILLS SECTION  |  45 pts total")

row("SKL","S1 — Skill List / Years\n(6 pts max — Recruiter validates)","Skills Listed on Resume",
    "Look for: Explicit skills section listing tech skills\n"
    "e.g.  'Python, SQL, Apache Spark, TensorFlow, AWS, Tableau'\n"
    "ALSO look in: Role descriptions (skills used in context), Project descriptions\n"
    "MISSING if: No explicit skills section — skills must be inferred from descriptions",
    "Each skill validated as APPLIED+ with clear years of use = 1pt. Max 6pts.",
    "You've listed [Skill] — can you confirm approximately how many years you've been using it and when you last used it?",
    "—",
    "CRITICAL")

row("SKL","S1 — Skill List / Years\n(6 pts max — Recruiter validates)","Years of Use per Skill",
    "Look for: Duration next to each skill — '5 yrs Python', 'Python (2018–present)'\n"
    "RARELY on resume — most candidates just list skills without years\n"
    "MUST ASK recruiter for every key skill",
    "Only APPLIED+ skills with confirmed years count. Unconfirmed years = MENTION evidence only.",
    "For [each key skill] — when did you first start using it professionally, and when did you last use it? Are you currently using it?",
    "—",
    "CRITICAL")

row("SKL","S5 — Skill Depth\n(8 pts — Panel validates)","Depth of Each Skill\n(MENTION vs APPLIED vs EXPERT)",
    "Look for in resume:\n"
    "EXPERT signal: 'Architected X using [Skill]', 'designed the [Skill] pipeline', 'OSS contributor to [Skill] library'\n"
    "DEEP signal: Used in multiple complex projects, architecture decisions mentioned\n"
    "APPLIED signal: Used in specific projects with context\n"
    "WEAK signal: Mentioned in role description without context\n"
    "MENTION: Listed in skills section only — no usage evidence anywhere\n"
    "MISSING: Not mentioned at all",
    "(avg_blended_depth_score / 5) x 8 for top 5 skills.\nMENTION only = 0.5/5 per skill.",
    "You've listed [Skill] on your resume — can you describe a specific project where you used it? What did you build with it?",
    "Walk me through the most complex thing you've built using [Skill]. What design decisions did you make, and what trade-offs did you consider?",
    "CRITICAL")

row("SKL","S5 — Skill Depth\n(8 pts — Panel validates)","Coding Signal per Skill\n(actually wrote code vs just used it)",
    "Look for: Action verbs connected to the skill\n"
    "'built X using Python', 'developed Y pipeline in Spark', 'implemented Z model'\n"
    "NOT ENOUGH: 'worked with Python', 'familiar with Spark', 'exposure to AWS'\n"
    "MISSING if: Skills listed but no verbs showing actual coding",
    "Coding signal detected → evidence_level upgraded to APPLIED minimum.\nNo coding signal → capped at WEAK.",
    "When you used [Skill] — were you writing the code yourself, or more configuring/managing it at a higher level?",
    "Can you show me or describe the actual code you wrote using [Skill]?",
    "CRITICAL")

row("SKL","S5 — Skill Depth\n(8 pts — Panel validates)","Architecture Signal per Skill\n(designed systems using it)",
    "Look for: Design/architecture verbs with the skill:\n"
    "'designed the data pipeline architecture using Spark', 'architected the ML platform on AWS'\n"
    "MISSING if: Candidate implemented but didn't design",
    "Architecture signal → evidence_level upgraded to DEEP or higher.",
    "Did you design the architecture for the system that used [Skill], or were you implementing someone else's design?",
    "Walk me through the system architecture for a project where you used [Skill] — what decisions did you make and why?",
    "IMPORTANT")

row("SKL","S6 — Skill Recency\n(6 pts)","Last Used Date per Skill",
    "Look for: End date of roles where skill appears\n"
    "RECENT: Skill appears in current role or role ended within last 2 years\n"
    "MID: Last used 2–5 years ago\n"
    "OLD: Last used >5 years ago\n"
    "MISSING if: Dates not given or skill only in skills section with no role context",
    "(count_RECENT_skills / total_skills) x 6.\nAll MENTION-only skills = OLD by default.",
    "When did you last actively use [Skill] in a real project? Is it something you use currently in your day-to-day work?",
    "Are you currently using [Skill] in your role? If not, when was the last time you used it and in what context?",
    "CRITICAL")

row("SKL","S4 — Certifications\n(3 pts max)","Certifications Listed",
    "Look for: Certifications section or listed under education\n"
    "e.g.  'AWS Certified Solutions Architect – Professional (2022)', 'Google Professional Data Engineer'\n"
    "MISSING if: Candidate has certs but didn't list them, or listed without year/validity",
    "1 valid cert = 1pt. Max 3pts.\nExpired certs = 0pts. Irrelevant certs = 0pts.",
    "Do you have any certifications? Can you confirm which ones are currently valid and relevant to this role?",
    "—",
    "IMPORTANT")

row("SKL","S4 — Certifications\n(3 pts max)","Certification Validity",
    "Look for: Year obtained + expiry (if mentioned)\n"
    "AWS/GCP/Azure certs expire after 3 years — need year to check\n"
    "MISSING if: Year not listed next to certification",
    "Expired cert = 0pts for that cert.",
    "When did you obtain your [Certification] and is it still currently valid?",
    "—",
    "IMPORTANT")

row("SKL","S7 — Learning Acumen\n(3 pts)","New Skills Acquired Year-on-Year",
    "Look for: Skills that appear only in recent roles (not in earlier roles)\n"
    "e.g.  LangChain appears only from 2023, not in 2020 roles = new skill picked up\n"
    "MISSING if: Skills section is flat — same skills listed without any timeline context",
    "Fast learner (>=2 new skills/yr for >=2 yrs) = 3pts.\nNew skills in >=3 yrs = 2pts.",
    "In the last 2–3 years, what new technologies have you learned and actually applied in your work — not just done a tutorial on?",
    "How do you typically stay current with new developments? Give me an example of something you learned in the last 6 months and used in a project.",
    "IMPORTANT")

row("SKL","S8 — Coding Community\n(3 pts)","Online Platform Presence",
    "Look for: Profile URLs or platform names in resume header or footer\n"
    "e.g.  'github.com/username', 'kaggle.com/username', 'leetcode.com/username'\n"
    "MISSING for most — candidates don't add community links",
    ">=3 platforms = 3pts | 2=2pts | 1=1pt | 0=0pts",
    "Do you have a GitHub profile? Can you share the URL? Are you active on Kaggle, LeetCode, or Stack Overflow?",
    "Walk me through your most significant GitHub repository — what problem does it solve and how many people use it?",
    "IMPORTANT")

row("SKL","S9 — Communication\n(5 pts — Panel only)","Communication Skills",
    "Look for: Cannot be assessed from resume.\n"
    "Proxy signals in resume: Well-structured descriptions, concise language, clear project narratives\n"
    "POOR proxy signal: Bullet points that are just tool lists, grammatical errors",
    "5pts — entirely Panel-assessed. No resume auto-score.",
    "N/A — Recruiter gives a rough impression from the call",
    "Explain [a technical concept from their work] as if I'm a business stakeholder with no technical background.\nThen: How would you explain it differently to a junior engineer?",
    "CRITICAL")

row("SKL","S10 — Domain Skills\n(5 pts — Panel only)","Domain Depth",
    "Look for: Cannot be fully assessed from resume.\n"
    "Partial signals: Domain-specific terminology used correctly in descriptions\n"
    "e.g.  'calculated LTV for NBFC portfolio', 'built credit scorecard using logistic regression' = BFSI depth signal\n"
    "MISSING: Generic descriptions like 'built ML models' without domain context",
    "5pts — Panel-assessed via scenario questions in candidate's claimed domain.",
    "N/A — assess from domain terminology used during call",
    "BFSI: Walk me through how you would build a credit risk model — what data, what approach, what regulatory constraints?\nHEALTHCARE: How would you approach patient readmission prediction — what features, what validation strategy?\n(Tailor question to candidate's domain)",
    "CRITICAL")

row("SKL","S11 — Project Explanation\n(3 pts — Recruiter + Panel)","Project Walk-through Quality",
    "Look for: Richness of project descriptions on resume — can you reconstruct the project from the text?\n"
    "GOOD: Problem + Approach + Role + Outcome described\n"
    "POOR: 'Worked on ML projects', 'Used Python and SQL'\n"
    "MISSING: No project descriptions at all",
    "3pts: 0=cannot explain | 1=disjointed | 2=good structure | 3=clear P-to-Outcome",
    "Walk me through your most recent project — what was the business problem, your technical approach, your specific role, and what did it achieve?",
    "You mentioned [project]. Walk me through the technical architecture — how did the different components connect and what was YOUR specific contribution to each piece?",
    "CRITICAL")

row("SKL","S12 — Coding Skills\n(Qualitative — Panel only)","Coding Ability",
    "Look for: Signals of actual coding in resume:\n"
    "- Specific libraries, frameworks, and code-level tools mentioned\n"
    "- GitHub links\n"
    "- 'Wrote', 'coded', 'scripted', 'implemented' verbs\n"
    "MISSING: Vague terms like 'worked with Python' — no indication of actual code",
    "Qualitative only — no numeric score. Panel narrative added to record.",
    "When you use Python/SQL — are you writing code yourself from scratch, or using notebooks and scripts written by others?",
    "Write a Python function that [relevant problem for their role level]. Walk me through your approach, edge cases, and the time complexity.",
    "CRITICAL")

row("SKL","S14 — Problem Solving\n(3 pts — Panel only)","Problem Solving",
    "Look for: Cannot be assessed from resume.\n"
    "Weak proxy: 'Solved X problem by Y approach' language in descriptions shows structured thinking",
    "3pts — entirely Panel-assessed via live scenarios.",
    "N/A",
    "Here's a scenario: [Relevant open-ended problem for their domain]. Walk me through how you would approach this. Take your time.",
    "CRITICAL")

# ─────────────────────────────────────────────────────────────────────────
# WRITE ROWS
# ─────────────────────────────────────────────────────────────────────────
SEC_CFG = {
    "EXP": ("1B5E20", "E8F5E9", "DCEDC8"),
    "EDU": ("5D4037", "FBE9E7", "FFCCBC"),
    "SKL": ("4A148C", "F3E5F5", "E1BEE7"),
}
PRIO_BG2 = {"CRITICAL": "FFCDD2", "IMPORTANT": "FFF9C4", "NICE": "F1F8E9"}

row_n = 6
alt = False
for record in ROWS:
    sec = record[0]

    if sec == "SEC":
        _, sec_key, label = record[0], record[1], record[2]
        cfg = SEC_CFG.get(sec_key, ("263238","ECEFF1","CFD8DC"))
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=9)
        cl = ws.cell(row=row_n, column=1, value=f"  {label}")
        cl.fill = fill(cfg[0])
        cl.font = f(bold=True, color="FFFFFF", size=10)
        cl.alignment = al("left", False)
        cl.border = BM
        ws.row_dimensions[row_n].height = 18
        row_n += 1
        alt = False
        continue

    (sec, param, dp, resume_look, score_imp, rq, pq, prio) = record
    cfg = SEC_CFG.get(sec, ("263238","ECEFF1","CFD8DC"))
    bg = cfg[2] if alt else cfg[1]
    alt = not alt

    ws.row_dimensions[row_n].height = 90

    c(row_n, 1, sec, bg, bold=True, ha="center", wrap=False)
    c(row_n, 2, param, bg, bold=True, size=8)
    c(row_n, 3, dp, bg, bold=True)
    c(row_n, 4, resume_look, "FFFFFF")
    c(row_n, 5, score_imp, "FFFDE7", size=8)

    # Column F — blank input cell
    cl = ws.cell(row=row_n, column=6, value="")
    cl.fill = fill("FFFFFF")
    cl.font = f(bold=True, size=12)
    cl.alignment = al("center", False)
    cl.border = Border(
        left=Side(style="medium", color="FF8F00"),
        right=Side(style="medium", color="FF8F00"),
        top=Side(style="medium", color="FF8F00"),
        bottom=Side(style="medium", color="FF8F00"))

    c(row_n, 7, rq, "E3F2FD" if rq != "N/A" else "F5F5F5")
    c(row_n, 8, pq, "FCE4EC" if pq not in ("—","N/A","") else "F5F5F5")
    c(row_n, 9, prio, PRIO_BG2.get(prio, "FFFFFF"),
      bold=(prio == "CRITICAL"), ha="center", wrap=False, size=8)

    row_n += 1

# widths
for i, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for i in range(1, 5):
    ws.row_dimensions[i].height = 22

out = "E:/Dev/resume_intelligence/Resume_DataPoint_Checklist.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Data rows: {row_n - 6}")
