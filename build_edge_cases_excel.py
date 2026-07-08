"""Build Edge_Cases_Signals.xlsx — archetypes, edge cases, red flags with datapoints."""
from __future__ import annotations
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Palette ──────────────────────────────────────────────────────────────────
C_HEADER_DARK  = "353395"   # primary indigo
C_HEADER_MED   = "6366F1"   # medium indigo
C_HEADER_LIGHT = "EEF0FF"   # light indigo tint
C_ARCHETYPE    = "0F4C81"   # dark blue for archetype sheet
C_ARCH_LIGHT   = "DBEAFE"   # light blue tint
C_EDGE_HDR     = "065F46"   # dark green for edge cases
C_EDGE_LIGHT   = "D1FAE5"   # light green
C_RED_HDR      = "991B1B"   # dark red for red flags
C_RED_LIGHT    = "FEE2E2"   # light red
C_SOFT         = "FEF3C7"   # amber for soft flags
C_HARD         = "FEE2E2"   # red for hard flags
C_WHITE        = "FFFFFF"
C_ROW_ALT      = "F8FAFC"
C_MUTED        = "62748E"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=10, color="262626", name="Calibri"):
    return Font(bold=True, size=size, color=color, name=name)

def reg_font(size=10, color="262626", name="Calibri"):
    return Font(bold=False, size=size, color=color, name=name)

def thin_border():
    s = Side(style="thin", color="CAD5E2")
    return Border(left=s, right=s, top=s, bottom=s)

def wrap_align(h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def write_header_row(ws, row, cols, bg, fg="FFFFFF", size=10):
    for c_idx, (label, width) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c_idx, value=label)
        cell.fill = fill(bg)
        cell.font = Font(bold=True, size=size, color=fg, name="Calibri")
        cell.alignment = wrap_align("center", "center")
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(c_idx)].width = width

def write_data_row(ws, row, values, bg=C_WHITE, font_color="262626", bold_cols=None):
    bold_cols = bold_cols or []
    for c_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c_idx, value=val)
        cell.fill = fill(bg)
        cell.font = Font(bold=(c_idx in bold_cols), size=10, color=font_color, name="Calibri")
        cell.alignment = wrap_align()
        cell.border = thin_border()

def freeze_and_filter(ws, pane="A2"):
    ws.freeze_panes = pane
    ws.auto_filter.ref = ws.dimensions

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 — Archetypes (A1-A10)
# ─────────────────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Archetypes (A1-A10)"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 32

ARCH_COLS = [
    ("Code",          7),
    ("Archetype Name",22),
    ("Trigger Condition",34),
    ("Key Datapoints / Signals",38),
    ("Edu Weight",    10),
    ("Exp Weight",    10),
    ("Skills Weight", 11),
    ("Total",          7),
    ("Scoring Intent / Why this reallocation",40),
    ("Priority Order",12),
]

write_header_row(ws1, 1, ARCH_COLS, C_ARCHETYPE, "FFFFFF", 10)
ws1.row_dimensions[1].height = 32

ARCHETYPES = [
    (
        "A1","Baseline","Default — no special-case match",
        "YoE: 2-9 yrs\nEdu tier: TIER_2/3\nCompany tier: 2-4\nNo founder/PhD/domain-switch signals",
        15, 40, 45, 100,
        "Standard weighting — no reallocation needed. Education, experience, and skills all contribute normally.",
        "10 (last — fallback)",
    ),
    (
        "A2","Weak Edu + FAANG","Edu TIER_4/5/UNKNOWN and best company = TIER_1",
        "highest_institute_tier: TIER_4/5/UNKNOWN\nbest_company_tier: 1 (Google, Meta, Amazon, etc.)\nYoE: 2+ yrs",
        5, 52, 43, 100,
        "Weak college is irrelevant when the candidate has been hired and retained by a FAANG company. Shift weight to experience and skills; nearly zero weight to education.",
        "7",
    ),
    (
        "A3","Weak Edu + Mid-Tier","Edu TIER_4/5/UNKNOWN and best company = TIER_2",
        "highest_institute_tier: TIER_4/5/UNKNOWN\nbest_company_tier: 2 (Flipkart, Razorpay, Zerodha, etc.)\nYoE: 2+ yrs",
        8, 47, 45, 100,
        "Similar to A2 but slightly less extreme. Mid-tier product company validates the candidate beyond edu pedigree. Edu weight reduced but not zeroed.",
        "8",
    ),
    (
        "A4","Elite Edu + Weak Company","Edu TIER_1 and best company TIER_3+",
        "highest_institute_tier: TIER_1 (IIT/IIM/NIT/equiv)\nbest_company_tier: 3+ (no product/FAANG experience)\nYoE: 3+ yrs",
        20, 37, 43, 100,
        "IIT/IIM pedigree signals intellectual capability even if career hasn't reached top-tier companies yet. Upweight education; downweight experience slightly to avoid penalising the pedigree candidate.",
        "9",
    ),
    (
        "A5","Fresh Graduate","YoE ≤ 1 yr OR single job with YoE ≤ 2 yrs",
        "total_experience_years: 0-1\nOR (n_jobs=1 AND yoe ≤ 2)\neducation_entries: active\nNo significant work history",
        30, 10, 60, 100,
        "For fresh grads, education pedigree and skills portfolio matter far more than work history. Heavily upweight education + skills; experience weight is minimal.",
        "1 (highest priority)",
    ),
    (
        "A6","Senior 10+ YoE","total_experience_years ≥ 10",
        "total_experience_years: 10+\nMultiple roles/companies\nSeniority signals in titles",
        8, 47, 45, 100,
        "For senior candidates, experience quality and depth dominate. Education becomes nearly irrelevant for a 10+ year career. Upweight experience.",
        "5",
    ),
    (
        "A7","PhD / Research Track","PhD degree detected OR 2+ publications listed",
        "degree_level_tag: contains 'phd'/'doctoral'\nOR publications list: len ≥ 2\neducation_entries: PhD institution",
        25, 30, 45, 100,
        "Research candidates are defined by their academic rigor, theoretical foundation, and published output. Education is a first-class signal. Experience weight is reduced (research roles are structured differently).",
        "2",
    ),
    (
        "A8","Domain Switcher","education_job_relevance: FOUNDATIONAL/UNKNOWN_DOMAIN AND YoE ≥ 3",
        "education_job_relevance: FOUNDATIONAL or UNKNOWN_DOMAIN\ntotal_experience_years: 3+\nSkill cluster diverges from degree field",
        5, 45, 50, 100,
        "When education is unrelated to the role, its signal value is near-zero. Skills portfolio is the primary evidence of capability for a domain switcher. Upweight skills significantly.",
        "4",
    ),
    (
        "A9","Founder / Serial Entrepreneur","2+ Founder/CEO title entries detected",
        "founder_title_count: ≥ 2\nTitles contain: founder/co-founder/cofounder\nMultiple company entries",
        10, 42, 48, 100,
        "Founders are often self-taught and domain-jumpers. Education matters less than what they built. Experience (companies built, exits) and skills (breadth) are the key signals.",
        "3",
    ),
    (
        "A10","Consultant / Contractor","3+ stints < 18m AND consultant/contractor keyword in titles",
        "short_stints_18m: ≥ 3\nn_roles: ≥ 3\nTitles contain: consultant/contractor/freelance/advisor",
        12, 40, 48, 100,
        "Consultants have many short engagements by design — don't penalise for 'job-hopping'. Skills breadth and project variety are their value proposition. Slight upweight on skills.",
        "6",
    ),
]

for i, row_data in enumerate(ARCHETYPES):
    r = i + 2
    bg = C_ARCH_LIGHT if i % 2 == 0 else C_WHITE
    ws1.row_dimensions[r].height = 68
    for c_idx, val in enumerate(row_data, 1):
        cell = ws1.cell(row=r, column=c_idx, value=val)
        cell.fill = fill(bg)
        cell.border = thin_border()
        cell.alignment = wrap_align()
        if c_idx == 1:
            cell.font = Font(bold=True, size=11, color=C_ARCHETYPE, name="Calibri")
        elif c_idx in (5, 6, 7, 8):
            cell.font = Font(bold=True, size=10, color="262626", name="Calibri")
            cell.alignment = wrap_align("center", "center")
        else:
            cell.font = reg_font()

freeze_and_filter(ws1)

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 — Edge Cases (E1-E20)
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Edge Cases (E1-E20)")
ws2.sheet_view.showGridLines = False

EDGE_COLS = [
    ("Code",       6),
    ("Edge Case Name",           22),
    ("Scenario Description",     34),
    ("Trigger Datapoints",       32),
    ("Trigger Signal Condition", 30),
    ("Current Behaviour (before fix)", 28),
    ("Correct Behaviour (after fix)",  28),
    ("Scoring Impact",           18),
    ("Implementation Status",    16),
    ("Rubric Section Affected",  18),
]

write_header_row(ws2, 1, EDGE_COLS, C_EDGE_HDR, "FFFFFF", 10)
ws2.row_dimensions[1].height = 32

EDGE_CASES = [
    (
        "E1","IIT → Tier-5 career path",
        "TIER_1 grad (IIT/NIT/IIM) who joins only IT services / unknown companies post-graduation.",
        "highest_institute_tier = TIER_1\nbest_company_tier = 4 or 5",
        "edu_tier_num == 1 AND best_tier >= 3",
        "Education score high but experience/company score drags total down severely.",
        "Archetype A4 fires: edu weight boosted to 20, exp weight reduced to 37. Candidate scored on pedigree signal.",
        "Edu: +5 pts | Exp: -3 pts",
        "IMPLEMENTED (A4 archetype)",
        "Education + Company Tier",
    ),
    (
        "E2","Self-taught developer, no degree",
        "No formal degree or degree from unknown institution, but strong GitHub + skills portfolio.",
        "edu_tier = UNKNOWN/TIER_5\noss_signal_count > 0\nstrong_rows count high",
        "edu_tier_num == 5 AND oss_signals >= 1",
        "Education score collapses the total even if skills are excellent.",
        "Archetype A2/A3 fires (if FAANG) or A8 (domain switch). Education weight dropped to 5-8 pts.",
        "Edu: -7 to -10 pts | Skills: +3-5 pts",
        "IMPLEMENTED (A2/A3 via archetype)",
        "Education",
    ),
    (
        "E3","FAANG hopper (Google → Amazon → Meta)",
        "Candidate moves between TIER_1 companies every 2-3 years. Pattern looks like job-hopping on naive parse.",
        "all companies in tenure_with_dates have tier = 1\nshort_stints count > 0",
        "ALL _pretier_map values == 1",
        "Stability penalised for short stints despite all moves being TIER_1 → TIER_1.",
        "Short stints zeroed, hop_rate zeroed. +0.5 stability pts. Note: 'E3: FAANG→FAANG trajectory is industry-normal.'",
        "Stability: +0.5 pts",
        "IMPLEMENTED (stability section)",
        "Experience > Stability",
    ),
    (
        "E4","Career break = MBA / further study",
        "Gap between jobs exactly overlaps an education entry (e.g. quit job to do MBA for 18m).",
        "gap_start / gap_end overlaps\neducation entry start_date / end_date\n≥50% temporal overlap",
        "_is_edu_overlap(gap, edu_periods) == True",
        "Gap counted as hard career break, penalising stability and career_breaks score.",
        "Gap moved to edu_breaks list. Not counted in break_count. Reason note: 'E4: educational break verified.'",
        "Career Breaks: +0.5 to +2 pts",
        "IMPLEMENTED (_detect_career_breaks)",
        "Experience > Career Breaks",
    ),
    (
        "E5","Maternity / paternity leave gap",
        "Gap of 4-18 months between consecutive jobs, no education overlap, no hostile context.",
        "gap_months: 3 < gap ≤ 18\nno education overlap\nnot classified as E4",
        "gap_months > 3 AND gap_months <= 18 AND NOT edu_overlap",
        "Gap penalised as short/moderate career break (1.0-1.5 pts deducted from full 2.0).",
        "Classified as possible_parental_break. Not counted in break_count. Soft note added. Full 2.0 pts retained.",
        "Career Breaks: +0.5 to +1.5 pts",
        "IMPLEMENTED (_detect_career_breaks)",
        "Experience > Career Breaks",
    ),
    (
        "E6","BERT career progression override",
        "BERT model classifies career as FAST_TRACK/GROWING but heuristic trajectory score is low.",
        "career_progression_prior.label = FAST_TRACK or GROWING\ncareer_progression_prior.confidence ≥ 0.60",
        "bert_cp_pts is not None AND cp_confidence >= 0.60",
        "Heuristic trajectory only. May miss nuance in title patterns.",
        "BERT blended at 60%: prog_pts = bert_pts*0.60 + heuristic*0.40 when confidence ≥ 0.60.",
        "Career Progression: ±0.5 to ±1.5 pts",
        "IMPLEMENTED (progression section)",
        "Experience > Career Progression",
    ),
    (
        "E7","Startup shutdown / involuntary exit",
        "Short stint (<18m) at a TIER_4/5 company with no gap after (immediately joined next role).",
        "company_tier ≥ 4 (TIER_4/5)\nstint_months < 18\ngap_after_months ≤ 3",
        "coy_tier >= 4 AND stint_m < 18 AND gap_after <= 3",
        "Startup exit counted as job-hopping, reducing stability score.",
        "Startup exits subtracted from adjusted_short_stints count. +0.5 stability pts if any exits detected.",
        "Stability: +0.5 pts per exit",
        "IMPLEMENTED (E7 in stability section)",
        "Experience > Stability",
    ),
    (
        "E8","Cert farming (shallow certifications)",
        "Candidate lists 5+ certifications but has no APPLIED/DEEP/EXPERT skill evidence and BERT depth ≤ FOUNDATIONAL.",
        "cert_count ≥ 5\nstrong_rows count == 0\nbert_avg_depth ≤ FOUNDATIONAL (score ≤ 1.5)",
        "cert_count >= 5 AND len(strong_rows) == 0 AND avg_bert_depth <= 1.5",
        "Certification score inflated to 5/5 pts without depth evidence.",
        "Cert score capped at 2/5 pts. R6 soft flag added: 'Cert farming detected.'",
        "Skills > Certifications: -3 pts cap",
        "IMPLEMENTED (compute_rubric_score cert guard)",
        "Skills > Certifications",
    ),
    (
        "E9","Consulting-style project titles",
        "Projects use consulting terminology (Engagement, Workstream, Deliverable) but system scores as product projects.",
        "project type detected as PRODUCT by rules\nllm_confirmed_type = CONSULTING",
        "llm_confirmed_type != rule_detected project_type",
        "Wrong project type, potentially wrong complexity score.",
        "LLM confirmed type overrides rule type in project reason string. Type note shows override clearly.",
        "Project score: ±0-1 pt",
        "IMPLEMENTED (LLM deep judgment)",
        "Experience > Project 1/2",
    ),
    (
        "E10","Multiple degrees (MBA after BTech)",
        "Candidate has both BTech (TIER_3) and MBA (TIER_1 IIM). System may only see one.",
        "education_entries count ≥ 2\nhighest_institute_tier = best across all entries",
        "max(tier across all education_entries) used",
        "Lower-tier first degree dominates if highest_tier not computed correctly.",
        "education_engine uses highest tier across all entries. IIM MBA correctly upgrades to TIER_1.",
        "Education: +2 to +3 pts",
        "IMPLEMENTED (education_engine)",
        "Education > Institute Tier",
    ),
    (
        "E11","Part-time / contract role listed as full-time",
        "Short 3-month stints that are actually consulting contracts, not permanent roles.",
        "stint_months < 4\ncompany_tier = 4/5\ntitle contains consultant/contract",
        "is_consultant_pattern == True (used for A10)",
        "Counted as instability / short stints.",
        "Archetype A10 fires. Consultant pattern context applied. Stints not penalised.",
        "Stability: +0.5 pts",
        "IMPLEMENTED (A10 archetype + E7)",
        "Experience > Stability",
    ),
    (
        "E12","PhD with no industry experience",
        "Candidate has PhD and 2+ publications but 0-2 years of work experience.",
        "is_phd = True\ntotal_experience_years ≤ 2",
        "is_phd AND yoe <= 2 → A7 (PhD takes priority over A5)",
        "May be scored as fresh grad with no archetype benefit.",
        "A7 fires (PhD priority over A5). Education weight = 25, exp weight = 30, skills = 45.",
        "Total: +5 to +8 pts vs naive scoring",
        "IMPLEMENTED (A7 > A5 in priority)",
        "Archetype Detection",
    ),
    (
        "E13","International candidate — degree not in DB",
        "Degree from foreign institution not in INSTITUTE_DICTIONARY. LLM search resolves it.",
        "institution_source = 'llm_search'\ntier resolved via AI",
        "llm_resolved = True in education engine output",
        "Institute defaults to TIER_4/5 (unknown). Education score collapses.",
        "LLM search resolves institute. Tier updated. Note '[tier via AI search]' shown in reason.",
        "Education: +2 to +4 pts",
        "IMPLEMENTED (education_engine LLM search)",
        "Education > Institute Tier",
    ),
    (
        "E14","Skills list in wrong section",
        "Candidate puts skills in 'Tools', 'Technologies', or 'Competencies' field instead of 'skills'.",
        "skills field empty\ntechnologies/tools/competencies fields non-empty",
        "Multiple source fields collected in _collect_skills()",
        "Skills not found; skill score = 0.",
        "engine.py _collect_skills() aggregates from 9 possible fields including nested structures.",
        "Skills: +10 to +20 pts",
        "IMPLEMENTED (engine.py _collect_skills)",
        "Skills",
    ),
    (
        "E15","Internal promotions — same company, multiple roles",
        "Candidate had 3 roles at Infosys over 6 years. Each role looks like a separate short job.",
        "tenure_with_dates: consecutive entries same company name\ninternal_promos count ≥ 1",
        "sorted_twd[j].company.lower() == sorted_twd[j+1].company.lower()",
        "Each promotion counted as a separate short stint; stability heavily penalised.",
        "Internal promos subtracted from adjusted_short_stints. +0.5 stability pts. Note: 'E15: internal promotions detected.'",
        "Stability: +0.5 to +1.5 pts",
        "IMPLEMENTED (E15 in stability section)",
        "Experience > Stability",
    ),
    (
        "E16","Tier-5 company that is well-funded startup",
        "Company name not in TIER_MAP but is actually a well-known funded startup.",
        "company not found in TIER_MAP\nllm_fallback = False → tier = 5",
        "classify_company_tier(name, llm_fallback=False) → 5",
        "Startup gets TIER_5 score (0/6 pts for company tier).",
        "LLM fallback (when enabled) or fuzzy matching resolves unknown companies. E7 also provides startup-exit context.",
        "Company Tier: situational",
        "PARTIAL (LLM fallback available; fuzzy match in tier taxonomy)",
        "Experience > Company Tier",
    ),
    (
        "E17","10+ year career, no certifications",
        "Senior candidate with 12+ years has zero certifications — not relevant for their level.",
        "total_experience_years ≥ 10\ncert_count = 0",
        "detect_archetype → A6 (yoe >= 10)",
        "5 pts lost on certifications for senior candidates who don't need them.",
        "A6 archetype fires. Skills weight stays 45 but exp weight boosted. Cert gap less impactful in relative total.",
        "Relative improvement via archetype weight",
        "IMPLEMENTED (A6 archetype)",
        "Archetype Detection",
    ),
    (
        "E18","1 job on resume, clearly a fresher",
        "Candidate has 1 job entry (1.5 yrs) — system might miscategorize if YoE > 1.",
        "n_jobs = 1\nyoe ≤ 2.0",
        "n_jobs <= 1 AND yoe <= 2.0 → A5",
        "Treated as mid-level with partial scores everywhere.",
        "A5 fires (n_jobs<=1 AND yoe<=2.0). Education + skills weighted up for fresh grad evaluation.",
        "Total: +5 to +10 pts in appropriate sections",
        "IMPLEMENTED (A5 trigger fix)",
        "Archetype Detection",
    ),
    (
        "E19","International exposure from resume language",
        "Resume mentions 'global team', 'cross-geo stakeholders', or lists foreign companies.",
        "experience.international_exposure = True\ncompanies contain non-India names",
        "intl = experience.get('international_exposure', False)",
        "International exposure shown as 0 pts (recruiter pending) with no resume context signal.",
        "Reason string notes: 'Resume signals: international exposure detected. Recruiter to validate.'",
        "Context note only (recruiter fills 0-2)",
        "IMPLEMENTED (international_exposure section)",
        "Experience > International Exposure",
    ),
    (
        "E20","COVID gap (2020-2022)",
        "Career break between March 2020 and Dec 2022 due to COVID layoffs.",
        "gap_start year = 2020/2021/2022\nbreak in breaks or possible_parental_breaks list",
        "gap_start[:4] in ('2020','2021','2022')",
        "Penalised as standard career break despite being force-majeure.",
        "R15 soft flag fires: 'Career gap aligns with COVID wave. Do not penalise automatically.' Recruiter informed.",
        "Soft flag note — no automatic score change",
        "IMPLEMENTED (R15 in detect_red_flags)",
        "Experience > Career Breaks",
    ),
]

for i, row_data in enumerate(EDGE_CASES):
    r = i + 2
    bg = C_EDGE_LIGHT if i % 2 == 0 else C_WHITE
    ws2.row_dimensions[r].height = 72
    for c_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c_idx, value=val)
        cell.fill = fill(bg)
        cell.border = thin_border()
        cell.alignment = wrap_align()
        if c_idx == 1:
            cell.font = Font(bold=True, size=11, color=C_EDGE_HDR, name="Calibri")
        elif c_idx == 9:
            is_impl = "IMPLEMENTED" in str(val)
            is_partial = "PARTIAL" in str(val)
            cell.font = Font(
                bold=True, size=10,
                color=("065F46" if is_impl else ("D97706" if is_partial else "991B1B")),
                name="Calibri",
            )
        else:
            cell.font = reg_font()

freeze_and_filter(ws2)

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 — Red Flags (R1-R20)
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Red Flags (R1-R20)")
ws3.sheet_view.showGridLines = False

RF_COLS = [
    ("Code",        6),
    ("Flag Name",  22),
    ("Severity",   10),
    ("Scenario",   34),
    ("Trigger Datapoints",        32),
    ("Trigger Signal Condition",  30),
    ("Recommended Action",        30),
    ("Auto-fires?", 12),
    ("Rubric / Section",          18),
]

write_header_row(ws3, 1, RF_COLS, C_RED_HDR, "FFFFFF", 10)
ws3.row_dimensions[1].height = 32

RED_FLAGS = [
    ("R1","3+ Career Breaks","HARD",
     "Candidate has more than 2 hard career breaks (>18m each, non-educational).",
     "break_count (hard, >18m, non-educational)\nedu_breaks excluded\npossible_parental excluded",
     "break_count > 2",
     "Auto-reject recommendation. Recruiter must have strong counter-evidence to override.",
     "Yes — auto reject flag",
     "Experience > Career Breaks"),
    ("R2","Downward Brand Trajectory","HARD",
     "Candidate started at TIER_1/2 company and ended at TIER_4/5 (last 3+ roles declining).",
     "companies list tier sequence\ntier_seq[0] <= 2\ntier_seq[-1] >= 4\ndelta >= 2",
     "tier_seq[0] <= 2 AND tier_seq[-1] >= 4 AND (tier_seq[-1]-tier_seq[0]) >= 2",
     "Probe reason in phone screen — could be layoff, lifestyle, or performance. Do not auto-reject.",
     "Yes — hard flag",
     "Experience > Company Tier"),
    ("R3","Title Inflation","HARD",
     "Candidate has Director/VP/Head title but shows zero team ownership or achievement signals.",
     "inflated_title_kws detected\nleadership_signal_score = 0\nachievements list empty",
     "inflated_title detected AND leadership_score == 0 AND len(achievements) == 0",
     "Probe org size, team count, and P&L ownership in phone screen. Request org chart.",
     "Yes — hard flag",
     "Experience > Project / Awards"),
    ("R4","Buzzword Resume, No Impact","HARD",
     "Resume has projects but no quantified metrics, no awards, no certifications.",
     "project_types list non-empty\nno numeric in project descriptions\nachievements = []\ncert_count = 0",
     "project_types exist AND no regex \\d+ in descriptions AND achievements=[] AND cert_count=0",
     "Ask for 1-2 specific business outcomes in phone screen. Request metrics from projects.",
     "Yes — hard flag",
     "Impact Score"),
    ("R5","Severe Overqualification","HARD",
     "Candidate has 2x+ the JD maximum years of experience.",
     "total_experience_years\nyoe_range.max from JD config\nyoe > yoe_max * 2",
     "jd_config set AND yoe > yoe_max * 2",
     "Discuss long-term intentions, salary band acceptance, and career stage. High early-exit risk.",
     "Yes (only when JD configured)",
     "Experience > Overall Experience"),
    ("R6","Cert Farming","HARD",
     "5+ certifications listed but skill depth is FOUNDATIONAL or below (BERT evidence).",
     "cert_count ≥ 5\nstrong_rows count = 0 (no APPLIED/DEEP/EXPERT)\nbert_avg_depth ≤ 1.5 (FOUNDATIONAL)",
     "cert_count >= 5 AND len(strong_rows) == 0 AND avg_bert_depth <= 1.5",
     "Run a 10-min live problem in technical screen. Certifications alone are insufficient evidence.",
     "Yes — hard + score cap",
     "Skills > Certifications"),
    ("R7","Role Family Mismatch","HARD",
     "BERT signals the candidate's role family is different from the JD's target role family.",
     "role_family_prior.label\nrole_family_prior.confidence ≥ 0.70\njd_config.role_family",
     "bert_rf_label != jd_family AND rf_confidence >= 0.70",
     "Verify role alignment in phone screen. Could be intentional career pivot — or wrong application.",
     "Yes (only when JD configured)",
     "Role Alignment"),
    ("R8","No Verifiable Output (5+ yr career)","HARD",
     "5+ year career with no projects, no awards, and no certifications.",
     "total_experience_years ≥ 5\nproject_types = []\nachievements = []\ncert_count = 0",
     "yoe >= 5 AND project_types=[] AND achievements=[] AND cert_count=0",
     "Strong reject signal. Candidate cannot demonstrate deliverables. Reject unless exceptional reference.",
     "Yes — hard flag",
     "Impact Score"),
    ("R9","Frequent Lateral Moves","SOFT",
     "BERT classifies career progression as LATERAL (no upward mobility across roles).",
     "career_progression_prior.label = LATERAL\ncareer_progression_prior.confidence any",
     "cp_label == 'LATERAL'",
     "Assess growth mindset, ambition, and reasons for lateral moves in phone screen.",
     "Yes — soft flag",
     "Experience > Career Progression"),
    ("R10","Job-Hopping Pattern","SOFT",
     "Hop rate >1.5 roles/year OR 2+ stints under 12 months.",
     "tenures list\nhop_rate = n_roles / (total_months/12)\nshort_stints_12 count",
     "hop_rate > 1.5 OR short_stints_12 >= 2",
     "Ask for context (layoffs, startups, contracts). E3/E7/E15 may reduce this if edge cases apply.",
     "Yes — soft flag",
     "Experience > Stability"),
    ("R11","Large Education Gap","SOFT",
     "More than 12 months between graduation and first job.",
     "education_gap_months from education engine\nor gap_to_first_job_months",
     "education_gap_months > 12",
     "Ask what candidate did during that period. May be legitimate (research, visa, personal).",
     "Yes — soft flag",
     "Education > Education Gap"),
    ("R12","Inflated Skills List","SOFT",
     "Candidate lists more than 30 skills — breadth likely too wide for real depth.",
     "total skills count (experience.skills / evidence_map keys)\nlen(_all_skills) > 30",
     "len(all_skills) > 30",
     "Focus technical screen on top 5 core skills. Ask to rank proficiency.",
     "Yes — soft flag",
     "Skills"),
    ("R13","Stagnation at Single Company","SOFT",
     "5+ years at one company with only 1-2 distinct title levels.",
     "unique_companies count = 1\nyoe ≥ 5\ntitle_set count ≤ 2",
     "len(unique_companies)==1 AND yoe>=5 AND len(title_set)<=2",
     "Probe internal scope: team size, P&L ownership, cross-functional work, external visibility.",
     "Yes — soft flag",
     "Experience > Stability"),
    ("R14","Domain Switch Without Bridge","SOFT",
     "Education is unrelated to role AND less than 30% of skills are APPLIED/DEEP/EXPERT.",
     "education_job_relevance: FOUNDATIONAL or NONE\nstrong_ratio = strong_rows / skill_rows < 0.30",
     "edu_relevance in ('FOUNDATIONAL','NONE') AND strong_ratio < 0.30",
     "Run deeper technical screen. Ask how they built domain skills without formal education base.",
     "Yes — soft flag",
     "Skills"),
    ("R15","COVID / Recession Gap","SOFT",
     "Career gap starts in 2020, 2021, or 2022 — aligns with COVID layoff wave.",
     "gap_start[:4] in ('2020','2021','2022')\nbreak in breaks or possible_parental_breaks",
     "gap_start year in {2020, 2021, 2022}",
     "Do NOT penalise automatically. Confirm context. Almost all 2020-2022 gaps are force-majeure.",
     "Yes — soft note only",
     "Experience > Career Breaks"),
    ("R16","Founder-Only History","SOFT",
     "All career entries are Founder or CEO roles — candidate has never been a team member.",
     "titles_raw all contain founder/co-founder/ceo\nlen(titles) ≥ 2",
     "ALL titles contain founder/ceo keyword",
     "Assess IC contribution ability and peer collaboration in phone screen. Cultural fit risk.",
     "Yes — soft flag",
     "Archetype / Career Progression"),
    ("R17","Elite College, Mediocre Career","SOFT",
     "TIER_1 education institution but best company across entire career is TIER_4/5.",
     "edu_tier = TIER_1\nbest_company_tier ≥ 4\nyoe ≥ 5",
     "edu_tier_str == 'TIER_1' AND best_tier_rf >= 4 AND yoe >= 5",
     "Investigate reason. Could be deliberate lifestyle choice, health, or underperformance. Neutral probe needed.",
     "Yes — soft flag",
     "Education vs Experience"),
    ("R18","Shallow Recent Skills","SOFT",
     "70%+ of applied skills have < 1.5 years of usage — all recently acquired.",
     "strong_rows with years_of_usage < 1.5\nratio: recent_only / strong_rows > 0.70",
     "len(recent_only) >= 3 AND len(recent_only)/len(strong_rows) > 0.70",
     "Probe depth of 2-3 core skills in technical screen. May be skills-washing (adding trending tech).",
     "Yes — soft flag",
     "Skills > Skill Recency"),
    ("R19","No OSS / Coding Community","SOFT",
     "No open-source signals and no coding platform links for a technical role.",
     "oss_signal_count = 0\ncoding platform links = 0\nrole_family in tech families",
     "oss_count==0 AND role_family in {DATA_ENGINEER, ML_ENGINEER, SOFTWARE_ENGINEER, ...}",
     "Ask for GitHub/Kaggle/HuggingFace profile in phone screen. Low signal for research/platform roles.",
     "Yes — soft (tech roles only)",
     "Skills > Coding Community"),
    ("R20","Templated Project Descriptions","SOFT",
     "Top 2 project descriptions have >65% text similarity — possible copy-paste.",
     "project_types[0].description\nproject_types[1].description\nSequenceMatcher similarity > 0.65",
     "SequenceMatcher(desc1, desc2).ratio() > 0.65",
     "Probe individual contribution and unique challenge for each project. Ask what they personally built.",
     "Yes — soft flag",
     "Experience > Project 1/2"),
]

for i, row_data in enumerate(RED_FLAGS):
    r = i + 2
    severity = row_data[2]
    bg = C_RED_LIGHT if (severity == "HARD" and i % 2 == 0) else (C_SOFT if (severity == "SOFT" and i % 2 == 0) else C_WHITE)
    ws3.row_dimensions[r].height = 70
    for c_idx, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c_idx, value=val)
        cell.fill = fill(bg)
        cell.border = thin_border()
        cell.alignment = wrap_align()
        if c_idx == 1:
            color = "991B1B" if severity == "HARD" else "D97706"
            cell.font = Font(bold=True, size=11, color=color, name="Calibri")
        elif c_idx == 3:
            color = "991B1B" if severity == "HARD" else "92400E"
            bg_pill = "FEE2E2" if severity == "HARD" else "FEF3C7"
            cell.font = Font(bold=True, size=10, color=color, name="Calibri")
            cell.alignment = wrap_align("center", "center")
        else:
            cell.font = reg_font()

freeze_and_filter(ws3)

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 4 — Signal / Datapoint Dictionary
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Signal Dictionary")
ws4.sheet_view.showGridLines = False

SD_COLS = [
    ("Signal / Datapoint",        26),
    ("Source",                    16),
    ("Field Path",                32),
    ("Possible Values",           26),
    ("Used By (Edge Cases / Flags)", 30),
    ("Scoring Section",           18),
    ("Notes",                     34),
]

write_header_row(ws4, 1, SD_COLS, C_HEADER_DARK, "FFFFFF", 10)
ws4.row_dimensions[1].height = 32

SIGNALS = [
    ("total_experience_years","experience_engine","experience.total_experience_years","float: 0.0 – 40.0","A5,A6,A7,A8,R5,R8,R13","Experience > Overall","Core YoE signal. Used in A5 (fresh grad) and A6 (senior) archetype triggers."),
    ("highest_institute_tier","education_engine","education.highest_institute_tier","TIER_1 / TIER_2 / TIER_3 / TIER_4 / TIER_5 / UNKNOWN","A2,A3,A4,E1,E10,R17","Education > Institute Tier","Best tier across all education entries. IIT/IIM/NIT = TIER_1."),
    ("education_job_relevance","education_engine","education.education_job_relevance","HIGH / MEDIUM / FOUNDATIONAL / UNKNOWN_DOMAIN","A8,R14","Skills / Archetype","Measures alignment between degree field and job domain. FOUNDATIONAL = domain switch."),
    ("education_entries","education_engine","education.education_entries[*]","list of dicts (degree, institution, start_date, end_date, gpa_band, ...)","E4,E10,E12,E13","Education","Used for career break cross-reference (E4) and multi-degree scenarios (E10)."),
    ("best_company_tier","company_tier_taxonomy","classify_company_tier(company_name)","int: 1 (FAANG) – 5 (unknown)","A2,A3,A4,E3,E7,R2,R17","Experience > Company Tier","Derived per company. Best (lowest) tier across all employers."),
    ("tenure_with_dates","experience_engine","experience.tenure_with_dates[*]","{company, start (YYYY-MM), end (YYYY-MM)}","E3,E4,E7,E15,R1,R10,R15","Experience > Career Breaks / Stability","Used for gap detection, E4 edu overlap, E7 startup exit, E15 internal promotions."),
    ("tenures","experience_engine","experience.tenures","list[int]: months per role","R10,R13","Experience > Stability","Parallel to companies list. Used for hop_rate and short_stints calculation."),
    ("companies","experience_engine","experience.companies","list[str]: company names","A2,A3,A4,E3,R2,R13,R17","Experience > Company Tier","Ordered by recency (most recent first)."),
    ("titles / job_titles","experience_engine","experience.titles","list[str]: role titles","A9,E3,R3,R16","Experience > Career Progression","Used for founder detection (A9), title inflation (R3), progression signal."),
    ("career_progression_prior","bert_signal_engine","bert_priors.career_progression_prior","{label: FAST_TRACK/GROWING/LATERAL/DECLINING, confidence: 0.0-1.0}","E6,R9","Experience > Career Progression","BERT-predicted career trajectory. Blended 60/40 with heuristic when conf ≥ 0.60."),
    ("skill_depth_priors","bert_signal_engine","bert_priors.skill_depth_priors[*]","{skill, predicted_depth_label: AWARENESS/.../ARCHITECT_LEVEL, confidence}","E8,R6","Skills > Skill Depth","BERT depth estimates per skill. avg_bert_depth used for cert farming guard (R6/E8)."),
    ("role_family_prior","bert_signal_engine","bert_priors.role_family_prior","{label: DATA_ENGINEER/ML_ENGINEER/etc., confidence}","R7,R19","Skills / Role Family","BERT-predicted role family. Used to detect JD mismatch (R7) and OSS signal gate (R19)."),
    ("stakeholder_prior","bert_signal_engine","bert_priors.stakeholder_prior","{label: NONE/INTERNAL/CLIENT_FACING/C_LEVEL, confidence}","stakeholder_management","Experience > Stakeholder Mgmt","BERT context for recruiter-stage stakeholder scoring."),
    ("mentorship_prior","bert_signal_engine","bert_priors.mentorship_prior","{label: NONE/IMPLIED/FORMAL/LEAD, confidence}","mentorship_signal","Experience > Mentorship","BERT context for recruiter-stage mentorship scoring."),
    ("project_types","experience_engine","experience.project_types[*]","{title, type, description, skills, ...}","R4,R20,E9","Experience > Project 1/2","LLM deep judgment enriches project scoring. Used for buzzword (R4) and template (R20) flags."),
    ("achievements","experience_engine","experience.achievements","list[str]: awards/recognitions","R3,R4,R8","Experience > Awards","Used for title inflation (R3) and no-deliverables (R4,R8) detection."),
    ("publications","experience_engine","experience.publications","list[str] or []","A7","Archetype A7","2+ publications triggers PhD/Research archetype regardless of degree."),
    ("cert_count","education_engine","len(education.certificates)","int: 0 – N","E8,R6","Skills > Certifications","Combined with skill depth for cert farming guard. Capped at 5 pts normally."),
    ("stability_score","experience_engine","experience.stability_score","float: 0.0 – 5.0","E3,E7,E15,R10,R13","Experience > Stability","Heuristic stability score. BERT career_progression blended on top."),
    ("leadership_signal_score","experience_engine","experience.leadership_signal_score","int: 0 – N","R3","Experience > Mentorship/Stability","Count of leadership language instances in resume. 0 = no lead signal."),
    ("international_exposure","experience_engine","experience.international_exposure","bool: True/False","E19","Experience > International Exposure","Resume signal only. Recruiter fills 0-2 pts after phone screen."),
    ("fast_learner","experience_engine","experience.fast_learner","bool: True/False","skills_learning_acumen","Skills > Learning Acumen","True if ≥2 new skills/year detected across ≥2 years."),
    ("oss_signal_count","evidence_map","evidence_row.open_source_signal","bool per skill","R19","Skills > Coding Community","Sum of open_source_signal=True across all skill evidence rows."),
    ("gpa_band","education_engine","education_entries[*].gpa_band","EXCELLENT / GOOD / ACCEPTABLE / LOW / UNKNOWN","Education > Institute Tier","Education","GPA bonus: +1 pt for TIER_1 with EXCELLENT/GOOD GPA; +0.5 for TIER_2 EXCELLENT."),
    ("education_gap_months","education_engine","education.education_gap_months","float: months","R11","Education > Education Gap","Gap between graduation and first job. > 12m triggers R11 soft flag."),
    ("founder_title_count","rubric_engine","detect_archetype (computed)","int: count of Founder/CEO entries","A9","Archetype A9","≥ 2 founder entries → serial entrepreneur archetype. Edu weight drops to 10."),
    ("short_stints_18m","rubric_engine","detect_archetype / stability (computed)","int: stints < 18m","A10,E7,E11","Experience > Stability / Archetype","≥3 short stints with consultant keyword → A10. Combined with tier for E7."),
    ("is_phd","rubric_engine","detect_archetype (computed)","bool","A7","Archetype A7","Detected from degree_level_tag containing 'phd'/'doctoral' in any education entry."),
    ("is_domain_switcher","rubric_engine","detect_archetype (computed)","bool","A8","Archetype A8","edu_relevance in (FOUNDATIONAL, UNKNOWN_DOMAIN) AND yoe >= 3."),
    ("hop_rate","rubric_engine","_score_experience_section (computed)","float: roles per year","R10","Experience > Stability","n_roles / (total_months / 12). > 1.5 triggers R10 soft flag."),
]

for i, row_data in enumerate(SIGNALS):
    r = i + 2
    bg = C_HEADER_LIGHT if i % 2 == 0 else C_WHITE
    ws4.row_dimensions[r].height = 52
    for c_idx, val in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=c_idx, value=val)
        cell.fill = fill(bg)
        cell.border = thin_border()
        cell.alignment = wrap_align()
        if c_idx == 1:
            cell.font = Font(bold=True, size=10, color=C_HEADER_DARK, name="Calibri")
        elif c_idx == 3:
            cell.font = Font(italic=True, size=9, color="353395", name="Courier New")
        else:
            cell.font = reg_font()

freeze_and_filter(ws4)

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 5 — Coverage Matrix (Edge Cases × Signals)
# ─────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("Coverage Matrix")
ws5.sheet_view.showGridLines = False

SIGNAL_NAMES = [
    "total_experience_years","highest_institute_tier","education_job_relevance",
    "best_company_tier","tenure_with_dates","tenures","career_progression_prior",
    "skill_depth_priors","role_family_prior","project_types","achievements",
    "publications","cert_count","stability_score","leadership_signal_score",
    "founder_title_count","is_phd","is_domain_switcher","hop_rate",
]

EC_NAMES = [f"E{i}" for i in range(1, 21)] + [f"R{i}" for i in range(1, 21)]

EC_SIGNAL_MAP: dict[str, list[str]] = {
    "E1":  ["highest_institute_tier","best_company_tier"],
    "E2":  ["highest_institute_tier","skill_depth_priors"],
    "E3":  ["best_company_tier","tenure_with_dates","stability_score"],
    "E4":  ["tenure_with_dates","education_entries"],
    "E5":  ["tenure_with_dates"],
    "E6":  ["career_progression_prior"],
    "E7":  ["tenure_with_dates","best_company_tier","stability_score"],
    "E8":  ["cert_count","skill_depth_priors"],
    "E9":  ["project_types"],
    "E10": ["highest_institute_tier"],
    "E11": ["tenure_with_dates","best_company_tier","stability_score"],
    "E12": ["is_phd","total_experience_years","publications"],
    "E13": ["highest_institute_tier"],
    "E14": [],
    "E15": ["tenure_with_dates","stability_score"],
    "E16": ["best_company_tier"],
    "E17": ["total_experience_years","cert_count"],
    "E18": ["total_experience_years","tenure_with_dates"],
    "E19": [],
    "E20": ["tenure_with_dates"],
    "R1":  ["tenure_with_dates"],
    "R2":  ["best_company_tier"],
    "R3":  ["leadership_signal_score","achievements"],
    "R4":  ["project_types","achievements","cert_count"],
    "R5":  ["total_experience_years"],
    "R6":  ["cert_count","skill_depth_priors"],
    "R7":  ["role_family_prior"],
    "R8":  ["total_experience_years","project_types","achievements","cert_count"],
    "R9":  ["career_progression_prior"],
    "R10": ["tenures","hop_rate"],
    "R11": [],
    "R12": [],
    "R13": ["total_experience_years","stability_score"],
    "R14": ["education_job_relevance","skill_depth_priors"],
    "R15": ["tenure_with_dates"],
    "R16": ["founder_title_count"],
    "R17": ["highest_institute_tier","best_company_tier","total_experience_years"],
    "R18": ["skill_depth_priors"],
    "R19": ["role_family_prior"],
    "R20": ["project_types"],
}

# Header row: signal names
ws5.cell(row=1, column=1, value="Code / Signal →").fill = fill(C_HEADER_DARK)
ws5.cell(row=1, column=1).font = bold_font(10, "FFFFFF")
ws5.cell(row=1, column=1).border = thin_border()
ws5.column_dimensions["A"].width = 9

for s_idx, sname in enumerate(SIGNAL_NAMES, 2):
    cell = ws5.cell(row=1, column=s_idx, value=sname)
    cell.fill = fill(C_HEADER_DARK)
    cell.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=45, wrap_text=True)
    cell.border = thin_border()
    ws5.column_dimensions[get_column_letter(s_idx)].width = 9
ws5.row_dimensions[1].height = 90

for r_idx, code in enumerate(EC_NAMES, 2):
    ws5.row_dimensions[r_idx].height = 18
    is_edge = code.startswith("E")
    row_bg = C_EDGE_LIGHT if is_edge else C_RED_LIGHT
    label_color = C_EDGE_HDR if is_edge else C_RED_HDR

    label_cell = ws5.cell(row=r_idx, column=1, value=code)
    label_cell.fill = fill(row_bg)
    label_cell.font = Font(bold=True, size=10, color=label_color, name="Calibri")
    label_cell.alignment = wrap_align("center", "center")
    label_cell.border = thin_border()

    triggers = EC_SIGNAL_MAP.get(code, [])
    for s_idx, sname in enumerate(SIGNAL_NAMES, 2):
        cell = ws5.cell(row=r_idx, column=s_idx)
        if sname in triggers:
            cell.value = "●"
            cell.fill = fill("353395" if is_edge else "991B1B")
            cell.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
        else:
            cell.fill = fill(C_WHITE)
        cell.alignment = wrap_align("center", "center")
        cell.border = thin_border()

ws5.freeze_panes = "B2"

# ─────────────────────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────────────────────
OUT = "Edge_Cases_Signals.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
