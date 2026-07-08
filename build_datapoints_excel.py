"""
Build comprehensive Resume Intelligence Data Points Excel
Covers all Resume / Recruiter / Panel data points with sub-parameters,
scoring logic, and methodology.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# COLOUR PALETTE
# ─────────────────────────────────────────────────────────────────────────────
C = {
    "hdr_dark":   "1F3864",   # dark navy  – sheet headers
    "hdr_med":    "2E75B6",   # medium blue
    "hdr_light":  "BDD7EE",   # pale blue
    "exp":        "E2EFDA",   # light green – experience rows
    "exp_hdr":    "375623",   # dark green
    "exp_hdr_l":  "70AD47",   # mid green
    "skill":      "FFF2CC",   # yellow – skills rows
    "skill_hdr":  "7F5700",   # dark amber
    "skill_hdr_l":"ED7D31",   # orange
    "edu":        "FCE4D6",   # peach – education rows
    "edu_hdr":    "833C00",   # dark orange
    "edu_hdr_l":  "C55A11",   # orange-red
    "rec":        "E2F0D9",   # pale green – recruiter
    "rec_hdr":    "375623",
    "panel":      "DAEEF3",   # pale cyan – panel
    "panel_hdr":  "17375E",
    "auto":       "EBF3FB",   # very pale blue – auto/python
    "bert":       "F2E7F5",   # pale purple – BERT
    "llm":        "FFF9E6",   # pale cream – LLM
    "white":      "FFFFFF",
    "gray_l":     "F2F2F2",
    "gray_d":     "D9D9D9",
    "red":        "FF0000",
    "orange":     "ED7D31",
    "green":      "70AD47",
    "gold":       "FFD966",
    "text_wh":    "FFFFFF",
    "reject":     "FFD7D7",
    "flag":       "FFC000",
}

thin = Side(style="thin", color="BFBFBF")
med  = Side(style="medium", color="7F7F7F")
thick= Side(style="medium", color="404040")
BORDER_THIN  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
BORDER_THICK = Border(left=thick, right=thick, top=thick, bottom=thick)
BORDER_HEAD  = Border(left=med,   right=med,   top=med,   bottom=med)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def hdr_cell(ws, row, col, value, bg, fg="FFFFFF", bold=True, size=10,
             halign="center", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size)
    c.alignment = align(h=halign, v="center", wrap=wrap)
    c.border = BORDER_HEAD
    return c

def data_cell(ws, row, col, value, bg="FFFFFF", bold=False, color="000000",
              halign="left", wrap=True, size=9):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=color, size=size)
    c.alignment = align(h=halign, v="center", wrap=wrap)
    c.border = BORDER_THIN
    return c

def set_col_width(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def merge_hdr(ws, row, c1, c2, value, bg, fg="FFFFFF", bold=True, size=11):
    ws.merge_cells(start_row=row, start_column=c1,
                   end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=value)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size)
    c.alignment = align(h="center", v="center")
    c.border = BORDER_THICK


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 1 — MASTER SCORING FRAMEWORK
# ═════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Master Scoring Framework"
ws1.freeze_panes = "A4"

merge_hdr(ws1, 1, 1, 15, "RESUME INTELLIGENCE — MASTER SCORING FRAMEWORK (100 pts)", "1F3864")
merge_hdr(ws1, 2, 1, 15,
    "3 Stages: Resume (Auto) → Recruiter Screening → Panel Interview  |  "
    "Methods: Pure Python · Python+BERT · Python+LLM · Recruiter/Panel Human Input",
    "2E75B6")

cols = ["#","Parameter","Module","Stage / Who Fills","Method",
        "Max Pts","Benchmark Score","Resume Auto\n(Score R)","Recruiter\n(Score A)",
        "Panel\n(Score B)","Final\n(Max Pts)","JD-Calibrated\nScore",
        "How Calculated","Key Sub-Parameters Tracked","Flag / Notes"]
for ci, col in enumerate(cols, 1):
    hdr_cell(ws1, 3, ci, col, C["hdr_dark"], size=9, wrap=True)

rows = [
    # ── EXPERIENCE 40 pts ──────────────────────────────────────────────────
    ("", "── EXPERIENCE SECTION ──", "", "", "", 40, "", "", "", "", 40, "", "", "", ""),
    ("E1","Companies Worked With","Experience","Resume + Recruiter Rescore","Python + LLM Judge",
     5, 5, "Y", "Y", "Y", 5, 5,
     "Best company tier across career → Tier 1=5 pts, Tier 2=4, Tier 3=3, Tier 4=2, Unknown=1",
     "Company Name · Tier (1–5) · Domain Tag · Company Type (Product/Services/Consulting) · "
     "Funding Stage · Headcount Band · Operating Model · Known Y/N · Signal Strength",
     "Probe if unknown company"),
    ("E2","Overall Experience / Relevant Experience","Experience","Resume","Python",
     3, 3, "Y", "Y", "Y", 3, 3,
     "With JD: ratio = min(total_yrs, jd_max)/total_yrs × 3. No JD bands: 10+=3, 6–10=2.5, 4–6=2, 2–4=1.5, <2=1",
     "Total YoE (years) · Relevant YoE · JD YoE Min · JD YoE Max · YoE Ratio · Band (<2/2-4/4-6/6-10/10+)",
     "REJECT if relevant_yoe < 70% of JD range"),
    ("E3","Career Progression","Experience","Resume + LLM Rescore","Python + BERT + LLM Judge",
     3, 3, "Y", "Y", "Y", 3, 3,
     "Seniority trajectory: IC→Lead→Sr→Arch scored 0–5, blended with BERT confidence tiers, LLM override possible",
     "Title per role · Seniority level (1=IC to 6=VP) · Promotion Y/N · Same-company growth · "
     "BERT class (FAST_TRACK/GROWING/LATERAL/DECLINING) · BERT confidence · Title velocity",
     "BERT guard clause applied"),
    ("E4","Stability","Experience","Resume","Python + LLM Judge",
     3, 3, "Y", "Y", "Y", 3, 3,
     "Base avg tenure: 36+m=5, 24–36=4, 18–24=3.5, 12–18=3, 8–12=2, <8=1.5; penalties for short stints + job-hopping",
     "Avg tenure (months) · Loyalty signal (LOW/MEDIUM/HIGH) · Short stints count · "
     "Hop rate (roles/yr) · Penalties applied · Loyalty bucket",
     "Soft flag if hop_rate > 1.5"),
    ("E5","Awards & Recognitions","Experience","Resume","Python + LLM Judge",
     3, 3, "Y", "Y", "Y", 3, 3,
     "Count named awards, promotions, patents, publications, conference talks. LLM validates genuine vs marketing",
     "Award count · Promotion count · Patent count · Publication count · Talk count · "
     "LLM verified (Y/N) · Types: AWARD/PROMOTION/PATENT/PUBLICATION/TALK",
     "LLM filters buzzwords"),
    ("E6","Mentorship / Code Reviews / Interviews","Experience","Resume + Recruiter","Python + BERT + LLM Judge",
     3, 3, "N", "Y", "Y", 3, 3,
     "Count instances of led/managed/mentored ≥2 roles=3pts, 1 instance=2, implied=1, none=0",
     "Mentored count · Led engineers count · Code review signal · Interview panel signal · "
     "BERT class (LEAD/FORMAL/IMPLIED/NONE) · BERT confidence · Roles with mentorship",
     "Recruiter validates during call"),
    ("E7","International Exposure","Experience","Resume + Recruiter","LLM Judge",
     2, 2, "N", "Y", "Y", 2, 2,
     "Python: scan for onsite/global/multi-country/relocation keywords → 2 or 0. LLM: implied vs explicit",
     "Keywords matched · Countries mentioned · Onsite signal (Y/N) · Relocation signal (Y/N) · "
     "Global team signal · Visa/work permit mention · LLM score (0/1/2)",
     "Pre-filled as 1 if keyword detected; recruiter confirms"),
    ("E8","Stakeholder Management","Experience","Resume + Recruiter","Python + BERT + LLM Judge",
     2, 2, "N", "Y", "Y", 2, 2,
     "BERT: C_LEVEL/CLIENT_FACING/INTERNAL/NONE; LLM final review. C-level=2, client-facing=1.5, internal=1, none=0",
     "Client-facing (Y/N) · C-level exposure (Y/N) · Cross-functional (Y/N) · "
     "BERT class (NONE/INTERNAL/CLIENT_FACING/C_LEVEL) · BERT confidence · Keywords matched",
     ""),
    ("E9","Career Breaks","Experience","Resume","Python",
     2, 2, "Y", "Y", "Y", 2, 2,
     "Gaps >3 months between jobs. 0 breaks=2pts, 1 break=1pt, 2 breaks=0, >2=0+REJECT",
     "Break count · Break durations (months each) · Break dates (start/end) · "
     "Reason classified (MBA/Maternity/Recession/Unknown) · Reject flag",
     "REJECT FLAG if >2 breaks. MBA overlap = no penalty. COVID 2020 = context flag"),
    ("E10","Project 1 — Latest Project","Experience","Resume + Recruiter + Panel","Python + LLM + Panel Feedback",
     8, 8, "Y", "Y", "Y", 8, 8,
     "8 criteria × 1pt each: (1)type known (2)title (3)desc>20c (4)duration≥3m (5)skills listed "
     "(6)domain tag (7)ownership verb+desc>50c (8)quantified impact",
     "Project Title · Project Type (DEVELOPMENT/MIGRATION/ANALYTICS/INFRA/RESEARCH/CONSULTING/MAINTENANCE) · "
     "Description · Duration (months) · Skills used · Domain tag · Ownership verbs · "
     "Quantified impact (% or numbers) · Complexity score (0–5) · Role played · "
     "Candidate signal (EXCELLENT/STRONG/AVERAGE/WEAK) · Implied skills · Green flags · Red flags · "
     "LLM confirmed type · Skills verified · Gaps detected",
     "Scope for change based on recruiter/panel input"),
    ("E11","Project 2 — 2nd Latest Project","Experience","Resume + Recruiter + Panel","Python + LLM + Panel Feedback",
     6, 6, "Y", "Y", "Y", 6, 6,
     "6 criteria × 1pt each: (1)type known (2)title (3)desc>20c (4)duration≥3m (5)skills listed (6)domain tag",
     "Project Title · Project Type · Description · Duration (months) · Skills used · "
     "Domain tag · Complexity score · Role signal",
     "Same as Project 1 but no depth/impact criteria"),

    # ── EDUCATION 15 pts ──────────────────────────────────────────────────
    ("", "── EDUCATION SECTION ──", "", "", "", 15, "", "", "", "", 15, "", "", "", ""),
    ("ED1","Institute Tier, GPA, Stream","Education","Resume","Python",
     5, 5, "Y", "Y", "Y", 5, 5,
     "TIER_1=4 base +1 if GPA EXCELLENT/GOOD. TIER_2=3+0.5 if EXCELLENT. TIER_3=2. TIER_4=1. Capped at 5",
     "Institution name · Tier (TIER_1/2/3/4/UNKNOWN) · GPA value · GPA scale (4pt/10pt/100pt) · "
     "GPA band (EXCELLENT/GOOD/ACCEPTABLE/LOW) · Canonical institution name · City · NIRF rank · "
     "Streams offered · IT stream (Y/N) · Category (IIT/IIM/NIT/etc.)",
     "Longest-match algorithm for institution lookup"),
    ("ED2","Highest Education + Stream","Education","Resume","Python",
     2, 2, "Y", "Y", "Y", 2, 2,
     "PhD/Master=2pts, Bachelor=1.5, Diploma=1, Unknown=0.5",
     "Degree level (PHD/MASTER/BACHELOR/DIPLOMA/UNKNOWN) · Field of study · "
     "Course family (ENGINEERING/ANALYTICS/SCIENCE/MANAGEMENT/ARTS) · "
     "Tech fit (TECH/SEMI_TECH/NON_TECH) · Course value signal (HIGH/MEDIUM/FOUNDATIONAL)",
     "IT stream: CS/CSE/IT/DS/AI/ML/MCA/Statistics/Analytics/Quant Econ/Econometrics"),
    ("ED3","Education Gaps","Education","Resume + Recruiter","Python + LLM Judge",
     1, 1, "Y", "Y", "Y", 1, 1,
     "≤6m=1pt, 6–12m=0.5, >12m=0+REJECT FLAG",
     "Gap start date · Gap end date · Gap duration (months) · Gap reason (if detected) · "
     "Education start date · Education end date · REJECT flag",
     "REJECT if gap > 12 months"),
    ("ED4","Education to Job Relevance","Education","Resume","Python + LLM Judge",
     2, 2, "Y", "Y", "Y", 2, 2,
     "HIGH=2, MEDIUM=1.5, FOUNDATIONAL=0.5, UNKNOWN=1. Mapped from strongest course value signal",
     "Course relevance (HIGH/MEDIUM/FOUNDATIONAL) · Stream relevance rank (1–5) · "
     "Domain match (Y/N) · Tech degree (Y/N) · Non-IT to IT switch signal",
     ""),
    ("ED5","Executive / Distance Learning","Education","Resume","Python or LLM Judge",
     1, 1, "Y", "Y", "Y", 1, 1,
     "Keywords: executive, continuing, distance, certification, online, mooc → 1pt",
     "Exec education detected (Y/N) · Keywords matched · Institution type · Year · Mode (online/offline)",
     "Boolean signal"),
    ("ED6","Patents / Publications","Education","Resume","Python",
     2, 2, "Y", "Y", "Y", 2, 2,
     "Boolean × 2pts. TIER_1 institution bonus 0.5pt even without patents. TIER_2 gets 0.25pt",
     "Patent count · Publication count · Conference talk count · Year · Title · "
     "Co-authors · TIER bonus applied · Patent detected (Y/N)",
     "Tier 1 always gets 0.5 base credit"),
    ("ED7","LinkedIn / Social Media Activeness","Education","Recruiter","Python",
     1, 1, "N", "Y", "Y", 1, 1,
     "Recruiter checks profile: 1=active, 0=absent",
     "LinkedIn URL · Profile active (Y/N) · Last post recency · Connections approx · "
     "GitHub link · Kaggle link · Stack Overflow link",
     "Recruiter fills during screening call"),
    ("ED8","Extra-Curricular Activities","Education","Recruiter","Python",
     1, 1, "Y", "Y", "Y", 1, 1,
     "Resume signal: 1=confirmed, 0=none",
     "Activities listed · Type (sports/volunteering/hackathon/club) · "
     "Leadership role (Y/N) · Relevance to tech",
     ""),

    # ── SKILLS 45 pts ──────────────────────────────────────────────────────
    ("", "── SKILLS SECTION ──", "", "", "", 45, "", "", "", "", 45, "", "", "", ""),
    ("S1","Skill List — Years of Experience & Timeline","Skills","Resume","Python",
     6, 6, "Y", "Y", "Y", 6, 6,
     "Each validated APPLIED+ skill with clear years = 1pt, max 6",
     "Skill name · Canonical cluster · Raw tenure (years) · Weighted evidence tenure · "
     "Recency (RECENT/MID/OLD/UNKNOWN) · Evidence level (NONE/MENTION/WEAK/APPLIED/DEEP/EXPERT) · "
     "First used year · Last used year · Timeline start · Timeline end · Active (Y/N)",
     "Only APPLIED+ skills counted"),
    ("S2","Mandatory Skills (per JD)","Skills","Resume","Python",
     0, 0, "Y", "Y", "Y", 0, 0,
     "Flag only — no score. ✅ if matched, ❌ if missing",
     "Skill name · Matched (Y/N) · Evidence level · Years · Weight (from JD) · Gap flag",
     "Flag shown to recruiter; no numeric score"),
    ("S3","Good-to-Have Skills (per JD)","Skills","Resume","Python",
     0, 0, "Y", "Y", "Y", 0, 0,
     "Flag only — no score. ✅ if matched, ❌ if missing",
     "Skill name · Matched (Y/N) · Evidence level · Years · Weight (from JD)",
     "Flag shown to recruiter"),
    ("S4","Certifications — Validity, Type, Relevance","Skills","Resume","Python + LLM Judge",
     3, 3, "Y", "Y", "Y", 3, 3,
     "clamp(cert_count, 0, 3). LLM checks validity and relevance",
     "Certification name · Issuing body · Year obtained · Expiry year · "
     "Still valid (Y/N) · Skill mapped · Relevance (HIGH/MEDIUM/LOW) · "
     "Type (cloud/data/security/mgmt) · LLM verified",
     "Max 3 pts. Cert farming red flag if ≥5 certs + depth ≤ FOUNDATIONAL"),
    ("S5","Skill Depth","Skills","Panel + BERT","Panel Feedback + LLM Judge",
     8, 8, "Y", "N (Panel)", "Y", 8, 8,
     "Top 5 skills via evidence_level→0–5 raw. BERT blended by confidence. (avg_blended/5)×8. JD: role-weighted",
     "Depth per skill (AWARENESS/FOUNDATIONAL/HANDS_ON/ADVANCED/ARCHITECT_LEVEL) · "
     "Evidence level per skill · BERT depth prediction · BERT confidence · Blend ratio · "
     "Blended score per skill · Architecture signal (Y/N) · Coding signal (Y/N) · "
     "Project context (DEVELOPMENT/MAINTENANCE/UNKNOWN) · Open source signal (Y/N) · "
     "Attributed roles · Upskill signal",
     "BERT confidence tiers: ≥0.65→65% BERT, 0.45–0.64→50/50, <0.45→evidence only"),
    ("S6","Skill Recency","Skills","Panel + Python","Panel Feedback + LLM Judge",
     6, 6, "Y", "N", "Y", 6, 6,
     "(count RECENT or CURRENT / total skills) × 6",
     "Recency per skill (RECENT/MID/OLD/UNKNOWN) · RECENT count · Total skill count · "
     "RECENT % · Last used year per skill · Current active flag",
     "RECENT = used in last 2 years"),
    ("S7","Skills Learning Acumen","Skills","Resume + Recruiter","LLM Judge",
     3, 3, "Y", "Y", "Y", 3, 3,
     "fast_learner (≥2 new/yr for ≥2 yrs)=3pts. New skills ≥3 yrs=2. 1–2 yrs=1. None=0",
     "New skills per year · Yearly skill learning list · Fast learner flag · "
     "Years with new skill uptake · Skill growth rate · Learning acumen score",
     ""),
    ("S8","Coding Platforms / Community Contributions","Education","Resume","Python + LLM Judge",
     3, 3, "Y", "Y", "Y", 3, 3,
     "Count OSS signals. ≥3=3pts, 2=2, 1=1, 0=0",
     "GitHub (Y/N) · Stack Overflow (Y/N) · LeetCode (Y/N) · HackerRank (Y/N) · "
     "Kaggle (Y/N) · Other (Y/N) · Total platforms · OSS contributor (Y/N)",
     ""),
    ("S9","Communication & Presentation Skills","Skills","Panel","Panel Feedback + LLM Judge",
     5, 5, "N", "N", "Y", 5, 5,
     "Panel assigns 0–5. Verbal clarity, structure, audience adaptability",
     "Clarity score · Structure score · Confidence level · Audience adaptability · "
     "Jargon use (appropriate Y/N) · Panel notes · Recruiter pre-screen note",
     "Panel-only parameter"),
    ("S10","Domain Skills","Skills","Panel","Panel Feedback + LLM Judge",
     5, 5, "N", "N", "Y", 5, 5,
     "Panel assigns 0–5. Domain-specific knowledge depth via scenario questions",
     "Domain tested · Questions asked · Score per question · Depth demonstrated · "
     "Business context understanding · Industry knowledge · Panel notes",
     "Panel-only parameter"),
    ("S11","Project Explanation Skills","Skills","Panel + Recruiter","Panel Feedback + LLM Judge",
     3, 3, "N", "Y", "Y", 3, 3,
     "3=clear problem→design→outcome. 2=good structure. 1=disjointed. 0=cannot explain",
     "Problem statement clarity · Solution design quality · Role played clarity · "
     "Outcome quantified (Y/N) · Ownership clarity · Trade-off awareness · "
     "Score breakdown · Panel notes",
     "Scale: 0=can't explain, 1=disjointed, 2=good, 3=clear P→D→O"),
    ("S12","Coding Skills","Skills","Panel","Panel Feedback + LLM Judge",
     0, 0, "N", "N", "Y", 0, 0,
     "Qualitative — no numeric score. Panel narrative only",
     "Language tested · Problem level (easy/medium/hard) · Approach quality · "
     "Clean code (Y/N) · Edge cases handled · Time complexity awareness · Panel notes",
     "No numeric score; qualitative narrative"),
    ("S13","Conceptual Skills","Skills","Panel","Panel Feedback + LLM Judge",
     0, 0, "N", "N", "Y", 0, 0,
     "Qualitative — no numeric score. Panel narrative only",
     "Concepts tested · Depth of understanding · Theory vs practice balance · "
     "First-principles thinking · Panel notes",
     "No numeric score"),
    ("S14","Problem Solving Skills","Skills","Panel","Panel Feedback + LLM Judge",
     3, 3, "N", "N", "Y", 3, 3,
     "Panel assigns 0–3. Live problem-solving ability",
     "Problem type · Approach taken · Structure of solution · Creativity · "
     "Time to solution · Hints needed (Y/N) · Panel score · Panel notes",
     "Panel-only parameter"),

    # ── TOTALS ────────────────────────────────────────────────────────────
    ("", "── TOTALS ──", "", "", "", 100, "", "", "", "", 100, "", "", "", ""),
    ("", "Experience Subtotal", "", "", "", 40, "", "~29 auto", "~37 max", "~40 max", 40, 40, "", "", ""),
    ("", "Education Subtotal", "", "", "", 15, "", "~13 auto", "~14 max", "~15 max", 15, 15, "", "", ""),
    ("", "Skills Subtotal", "", "", "", 45, "", "~23 auto", "~32 max", "~45 max", 45, 45, "", "", ""),
    ("", "GRAND TOTAL", "", "", "", 100, "", "Stage R: /76→norm", "Stage A: /87→norm", "Stage B: /100", 100, 100, "", "", ""),
]

exp_params  = {"E1","E2","E3","E4","E5","E6","E7","E8","E9","E10","E11"}
edu_params  = {"ED1","ED2","ED3","ED4","ED5","ED6","ED7","ED8"}
skill_params= {"S1","S2","S3","S4","S5","S6","S7","S8","S9","S10","S11","S12","S13","S14"}
section_rows= {"── EXPERIENCE SECTION ──","── EDUCATION SECTION ──","── SKILLS SECTION ──","── TOTALS ──",
               "Experience Subtotal","Education Subtotal","Skills Subtotal","GRAND TOTAL"}

for r_idx, row_data in enumerate(rows, 4):
    param_id = row_data[0]
    param_name = row_data[1]

    if param_name in section_rows:
        if param_name.startswith("──"):
            section = param_name.replace("──","").strip()
            bg = (C["exp_hdr"] if "EXPERIENCE" in section
                  else C["edu_hdr"] if "EDUCATION" in section
                  else C["skill_hdr"] if "SKILLS" in section else C["hdr_dark"])
            merge_hdr(ws1, r_idx, 1, 15, f"  {param_name}", bg, size=10)
        else:
            bg = C["gray_l"]
            for ci, val in enumerate(row_data, 1):
                data_cell(ws1, r_idx, ci, val, bg=bg, bold=(param_name=="GRAND TOTAL"),
                          color="000000", size=9)
        continue

    # colour per module
    if param_id in exp_params:
        row_bg = C["exp"]
    elif param_id in edu_params:
        row_bg = C["edu"]
    elif param_id in skill_params:
        row_bg = C["skill"]
    else:
        row_bg = C["white"]

    for ci, val in enumerate(row_data, 1):
        wrap = ci in (14, 15, 13)  # sub-parameters and notes columns — wrap
        data_cell(ws1, r_idx, ci, val, bg=row_bg, bold=(ci == 2),
                  halign="center" if ci in (1,6,7,8,9,10,11,12) else "left",
                  wrap=wrap, size=9)

set_col_width(ws1, [5, 28, 12, 22, 22, 7, 9, 9, 9, 9, 7, 9, 35, 60, 28])
ws1.row_dimensions[1].height = 22
ws1.row_dimensions[2].height = 18
ws1.row_dimensions[3].height = 36


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 2 — WORK EXPERIENCE DATA POINTS (per company)
# ═════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Work Experience Data Points")
ws2.freeze_panes = "A4"

merge_hdr(ws2, 1, 1, 7, "WORK EXPERIENCE — COMPLETE DATA POINTS PER COMPANY / ROLE", C["exp_hdr"])
merge_hdr(ws2, 2, 1, 7, "Every field extracted / inferred from the resume for each job role", C["exp_hdr_l"])

exp_sections = [
    ("COMPANY IDENTIFICATION", [
        ("Company Name",          "Raw name as written on resume",                     "Text",                     "Python extract"),
        ("Canonical Company Name","Normalised name after fuzzy match",                  "Text",                     "Python lookup"),
        ("Company Known",         "Is company in our database?",                        "Y / N",                    "Python lookup"),
        ("Company Tier",          "Quality tier of employer",                           "1 (FAANG) to 5 (Unknown)", "Python + LLM fallback"),
        ("Signal Strength",       "Employer quality signal",                            "HIGH / MEDIUM / LOW / UNKNOWN","Python lookup"),
        ("Company Domain",        "Industry sector",                                    "ECOMMERCE / BFSI / HEALTHCARE / SAAS / EDTECH / FINTECH / etc.","Python + LLM"),
        ("Sub-Domain",            "More specific sector",                               "Text",                     "LLM"),
        ("Company Type",          "Business model",                                     "PRODUCT / CONSULTING / SERVICES / STARTUP / MNC","Python + LLM"),
        ("Is Funded",             "Has external funding",                               "Y / N",                    "Python lookup"),
        ("Funding Stage",         "Stage of funding",                                   "SEED / SERIES_A / SERIES_B / SERIES_C / IPO / LISTED / N/A","Python lookup"),
        ("Headcount Band",        "Company size",                                       "<50 / 50–200 / 200–1000 / 1000–5000 / 5000+","Python lookup"),
        ("Work Type",             "Type of work done",                                  "PRODUCT_BUILD / CONSULTING / OUTSOURCING / HYBRID","Python lookup"),
        ("Culture Signals",       "Notable cultural tags",                              "List: fast-paced / remote-first / equity-driven / etc.","Python lookup"),
        ("Operating Model",       "Consulting vs Product vs Platform",                  "CONSULTING / PRODUCT / PLATFORM_INFRA / DOMAIN_SPECIALIST","Python heuristic"),
        ("Similar Companies",     "Peer companies at same tier",                        "List (max 5)",             "LLM"),
    ]),
    ("ROLE / TENURE", [
        ("Job Title",             "Title as written on resume",                         "Text",                     "Python extract"),
        ("Seniority Level",       "Numeric seniority for progression tracking",         "1=IC, 2=SrIC, 3=Lead, 4=Manager, 5=Director, 6=VP","Python heuristic"),
        ("Start Date",            "Role start date",                                    "MM/YYYY or YYYY",          "Python extract"),
        ("End Date",              "Role end date (or 'Present')",                       "MM/YYYY / Present",        "Python extract"),
        ("Duration (months)",     "How long in this role",                              "Integer (months)",         "Python calc"),
        ("Duration (years)",      "Rounded years",                                      "Float",                    "Python calc"),
        ("Is Current Role",       "Still in this role?",                                "Y / N",                    "Python flag"),
        ("Same Company Growth",   "Promotion within same employer",                     "Y / N",                    "Python detect"),
        ("Internal Title Changes","Number of title changes at same company",            "Integer",                  "Python count"),
        ("Employment Gap After",  "Months gap before next role",                        "Integer (months) or 0",    "Python calc"),
        ("Gap Reason",            "Classified reason for gap",                          "MBA / MATERNITY / RECESSION / LAYOFF / UNKNOWN","Python + LLM"),
    ]),
    ("ROLE CONTENT / SKILLS", [
        ("Skills Used",           "Skills mentioned in this role",                      "List of skill names",      "Python extract"),
        ("Primary Skill Cluster", "Dominant skill family",                              "PROGRAMMING / ML / CLOUD / BIG_DATA / etc.","Python taxonomy"),
        ("Project Types",         "Types of work done (up to 8)",                       "DEVELOPMENT / MIGRATION / MAINTENANCE / SUPPORT / POC / ANALYTICS","Python heuristic"),
        ("Domain Tags",           "Business domains touched",                           "ECOMMERCE / BFSI / HEALTHCARE / RETAIL / etc.","Python heuristic"),
        ("Leadership Signal",     "Led a team or initiative",                           "Score 0–5",                "Python keyword scan"),
        ("Complexity Signal",     "Work complexity (distributed, scalable, real-time)", "Score 0–5",                "Python keyword scan"),
        ("Ownership Signal",      "End-to-end ownership verbs",                        "Score 0–5",                "Python keyword scan"),
        ("Problem Solving Signal","Solved / improved / optimised verbs",               "Score 0–5",                "Python keyword scan"),
        ("Client-Facing",         "Worked directly with clients",                       "Y / N",                    "Python keyword scan"),
        ("International Exposure","Global or onsite work signal",                       "Y / N",                    "Python keyword scan"),
        ("Business Impacts",      "Quantified impact statements",                       "List of text",             "Python regex"),
        ("Impact Count",          "Number of quantified impacts",                       "Integer",                  "Python count"),
        ("Support/Maintenance %", "Proportion of support vs build work",                "LOW / MEDIUM / HIGH",      "Python heuristic"),
        ("Decision Maker",        "Identified as decision maker (6+ yrs)",              "Y / N",                    "Python flag"),
    ]),
    ("DNA & FIT SIGNALS", [
        ("DNA Contribution",      "Which DNA type this role supports",                  "CONSULTING / PRODUCT / PLATFORM_INFRA / DOMAIN_SPECIALIST / RESEARCH","Python score"),
        ("Company-Skill Align",   "How well company type aligns with role skills",      "ALIGNED / PARTIAL / MISMATCH","Python"),
        ("Role Family Hints",     "Which role family this experience supports",         "DATA_SCIENTIST / ML_ENGINEER / DE / etc.","Semantic taxonomy"),
        ("Consulting Hints",      "Consulting-flavoured signals",                       "Y / N",                    "Python keyword"),
        ("Product Hints",         "Product-flavoured signals",                          "Y / N",                    "Python keyword"),
    ]),
]

row_n = 3
for section_name, fields in exp_sections:
    merge_hdr(ws2, row_n, 1, 7, f"  {section_name}", C["exp_hdr"], size=10)
    row_n += 1
    cols2 = ["#","Data Point / Field","Description","Sample Values / Options","How Detected","Stored As","Used In Parameter(s)"]
    for ci, col in enumerate(cols2, 1):
        hdr_cell(ws2, row_n, ci, col, C["hdr_med"], size=9)
    row_n += 1
    for fi, (fname, fdesc, fvals, fmethod) in enumerate(fields, 1):
        data_cell(ws2, row_n, 1, fi, bg=C["exp"], halign="center")
        data_cell(ws2, row_n, 2, fname, bg=C["exp"], bold=True)
        data_cell(ws2, row_n, 3, fdesc, bg=C["white"], wrap=True)
        data_cell(ws2, row_n, 4, fvals, bg=C["gray_l"], wrap=True)
        data_cell(ws2, row_n, 5, fmethod, bg=C["auto"], wrap=True)
        data_cell(ws2, row_n, 6, "JSON field in experience_engine output", bg=C["white"], wrap=True)
        row_n += 1

set_col_width(ws2, [5, 28, 38, 35, 22, 28, 25])


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 3 — SKILLS DATA POINTS (per skill)
# ═════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Skills Data Points")
ws3.freeze_panes = "A4"

merge_hdr(ws3, 1, 1, 7, "SKILLS — COMPLETE DATA POINTS PER SKILL", C["skill_hdr"])
merge_hdr(ws3, 2, 1, 7,
    "Every field tracked for each skill found on the resume — far beyond simple keyword matching",
    C["skill_hdr_l"], fg="FFFFFF")

skill_sections = [
    ("SKILL IDENTITY & CLASSIFICATION", [
        ("Raw Skill Name",        "Skill as written on resume",             "e.g. 'Python 3', 'Py'",   "Python extract"),
        ("Canonical Skill Name",  "Normalised skill name",                   "e.g. 'Python'",           "Taxonomy lookup"),
        ("Canonical Cluster",     "High-level skill family / cluster",       "PROGRAMMING / STATISTICS_ML / DEEP_LEARNING_GENAI / BIG_DATA / CLOUD_INFRA / MLOPS_DEPLOYMENT / VISUALIZATION_BI / DATA_MANAGEMENT / SYSTEMS_ARCHITECTURE / PRODUCT_ANALYTICS / EXPERIMENTATION_RCA / domain clusters","Taxonomy"),
        ("Alias Matched",         "Which alias triggered the match",         "Text",                     "Python lookup"),
        ("Detection Zone",        "Where skill was found on resume",         "skills_section / role_description / project / certification","Python"),
    ]),
    ("EVIDENCE & DEPTH", [
        ("Evidence Level",        "Raw evidence strength",                   "NONE / MENTION / WEAK / APPLIED / DEEP / EXPERT","Python rule engine"),
        ("Evidence Score (0–5)",  "Numeric evidence score",                  "NONE=0, MENTION=0.5, WEAK=1.5, APPLIED=3.0, DEEP=4.0, EXPERT=5.0","Python lookup"),
        ("Depth Label",           "Skill depth category",                    "AWARENESS / FOUNDATIONAL / HANDS_ON / ADVANCED / ARCHITECT_LEVEL","Python + BERT"),
        ("Depth Score (0–5)",     "Numeric depth score",                     "AWARENESS=0.5, FOUNDATIONAL=1.5, HANDS_ON=3.0, ADVANCED=4.0, ARCHITECT_LEVEL=5.0","Python lookup"),
        ("BERT Depth Prediction", "BERT classifier prediction",              "AWARENESS / FOUNDATIONAL / HANDS_ON / ADVANCED / ARCHITECT_LEVEL","BERT model"),
        ("BERT Confidence",       "BERT prediction confidence",              "0.0 – 1.0",                "BERT model"),
        ("BERT Blend Tier",       "Which blend ratio applied",               "≥0.65 → 65% BERT / <0.45 → evidence only","Python"),
        ("Blended Score",         "Final blended depth score",               "0.0 – 5.0",                "Python blend formula"),
        ("Guard Clause Applied",  "BERT AWARENESS overridden by evidence",   "Y / N",                    "Python guard"),
    ]),
    ("TENURE & RECENCY", [
        ("Raw Tenure (years)",    "Total years skill appears on resume",     "Float",                    "Python calc"),
        ("Weighted Evidence Tenure","Tenure weighted by evidence quality",   "Float (credible use years)","Python calc"),
        ("Recency",               "How recently skill was used",             "RECENT (<2 yrs) / MID (2–5 yrs) / OLD (>5 yrs) / UNKNOWN","Python calc"),
        ("First Used Year",       "Year skill first appeared",               "Integer",                  "Python extract"),
        ("Last Used Year",        "Year skill last used",                    "Integer",                  "Python extract"),
        ("Is Currently Active",   "Still using this skill",                  "Y / N",                    "Python flag"),
        ("Matched Context Count", "Number of contexts (roles/projects) where skill appears","Integer",   "Python count"),
    ]),
    ("CONTEXT SIGNALS", [
        ("Coding Signal",         "Used in building/coding context",         "Y / N",                    "Python keyword: built/developed/implemented/script/code/API/pipeline"),
        ("Architecture Signal",   "Used in design/architecture context",     "Y / N",                    "Python keyword: system design/architecture/scalable/migration"),
        ("Project Context",       "Type of project where skill was used",    "DEVELOPMENT / MAINTENANCE_SUPPORT / UNKNOWN","Python heuristic"),
        ("Open Source Signal",    "Contributed to OSS using this skill",     "Y / N",                    "Python detect: GitHub/contributor/pull request"),
        ("Attributed Roles",      "Roles where skill is attributed",         "List of role titles",       "Python extract"),
        ("Project Contexts",      "Projects where skill appears",            "List of project names",     "Python extract"),
        ("Upskill Signal",        "Evidence of growing proficiency over time","Y / N",                   "Python: recent evidence + repeated roles + certs + timeline"),
    ]),
    ("ARTIFACT EVIDENCE", [
        ("Certification for Skill","Certification specifically for this skill","Text or None",            "Python cross-ref"),
        ("Patent Using Skill",    "Patent involving this skill",              "Y / N",                    "Python cross-ref"),
        ("Achievement Mentioning Skill","Award/recognition citing this skill","Y / N",                   "Python cross-ref"),
        ("Talk / Publication",    "Presented/published on this skill",        "Y / N",                    "Python cross-ref"),
    ]),
    ("SCORING OUTPUT", [
        ("Skill JD Match",        "Matches a JD mandatory skill",            "MANDATORY / GOOD_TO_HAVE / NONE","Python JD match"),
        ("JD Skill Weight",       "Weight assigned in JD configuration",     "Float 0–2",                "Client config"),
        ("Skill Score Contribution","Points this skill contributes to S5",   "Float",                    "Python calc"),
        ("Skill Recency Flag",    "Counted toward recency score",            "Y / N",                    "Python flag"),
        ("Counted in Top 5",      "In the top-5 skills for depth scoring",   "Y / N",                    "Python selection"),
        ("Unique Combination",    "Rare skill combination signal",            "Y / N",                    "Python heuristic"),
    ]),
]

row_n3 = 3
for section_name, fields in skill_sections:
    merge_hdr(ws3, row_n3, 1, 7, f"  {section_name}", C["skill_hdr"], fg="FFFFFF", size=10)
    row_n3 += 1
    cols3 = ["#","Data Point / Field","Description","Values / Options","Detection Method","Formula / Logic","Used In Parameter(s)"]
    for ci, col in enumerate(cols3, 1):
        hdr_cell(ws3, row_n3, ci, col, C["hdr_med"], size=9)
    row_n3 += 1
    for fi, (fname, fdesc, fvals, fmethod) in enumerate(fields, 1):
        data_cell(ws3, row_n3, 1, fi, bg=C["skill"], halign="center")
        data_cell(ws3, row_n3, 2, fname, bg=C["skill"], bold=True)
        data_cell(ws3, row_n3, 3, fdesc, bg=C["white"], wrap=True)
        data_cell(ws3, row_n3, 4, fvals, bg=C["gray_l"], wrap=True)
        data_cell(ws3, row_n3, 5, fmethod, bg=C["auto"], wrap=True)
        data_cell(ws3, row_n3, 6, "", bg=C["white"])
        row_n3 += 1

set_col_width(ws3, [5, 28, 38, 38, 28, 22, 20])


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 4 — EDUCATION DATA POINTS (per entry)
# ═════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Education Data Points")
ws4.freeze_panes = "A4"
merge_hdr(ws4, 1, 1, 7, "EDUCATION — COMPLETE DATA POINTS PER DEGREE / INSTITUTION ENTRY", C["edu_hdr"])
merge_hdr(ws4, 2, 1, 7, "Every field tracked for each education entry on the resume", C["edu_hdr_l"])

edu_sections = [
    ("INSTITUTION", [
        ("Institution Name (Raw)","As written on resume",                    "Text",                     "Python extract"),
        ("Canonical Institution", "Normalised full name",                     "e.g. 'Indian Institute of Technology Bombay'","Taxonomy longest-match"),
        ("Institute Tier",        "Quality tier",                             "TIER_1 / TIER_2 / TIER_3 / TIER_4 / UNKNOWN","Python lookup (500+ institutions)"),
        ("Category",              "Institution category",                     "IIT / IIM / ISI / ISB / NIT / IIIT / STATE / REGIONAL / GLOBAL","Python lookup"),
        ("NIRF Rank",             "National ranking if known",                "Integer or N/A",           "Python lookup"),
        ("City",                  "Campus city",                              "Text",                     "Python lookup"),
        ("Streams",               "Disciplines offered",                      "List",                     "Python lookup"),
        ("Global Top 200",        "Is it a global top-200 university?",       "Y / N",                    "Python lookup"),
    ]),
    ("DEGREE / COURSE", [
        ("Degree Raw",            "Degree as written",                        "e.g. 'B.Tech', 'M.Sc'",   "Python extract"),
        ("Degree Level",          "Normalised degree level",                  "PHD / MASTER / BACHELOR / DIPLOMA / UNKNOWN","Python heuristic"),
        ("Field of Study (Raw)",  "Field as written",                         "e.g. 'Computer Science and Engineering'","Python extract"),
        ("Course Family",         "Normalised course family",                 "ENGINEERING / ANALYTICS / SCIENCE / MANAGEMENT / ARTS / COMMERCE / COMPUTER_APPLICATIONS / RESEARCH","Taxonomy longest-match"),
        ("Course Value Signal",   "Academic value for tech roles",            "HIGH / MEDIUM / FOUNDATIONAL / UNKNOWN","Python lookup"),
        ("Tech Fit",              "Degree tech-relevance",                    "TECH / SEMI_TECH / NON_TECH / UNKNOWN","Python heuristic"),
        ("IT Stream",             "Is it an IT/tech stream?",                 "Y / N",                    "Python: CS/CSE/IT/DS/AI/ML/MCA/Stats/Analytics/Quant Econ/Econometrics/OR"),
        ("Stream Relevance Rank", "Stream relevance ranking for tech (1=best)","1=ECE, 2=Mech/EE, 3=Chemistry, 4=Math, 5=Civil+","Python"),
    ]),
    ("GPA / ACADEMIC PERFORMANCE", [
        ("GPA Value (Raw)",       "GPA/percentage/CGPA as on resume",         "e.g. '8.5', '75%', '3.9/4.0'","Python extract"),
        ("GPA Numeric",           "Normalised numeric value",                  "Float",                    "Python parse"),
        ("GPA Scale",             "Scale detected",                           "10_POINT / 4_POINT / 100_POINT","Python infer"),
        ("GPA Band",              "Performance label",                        "EXCELLENT / GOOD / ACCEPTABLE / LOW","Python benchmarks"),
        ("GPA Present",           "Was a GPA given?",                         "Y / N",                    "Python flag"),
    ]),
    ("DATES & GAPS", [
        ("Start Year",            "Year education started",                   "Integer",                  "Python extract"),
        ("End Year",              "Year education completed",                  "Integer",                  "Python extract"),
        ("Duration (years)",      "Years of study",                           "Float",                    "Python calc"),
        ("Education Gap (months)","Gap between education end and job start",  "Integer (months)",         "Python calc"),
        ("Gap Flag",              "Was there a significant gap?",             "NONE / MINOR (<6m) / MAJOR (6–12m) / REJECT (>12m)","Python rule"),
        ("Distance Learning",     "Was this a distance/online degree?",       "Y / N",                    "Python keyword detect"),
    ]),
    ("ACADEMIC ACHIEVEMENTS", [
        ("GPA Bonus Applied",     "GPA bonus points added to score",          "Float",                    "Python calc"),
        ("Patents Count",         "Number of patents",                        "Integer",                  "Python extract"),
        ("Publications Count",    "Number of publications",                   "Integer",                  "Python extract"),
        ("Conference Talks Count","Number of conference presentations",       "Integer",                  "Python extract"),
        ("Exec Education",        "Executive/continuing education detected",   "Y / N",                    "Python keyword: executive/continuing/distance/mooc"),
        ("Elite Bonus Applied",   "TIER_1 base patent credit (0.5pt)",        "Y / N",                    "Python rule"),
    ]),
    ("COMPUTED EDUCATION SCORE", [
        ("Raw Education Score",   "Full 0–10 score before rubric normalisation","Float",                  "Python formula: base+gpa+course+degree"),
        ("Institute Tier Score",  "Contribution from institute tier",          "0–5",                     "Python formula"),
        ("Degree Level Score",    "Contribution from degree level",            "0–2",                     "Python formula"),
        ("Relevance Score",       "Contribution from course relevance",        "0–2",                     "Python formula"),
        ("Gap Score",             "Contribution from gap check",               "0–1",                     "Python rule"),
        ("Best Entry Used",       "Highest-scoring education entry selected",  "Y / N",                    "Python (best-of-multiple logic)"),
    ]),
]

row_n4 = 3
for section_name, fields in edu_sections:
    merge_hdr(ws4, row_n4, 1, 7, f"  {section_name}", C["edu_hdr"], size=10)
    row_n4 += 1
    for ci, col in enumerate(["#","Data Point / Field","Description","Values / Options","Detection Method","Score Impact","Notes"], 1):
        hdr_cell(ws4, row_n4, ci, col, C["hdr_med"], size=9)
    row_n4 += 1
    for fi, (fname, fdesc, fvals, fmethod) in enumerate(fields, 1):
        data_cell(ws4, row_n4, 1, fi, bg=C["edu"], halign="center")
        data_cell(ws4, row_n4, 2, fname, bg=C["edu"], bold=True)
        data_cell(ws4, row_n4, 3, fdesc, bg=C["white"], wrap=True)
        data_cell(ws4, row_n4, 4, fvals, bg=C["gray_l"], wrap=True)
        data_cell(ws4, row_n4, 5, fmethod, bg=C["auto"], wrap=True)
        data_cell(ws4, row_n4, 6, "", bg=C["white"])
        row_n4 += 1

set_col_width(ws4, [5, 28, 38, 38, 32, 20, 20])


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 5 — RECRUITER STAGE DATA POINTS
# ═════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Recruiter Stage")
ws5.freeze_panes = "A4"
merge_hdr(ws5, 1, 1, 8, "RECRUITER SCREENING STAGE — ALL DATA POINTS CAPTURED DURING PHONE / VIDEO SCREEN", C["rec_hdr"])
merge_hdr(ws5, 2, 1, 8,
    "These data points are captured by the recruiter after resume scoring and used to rescore / validate the auto-scored resume",
    "375623", fg="E2F0D9")

rec_cols = ["#","Data Point","Category","Input Type","Options / Format","Score Impact","Maps To Parameter","Purpose"]
for ci, col in enumerate(rec_cols, 1):
    hdr_cell(ws5, 3, ci, col, C["hdr_dark"], size=9)

recruiter_rows = [
    # General / Logistics
    ("","── GENERAL INFORMATION (Pre-Screen) ──","","","","","",""),
    (1,"Candidate Name","General","Text","Free text","None","Identifier","Confirm identity"),
    (2,"Current Location","General","Text / Dropdown","City, State, Country","None","Candidate profile","Location fit check"),
    (3,"Preferred Location(s)","General","Multi-select","Cities","None","Candidate profile","Can be more than 1"),
    (4,"Current CTC (annual)","General","Number + Currency","e.g. ₹12L, $80K","None","Candidate profile","Compensation alignment"),
    (5,"Expected CTC","General","Number + Currency","e.g. ₹18L","None","Candidate profile","Budget filter"),
    (6,"Preferred Work Mode","General","Dropdown","Remote / Hybrid / Onsite","None","Candidate profile","Mode fit"),
    (7,"Preferred Company Types","General","Multi-select","Product / Service / Startup / MNC","None","Candidate profile","Culture fit"),
    (8,"Reason for Change","General","Text / Dropdown","Growth / Compensation / Location / Culture / Layoff / Other","None","Candidate profile","Motivation signal"),
    (9,"Expected Job Role","General","Text","Free text","None","Candidate profile","Role alignment"),
    (10,"Offer Holding","General","Y/N + Text","Other offer details","None","Candidate profile","Counter-offer risk"),
    (11,"LWD (Last Working Date)","General","Date","MM/YYYY","None","Candidate profile","Joining timeline"),
    (12,"Expected Joining Date","General","Date","MM/YYYY","None","Candidate profile","Notice period"),
    # Missing Info from Resume
    ("","── FILLING MISSING RESUME INFORMATION ──","","","","","",""),
    (13,"Project 1 Clarification","Projects","Text","Recruiter fills if resume is vague: Duration, About project, Role played, Skills used, Domain depth","+/- on Project 1 score","project_1 (E10)","Critical — 8pt param"),
    (14,"Project 2 Clarification","Projects","Text","Duration, About project, Role played, Skills used, Domain depth","+/- on Project 2 score","project_2 (E11)","Important context"),
    (15,"Skill — Years & Timeline","Skills","Table: Skill | From | To | Active","Per validated APPLIED+ skill","+1pt per skill max 6","skill_list_years (S1)","Recruiter confirms timeline per skill"),
    (16,"Mentorship Confirmation","Experience","Y/N + Detail","Mentored/led engineers in how many roles? Team size?","Up to 3pts","mentorship_signal (E6)","Recruiter verifies"),
    (17,"International Exposure","Experience","Y/N + Detail","Onsite details: country, duration, type (client/conference/transfer)","0–2 pts","international_exposure (E7)","Pre-filled by system; recruiter confirms"),
    (18,"Stakeholder Management","Experience","Dropdown","None / Internal / Client-Facing / C-Level","0–2 pts","stakeholder_management (E8)","Recruiter validates from conversation"),
    (19,"LinkedIn Profile Active","Education","Y/N","Profile URL check","0 or 1 pt","linkedin_activity (ED7)","Recruiter checks profile"),
    (20,"Extra-Curricular Activities","Education","Y/N + Detail","Sports / Volunteering / Hackathon / Club","0 or 1 pt","extra_curriculars (ED8)","Resume signal; recruiter confirms"),
    (21,"Communication Skills (pre-screen)","Skills","Rating 1–5","1=Very Poor, 2=Poor, 3=Acceptable, 4=Good, 5=Excellent","Pre-fills panel param","communication_skills (S9)","Proxy from recruiter call"),
    (22,"Project Explanation Quality","Skills","Rating 0–3","0=Cannot explain, 1=Disjointed, 2=Good structure, 3=Clear P→Design→Outcome","0–3 pts","project_explanation (S11)","Recruiter evaluates walk-through"),
    (23,"Skills Learning Acumen","Skills","Y/N + Detail","Has candidate picked up new skills per year? Fast learner?","0–3 pts","skills_learning_acumen (S7)","Recruiter validates growth pattern"),
    # Interview Design
    ("","── INTERVIEW DESIGN / ROUTING ──","","","","","",""),
    (24,"Panel Interview — Proceed?","Admin","Y/N","Panel avail / Candidate willing / Cost OK","None","Routing","Panel interview gating"),
    (25,"Panel Interview Type","Admin","Dropdown","Client-JD Structured / Resume-based Pipeline","None","interview_question_engine","Method 1 vs Method 2"),
    (26,"Suggested Interview Duration","Admin","Dropdown","30 min / 45 min / 60 min","None","Panel routing","Resource planning"),
    (27,"Skill-Based Questions Generated","Questions","System auto-generates","Based on skills + gaps + depth","None","telephonic_question_engine","Recruiter uses for further screening"),
    (28,"Domain Questions Generated","Questions","System auto-generates","Based on domain tags on resume","None","telephonic_question_engine","Domain validation"),
    # Recruiter Notes
    ("","── RECRUITER ASSESSMENT & NOTES ──","","","","","",""),
    (29,"Overall Recruiter Impression","Assessment","Rating 1–5","1=Reject, 2=Weak, 3=Borderline, 4=Good, 5=Excellent","Adjusts score band","recruiter_summary","Holistic gut check"),
    (30,"Recruiter Notes (Free Text)","Assessment","Text","Observations, concerns, highlights","Stored as narrative","recruiter_summary","Stored with candidate record"),
    (31,"Decision Recommendation","Assessment","Dropdown","PROCEED / HOLD / REJECT / ESCALATE","None","pipeline_stage","Routing decision"),
    (32,"Strengths Confirmed","Feedback","Multi-text","List of confirmed strengths","None","feedback_data","Calibration data"),
    (33,"Gaps / Concerns","Feedback","Multi-text","List of gaps or concerns","None","feedback_data","Risk flags"),
    (34,"Corrected Role Family (if wrong)","Feedback","Dropdown","Role family list","Overrides auto","feedback_store","System learning"),
    (35,"Corrected Score (if wrong)","Feedback","Integer","0–100","Overrides auto","feedback_store","System learning"),
    (36,"Corrected Band (if wrong)","Feedback","Dropdown","EXCELLENT/GOOD/AVERAGE/WEAK","Overrides auto","feedback_store","System learning"),
    (37,"Call Outcome","Outcome","Dropdown","INTERESTED / NOT_INTERESTED / NO_SHOW / CALLBACK","None","pipeline_store","Pipeline tracking"),
]

for r_idx, row_data in enumerate(recruiter_rows, 4):
    if row_data[0] == "":
        # section header
        merge_hdr(ws5, r_idx, 1, 8, f"  {row_data[1]}", C["rec_hdr"], size=10)
        continue
    bg = C["rec"]
    for ci, val in enumerate(row_data, 1):
        data_cell(ws5, r_idx, ci, val, bg=bg, bold=(ci==2),
                  halign="center" if ci==1 else "left",
                  wrap=(ci in (4,5,6,7,8)), size=9)

set_col_width(ws5, [4, 30, 15, 16, 38, 20, 25, 28])


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 6 — PANEL STAGE DATA POINTS
# ═════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Panel Stage")
ws6.freeze_panes = "A4"
merge_hdr(ws6, 1, 1, 8, "PANEL INTERVIEW STAGE — ALL DATA POINTS CAPTURED DURING TECHNICAL INTERVIEW", C["panel_hdr"])
merge_hdr(ws6, 2, 1, 8,
    "Panel evaluates candidate on skill depth, domain knowledge, problem-solving, coding, and communication",
    "17375E", fg="DAEEF3")

panel_cols = ["#","Data Point","Category","Who Captures","Input Type","Score Range","Maps To Parameter","Scoring Scale / Notes"]
for ci, col in enumerate(panel_cols, 1):
    hdr_cell(ws6, 3, ci, col, C["hdr_dark"], size=9)

panel_rows = [
    ("","── PRE-PANEL SETUP (Auto-generated) ──","","","","","",""),
    ("A1","Interview Questions Set","Setup","System","Auto-generated (30–40 questions)","N/A","interview_question_engine","Based on resume: skill depth + gaps + domain + progression probes"),
    ("A2","Question Themes","Setup","System","SKILL_DEPTH / STAKEHOLDER / LEADERSHIP / DOMAIN / COMMUNICATION / PROBLEM_SOLVING","N/A","question_scoring_engine","Mapped to rubric parameters"),
    ("A3","Question Type per Skill","Setup","System","telephonic_question_engine output","N/A","telephonic_question_engine","Mandatory-skill focus + gap probes"),
    ("A4","Interview Method","Setup","Recruiter Selects","Method 1 (Client JD) / Method 2 (Resume-based pipeline)","N/A","Panel routing","Determines question set focus"),

    ("","── SKILL DEPTH ASSESSMENT (Per Skill Tested) ──","","","","","",""),
    (1,"Skill Name Tested","Skill Depth","Panel","Text","N/A","skill_depth (S5)","Which skill was tested"),
    (2,"Question Asked","Skill Depth","Panel","Text","N/A","question_scoring_engine","Actual question text"),
    (3,"Candidate Answer (Transcript)","Skill Depth","Panel","Text / Recording","N/A","transcript_scoring_engine","Raw answer text"),
    (4,"Answer Score","Skill Depth","Panel + LLM","0–10","0–10","skill_depth (S5)",
     "0–2=No answer/deflection | 3–4=Vague | 5–6=Adequate (1 example, no depth) | "
     "7–8=Good (specific, ownership, design) | 9–10=Excellent (quantified, ownership, failure/learning)"),
    (5,"Confidence Level","Skill Depth","LLM","HIGH/MEDIUM/LOW","N/A","confidence_score","LLM certainty in score"),
    (6,"What Was Strong","Skill Depth","LLM","Text (1–2 sentences)","N/A","panel_notes","Positive evidence"),
    (7,"What Was Missing","Skill Depth","LLM","Text (1–2 sentences)","N/A","panel_notes","Gap identified"),
    (8,"Follow-up Probe","Skill Depth","LLM auto-generated","Text (<25 words)","N/A","question_set","Next targeted question"),
    (9,"Recruiter Note (hiring brief)","Skill Depth","LLM","Text (1 sentence)","N/A","recruiter_summary","Capture signal for record"),
    (10,"Evidence Cited","Skill Depth","LLM","Specific text claim","N/A","evidence","Supports score"),
    (11,"Deductions Applied","Skill Depth","LLM rule","'We did' without 'I' / No outcome / Buzzwords / Short","Score max cap","skill_depth (S5)","Anti-inflation rules"),
    (12,"Skill Depth — Final Override","Skill Depth","Panel Aggregate","Float","0–8 pts","skill_depth (S5)","avg(all scores for skill_depth) / 10 × 8"),

    ("","── SKILL RECENCY VALIDATION ──","","","","","",""),
    (13,"Skill Recency Confirmed","Skill Recency","Panel","Dropdown per skill","RECENT/MID/OLD/UNKNOWN","skill_recency (S6)","Panel corrects auto-detected recency"),
    (14,"Skill Recency Score Override","Skill Recency","Panel Aggregate","Float","0–6 pts","skill_recency (S6)","avg(recency scores) / 10 × 6"),

    ("","── COMMUNICATION & PRESENTATION SKILLS ──","","","","","",""),
    (15,"Verbal Clarity","Communication","Panel","Rating 1–5","N/A","communication_skills (S9)","Clear sentences, no rambling"),
    (16,"Structure / Logical Flow","Communication","Panel","Rating 1–5","N/A","communication_skills (S9)","Beginning → Middle → End"),
    (17,"Confidence Level","Communication","Panel","Rating 1–5","N/A","communication_skills (S9)","Not nervous, not overconfident"),
    (18,"Audience Adaptability","Communication","Panel","Rating 1–5","N/A","communication_skills (S9)","Adjusts jargon to audience"),
    (19,"Communication Skills — Final Score","Communication","Panel","0–5","0–5 pts","communication_skills (S9)","Panel aggregate"),

    ("","── DOMAIN SKILLS ──","","","","","",""),
    (20,"Domain Area Tested","Domain","Panel","Text","N/A","domain_skills (S10)","e.g. BFSI / ECOMMERCE / HEALTHCARE"),
    (21,"Scenario Questions Asked","Domain","Panel / AI-generated","Text","N/A","domain_skills (S10)","Business domain scenario questions"),
    (22,"Domain Depth Score","Domain","Panel","0–5","N/A","domain_skills (S10)","Business context understanding"),
    (23,"Industry Knowledge Score","Domain","Panel","0–5","N/A","domain_skills (S10)","Domain-specific terminology and context"),
    (24,"Domain Skills — Final Score","Domain","Panel","0–5","0–5 pts","domain_skills (S10)","Panel aggregate"),

    ("","── PROJECT EXPLANATION ──","","","","","",""),
    (25,"Project 1 Walk-through","Project Explanation","Panel","Rating 0–3","N/A","project_explanation (S11)",
     "0=Cannot explain | 1=Disjointed | 2=Good structure | 3=Clear Problem→Solution→Outcome"),
    (26,"Project 1 — Problem Statement Clarity","Project","Panel","Rating 1–3","N/A","project_1 (E10)","Did they articulate the business problem?"),
    (27,"Project 1 — Solution Design Quality","Project","Panel","Rating 1–3","N/A","project_1 (E10)","Did they explain the technical approach?"),
    (28,"Project 1 — Role Played Clarity","Project","Panel","Rating 1–3","N/A","project_1 (E10)","Did they clearly state their ownership?"),
    (29,"Project 1 — Outcome Quantified","Project","Panel","Y/N","N/A","project_1 (E10)","Did they mention measurable outcome?"),
    (30,"Project 1 — Trade-off Awareness","Project","Panel","Y/N","N/A","project_1 (E10)","Did they discuss trade-offs/alternatives?"),
    (31,"Project 2 Walk-through","Project Explanation","Panel","Rating 0–3","N/A","project_2 (E11)","Same as above for Project 2"),

    ("","── CODING SKILLS (Qualitative) ──","","","","","",""),
    (32,"Language Tested","Coding","Panel","Text","N/A","coding_skills (S12)","e.g. Python, SQL"),
    (33,"Problem Level","Coding","Panel","Dropdown","Easy / Medium / Hard","coding_skills (S12)","LeetCode-style difficulty"),
    (34,"Approach Quality","Coding","Panel","Rating 1–5","N/A","coding_skills (S12)","Brute force vs optimal"),
    (35,"Clean Code","Coding","Panel","Y/N","N/A","coding_skills (S12)","Readable, named variables"),
    (36,"Edge Cases Handled","Coding","Panel","Y/N","N/A","coding_skills (S12)","Null, empty, overflow"),
    (37,"Time/Space Complexity Awareness","Coding","Panel","Y/N","N/A","coding_skills (S12)","Big-O awareness"),
    (38,"Coding Skills — Panel Narrative","Coding","Panel","Free text","None (qualitative)","coding_skills (S12)","No numeric score — narrative only"),

    ("","── CONCEPTUAL SKILLS (Qualitative) ──","","","","","",""),
    (39,"Concept Area Tested","Conceptual","Panel","Text","N/A","conceptual_skills (S13)","e.g. Transformer architecture, Spark internals"),
    (40,"Depth of Understanding","Conceptual","Panel","Rating 1–5","N/A","conceptual_skills (S13)","Surface vs deep understanding"),
    (41,"Theory vs Practice Balance","Conceptual","Panel","Rating 1–5","N/A","conceptual_skills (S13)","Can they apply concepts?"),
    (42,"First-Principles Thinking","Conceptual","Panel","Y/N","N/A","conceptual_skills (S13)","Derived from first principles?"),
    (43,"Conceptual Skills — Panel Narrative","Conceptual","Panel","Free text","None (qualitative)","conceptual_skills (S13)","No numeric score — narrative only"),

    ("","── PROBLEM SOLVING SKILLS ──","","","","","",""),
    (44,"Problem Type Given","Problem Solving","Panel / AI-generated","Text","N/A","problem_solving (S14)","Scenario from problem statements database"),
    (45,"Approach Taken","Problem Solving","Panel","Text","N/A","problem_solving (S14)","How they structured the solution"),
    (46,"Structure of Solution","Problem Solving","Panel","Rating 1–3","N/A","problem_solving (S14)","Logical decomposition"),
    (47,"Creativity","Problem Solving","Panel","Y/N","N/A","problem_solving (S14)","Novel approach?"),
    (48,"Time to Solution","Problem Solving","Panel","Minutes","N/A","problem_solving (S14)","Speed signal"),
    (49,"Hints Needed","Problem Solving","Panel","Y/N + Count","N/A","problem_solving (S14)","Independence signal"),
    (50,"Problem Solving — Final Score","Problem Solving","Panel","0–3","0–3 pts","problem_solving (S14)","Panel assigns"),

    ("","── TRANSCRIPT SCORING (Auto Post-Interview) ──","","","","","",""),
    (51,"Resume Skill Claims","Transcript","System","Extracted from resume","N/A","transcript_scoring_engine","Skills with evidence level"),
    (52,"Transcript Skill Evidence","Transcript","System","Extracted from interview transcript","N/A","transcript_scoring_engine","Skills mentioned in interview"),
    (53,"Validated Claims","Transcript","System","Skills upgraded: MENTION/WEAK→APPLIED/DEEP/EXPERT","N/A","transcript_scoring_engine","+2 pts per validated claim"),
    (54,"Weakened Claims","Transcript","System","Skills downgraded: APPLIED/DEEP/EXPERT→MENTION/WEAK","N/A","transcript_scoring_engine","-2 pts per weakened claim"),
    (55,"Score Delta","Transcript","System","Net points change","Float","transcript_scoring_engine","resume_score ± delta = final_score"),
    (56,"Recommendation Shift","Transcript","System","IMPROVED / WEAKENED / UNCHANGED","N/A","transcript_scoring_engine","Signal to hiring manager"),

    ("","── PANEL OUTCOME ──","","","","","",""),
    (57,"Panel Recommendation","Outcome","Panel","Dropdown","STRONG_HIRE / HIRE / MAYBE / NO_HIRE","None","pipeline_store","Final panel decision"),
    (58,"Panel Notes (Free Text)","Outcome","Panel","Text","None","panel_notes","Stored with candidate"),
    (59,"Interview Outcome","Outcome","Panel","Dropdown","SELECTED / REJECTED / ON_HOLD / SECOND_ROUND","None","pipeline_store","Pipeline routing"),
    (60,"Joined","Outcome","HR","Y/N","None","feedback_store","Outcome for ML training data"),
]

for r_idx, row_data in enumerate(panel_rows, 4):
    if row_data[0] == "":
        merge_hdr(ws6, r_idx, 1, 8, f"  {row_data[1]}", C["panel_hdr"], size=10)
        continue
    bg = C["panel"]
    for ci, val in enumerate(row_data, 1):
        data_cell(ws6, r_idx, ci, val, bg=bg, bold=(ci==2),
                  halign="center" if ci in (1,6) else "left",
                  wrap=(ci in (5,7,8)), size=9)

set_col_width(ws6, [4, 32, 16, 16, 18, 12, 24, 48])


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 7 — DNA & CANDIDATE INTELLIGENCE
# ═════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("DNA & Candidate Intelligence")
ws7.freeze_panes = "A3"
merge_hdr(ws7, 1, 1, 6, "DNA, RED FLAGS, ARCHETYPES & CANDIDATE INTELLIGENCE SIGNALS", C["hdr_dark"])

# DNA Table
hdr_cell(ws7, 2, 1, "DNA TYPE", C["hdr_med"], size=9)
hdr_cell(ws7, 2, 2, "DEFINITION", C["hdr_med"], size=9)
hdr_cell(ws7, 2, 3, "KEYWORDS THAT TRIGGER IT", C["hdr_med"], size=9)
hdr_cell(ws7, 2, 4, "ZONE WEIGHT", C["hdr_med"], size=9)
hdr_cell(ws7, 2, 5, "SCORE", C["hdr_med"], size=9)
hdr_cell(ws7, 2, 6, "CONFIDENCE", C["hdr_med"], size=9)

dna_data = [
    ("CONSULTING","Delivery/client/project-based work","client, stakeholder, advisory, consulting, engagement, client delivery, managed client, multiple clients, presented to","title=3x, skills=1.5x, desc=1x","Float","HIGH if ratio≥2.0 vs 2nd"),
    ("PRODUCT","Feature/roadmap/user/growth work","product, feature, roadmap, user, retention, growth, platform, saas, b2c, b2b, launch, product analytics","title=3x, skills=1.5x, desc=1x","Float","MEDIUM if ratio≥1.3"),
    ("PLATFORM_INFRA","Infrastructure/SRE/DevOps","infrastructure, platform engineering, sre, devops, reliability, kubernetes, terraform, on-call, incident, observability","title=3x, skills=1.5x, desc=1x","Float","LOW otherwise"),
    ("DOMAIN_SPECIALIST","Deep domain knowledge","hedge fund, fintech, healthcare, insurance, bfsi, manufacturing, quant, capital markets, supply chain","title=3x, skills=1.5x, desc=1x","Float",""),
    ("RESEARCH","Academic/R&D/publications","research, publications, arxiv, conference, thesis, academic, peer-reviewed, novel, ablation","title=3x, skills=1.5x, desc=1x","Float",""),
    ("HYBRID","Mixed signals — no dominant DNA","Secondary DNA > 30% of primary","—","Flag","When primary + secondary both present"),
]
for ri, row in enumerate(dna_data, 3):
    for ci, val in enumerate(row, 1):
        data_cell(ws7, ri, ci, val, bg=C["bert"], wrap=True, size=9)

# Red Flags
merge_hdr(ws7, 10, 1, 6, "  RED FLAGS — AUTO-DETECTED", C["hdr_dark"], size=10)
rf_cols = ["Flag ID","Flag Name","Detection Logic","Action","Severity","Parameters Affected"]
for ci, col in enumerate(rf_cols, 1):
    hdr_cell(ws7, 11, ci, col, C["hdr_med"], size=9)

red_flags = [
    ("R1","3+ Career Breaks (>3m each)","count_breaks ≥ 3","HARD REJECT FLAG","REJECT",  "career_breaks (E9)"),
    ("R2","Declining Company Tier","TIER_1→TIER_3→TIER_4+ across 3 roles","'Downward trajectory' flag","SOFT","company_tier (E1)"),
    ("R3","Title Inflation","Dir/VP title but team_size=0, no reports","'Title inflation' probe flag","SOFT","career_progression (E3)"),
    ("R4","Buzzword Resume No Impact","proj_count<4 AND awards=0 AND no numbers","'Low impact evidence' flag","SOFT","project_1 (E10), awards (E5)"),
    ("R5","Severe Overqualification","YoE ratio < 0.5 vs JD","'Early exit risk' flag","SOFT","overall_experience (E2)"),
    ("R6","Certification Farming","certs ≥ 5 AND depth ≤ FOUNDATIONAL","'Cert-heavy, low depth' flag","SOFT","certifications (S4), skill_depth (S5)"),
    ("R7","Skill Cluster Mismatch","BERT conf>0.7 AND role_family≠target JD","'Wrong-fit application' flag","SOFT","skill_depth (S5)"),
    ("R8","No Verifiable Output (5+ YoE)","proj_count=0 AND awards=0 AND certs=0","'No deliverables' flag","HARD","project_1 (E10)"),
    ("R9","Frequent Lateral Moves","career_progression=LATERAL in 2+ roles","'Assess growth mindset'","SOFT","career_progression (E3)"),
    ("R10","Job-hopping Pattern","hop_rate>1.5 OR 2+ roles <12m","'Explore context'","SOFT","stability (E4)"),
    ("R11","Large Education Gap","gap > 12 months","REJECT FLAG","REJECT","education_gap (ED3)"),
    ("R12","Overloaded Skills List","len(skills) > 30","'Depth shallow; probe'","SOFT","skill_depth (S5)"),
    ("R13","No Progression 5+ Years","Same title ≥ 5 years same company","'Stagnation risk'","SOFT","career_progression (E3)"),
    ("R14","Domain Switch No Bridge","A8 archetype + depth < HANDS_ON","'Onboarding risk'","SOFT","skill_depth (S5)"),
    ("R15","COVID Gap (2020)","Gap aligned with 2020 recession","'Do NOT penalise' context note","INFO","career_breaks (E9)"),
    ("R16","Identical Project Descriptions","High string similarity across projects","'Probe ownership'","SOFT","project_1 (E10)"),
    ("R17","Relevant Experience <70% of JD","rel_yoe / jd_yoe_min < 0.70","REJECT FLAG","REJECT","overall_experience (E2)"),
]
for ri, row in enumerate(red_flags, 12):
    bg = C["reject"] if row[4]=="REJECT" else C["flag"] if row[4]=="HARD" else C["white"]
    for ci, val in enumerate(row, 1):
        data_cell(ws7, ri, ci, val, bg=bg, bold=(row[4]=="REJECT"), wrap=True, size=9)

# Archetypes
merge_hdr(ws7, 31, 1, 6, "  DYNAMIC WEIGHT ARCHETYPES (10 types)", C["hdr_dark"], size=10)
arch_cols = ["Archetype","Trigger","Education Weight","Experience Weight","Skills Weight","Notes"]
for ci, col in enumerate(arch_cols, 1):
    hdr_cell(ws7, 32, ci, col, C["hdr_med"], size=9)

archetypes = [
    ("A1: Elite College + Strong Company","TIER_1/2 institute + TIER_1/2 employer","15","40","45","Baseline weights"),
    ("A2: Weak College, FAANG","TIER_4/5 institute + TIER_1 employer","5","52","43","Experience dominates"),
    ("A3: Weak College, Mid-Tier","TIER_4/5 + TIER_2/3 employer","8","47","45","Moderate reallocation"),
    ("A4: Elite College, No Brand Company","TIER_1 + TIER_3/4/5 employer","20","37","43","Education differentiates"),
    ("A5: Fresh Graduate (0–1 YoE)","≤1 yr total experience","30","10","60","Skills + Education dominate"),
    ("A6: Senior (10+ YoE)","≥10 years total experience","8","47","45","Track record primary"),
    ("A7: PhD / Research Track","PhD degree or ≥2 publications","25","30","45","Research depth matters"),
    ("A8: Domain Switcher","Education domain ≠ job domain for 2+ roles","5","45","50","Skills prove the switch"),
    ("A9: Founder / Serial Entrepreneur","2+ Founder/CEO titles","10","42","48","Scale + funding signals count"),
    ("A10: Consultant / Contractor","3+ stints <18m, consistent skill cluster","12","40","48","Stability algo adjusted"),
]
for ri, row in enumerate(archetypes, 33):
    for ci, val in enumerate(row, 1):
        data_cell(ws7, ri, ci, val, bg=C["gray_l"], bold=(ci==1), wrap=True, size=9)

set_col_width(ws7, [22, 38, 38, 18, 18, 32])


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 8 — SCORING CALCULATION REFERENCE
# ═════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Scoring Calculation Reference")
ws8.freeze_panes = "A3"
merge_hdr(ws8, 1, 1, 5, "SCORING CALCULATION — DETAILED FORMULAS & THRESHOLDS PER PARAMETER", C["hdr_dark"])

sc_cols = ["Parameter","Max Pts","Formula / Logic","Thresholds / Bands","Reject / Flag Conditions"]
for ci, col in enumerate(sc_cols, 1):
    hdr_cell(ws8, 2, ci, col, C["hdr_med"], size=9)

scoring_calc = [
    # SECTION HEADERS
    ("── EXPERIENCE ──","","","",""),
    ("Companies Worked With (E1)","5",
     "Best tier across all companies. tier1=5, tier2=4, tier3=3, tier4=2, unknown=1",
     "Tier 1: FAANG / Hyper-Scale\nTier 2: Unicorn / top regional\nTier 3: Mid-size funded\nTier 4: IT Services / large SME\nTier 5: Unknown / very small",
     "LLM fallback for unknown companies when ENABLE_COMPANY_TIER_LLM=true"),
    ("Overall Experience (E2)","3",
     "With JD: ratio = min(total_yrs, jd_yoe_max) / jd_yoe_max × 3\nNo JD: band scoring",
     "No JD bands: 10+=3.0 | 6–10=2.5 | 4–6=2.0 | 2–4=1.5 | 1–2=1.0 | <1=0.5",
     "REJECT if relevant_yoe < 70% of JD min range"),
    ("Career Progression (E3)","3",
     "Python title seniority 1–6 → score 0–5\nBERT blend (confidence tiers)\nLLM can override",
     "≥0.60 BERT conf: 60% BERT + 40% heuristic\n0.40–0.59: 50/50\n<0.40: heuristic only",
     "BERT AWARENESS guard clause: if evidence=APPLIED/DEEP/EXPERT, cap BERT conf at 0.3"),
    ("Stability (E4)","3",
     "Avg tenure → 1–5 raw score\nPenalties: <6m stints -0.8×w, <12m -0.4×w\nJob-hopping (>2/yr) -0.5\nUpward title +0.5 bonus",
     "36+m=5.0 | 24–36=4.0 | 18–24=3.5 | 12–18=3.0 | 8–12=2.0 | <8=1.5",
     "Soft flag if hop_rate > 1.5 roles/year"),
    ("Awards & Recognitions (E5)","3",
     "Python count → 0–3\nLLM validates genuine vs marketing",
     "0 awards=0 | 1=1 | 2=2 | 3+=3",
     "LLM filters buzzwords; promotion must be confirmed"),
    ("Mentorship (E6)","3",
     "Python count lead/managed/mentored instances\nBERT: LEAD/FORMAL/IMPLIED/NONE\nLLM final review",
     "Led ≥2 roles=3 | 1 instance=2 | implied=1 | none=0\nBERT ≥0.60 → BERT primary; else heuristic",
     ""),
    ("International Exposure (E7)","2",
     "Python: onsite/global/multi-country/relocation keywords → 2 or 0\nLLM: implied vs explicit",
     "Explicit (onsite/visa/relocation)=2\nImplied (global team/multi-timezone)=1\nNone=0",
     "Pre-filled as 1 if keyword; recruiter must confirm for full 2"),
    ("Stakeholder Management (E8)","2",
     "Python keyword scan + BERT (NONE/INTERNAL/CLIENT_FACING/C_LEVEL)\nLLM final review",
     "C-level=2 | Client-facing=1.5 | Internal=1 | None=0\nBERT ≥0.60 → BERT primary",
     ""),
    ("Career Breaks (E9)","2",
     "Count gaps >3 months between consecutive roles",
     "0 breaks=2 | 1 break=1 | 2 breaks=0 | >2 breaks=0+REJECT",
     "REJECT if >2 breaks. MBA/Maternity overlap with education → no penalty"),
    ("Project 1 (E10)","8",
     "8 criteria × 1pt:\n(1)type known (2)title present (3)desc>20c (4)duration≥3m (5)skills≥1\n(6)domain tag (7)ownership verb+desc>50c (8)quantified impact\n+ LLM PROJECT_JUDGE for complexity_score (0–5)",
     "Ownership verbs: built/led/owned/designed/architected/implemented/optimised/created/deployed/launched/migrated\nImpact: % or number + outcome word (reduced/increased/improved/saved/accelerated)",
     "LLM adds: complexity_score, candidate_signal, implied_skills, green/red flags"),
    ("Project 2 (E11)","6",
     "6 criteria × 1pt (criteria 1–6 only — no depth/impact criteria)",
     "Same criteria as Project 1 except no ownership verb or quantified impact checks",
     ""),

    ("── EDUCATION ──","","","",""),
    ("Institute Tier (ED1)","5",
     "Base: T1=4, T2=3, T3=2, T4=1\nGPA bonus: EXCELLENT→T1+1, T2+0.5\nGOOD→T1+1\nCapped at 5",
     "TIER_1: IIT/IIM/ISI/ISB/IIIT-H/global top-200\nTIER_2: Strong state/NIT/VIT/Manipal\nTIER_3: Mid-tier\nTIER_4: Below average",
     "TIER_1 gets 0.5 base patent credit even with no patents\nTIER_2 gets 0.25 base patent credit"),
    ("Degree Level (ED2)","2",
     "PhD/Master=2, Bachelor=1.5, Diploma=1, Unknown=0.5",
     "Degree level from degree text + course family",
     ""),
    ("Education Gap (ED3)","1",
     "≤6m gap=1, 6–12m=0.5, >12m=0+REJECT",
     "Gap = months between last education end and first job start",
     "REJECT if gap > 12 months"),
    ("Education Relevance (ED4)","2",
     "HIGH=2, MEDIUM=1.5, FOUNDATIONAL=0.5, UNKNOWN=1",
     "HIGH: CS/Eng/MCA/PhD/Research/Analytics/Stats/Quant Econ/OR/Data Science/AI\nMEDIUM: B.Sc/M.Sc/MBA\nFOUNDATIONAL: Arts/Commerce",
     ""),
    ("Executive Education (ED5)","1",
     "Keywords: executive/continuing/distance/certification/online/mooc → 1pt",
     "Boolean",
     ""),
    ("Patents/Publications (ED6)","2",
     "Boolean × 2pts\nTIER_1 base=0.5 even without patents\nTIER_2 base=0.25",
     "Patent OR publication detected → 2pts total",
     ""),
    ("LinkedIn Active (ED7)","1",
     "Recruiter checks profile: 1=active, 0=absent",
     "Boolean",
     ""),
    ("Extra-Curriculars (ED8)","1",
     "Resume or recruiter confirmed: 1=present, 0=none",
     "Boolean",
     ""),

    ("── SKILLS ──","","","",""),
    ("Skill List — Years (S1)","6",
     "Each validated APPLIED+ skill with clear years = 1pt\nclamp(count, 0, 6)",
     "Recruiter validates during call. Only APPLIED/DEEP/EXPERT evidence levels count",
     ""),
    ("Certifications (S4)","3",
     "clamp(cert_count, 0, 3)\nLLM checks validity and relevance",
     "1 cert=1pt | 2=2pts | 3+=3pts",
     "Red flag: ≥5 certs AND depth ≤ FOUNDATIONAL"),
    ("Skill Depth (S5)","8",
     "Top 5 skills: evidence_score → BERT blend\nWith JD: role-weighted\n(avg_blended_score / 5) × 8",
     "Evidence: NONE=0 | MENTION=0.5 | WEAK=1.5 | APPLIED=3.0 | DEEP=4.0 | EXPERT=5.0\nDepth: AWARENESS=0.5 | FOUNDATIONAL=1.5 | HANDS_ON=3.0 | ADVANCED=4.0 | ARCHITECT_LEVEL=5.0",
     "BERT blend tiers: ≥0.65 conf→65%BERT/35%evidence | 0.45–0.64→50/50 | <0.45→evidence only"),
    ("Skill Recency (S6)","6",
     "(count_RECENT_or_CURRENT / total_skills) × 6",
     "RECENT = last 2 years\nMID = 2–5 years\nOLD = >5 years",
     ""),
    ("Skills Learning Acumen (S7)","3",
     "fast_learner (≥2 new/yr for ≥2 yrs)=3\nNew skills ≥3 yrs=2 | 1–2 yrs=1 | none=0",
     "Derived from yearly_skill_learning history",
     ""),
    ("Coding Community (S8)","3",
     "Count OSS platforms: GitHub/Stack Overflow/LeetCode/HackerRank/Kaggle\n≥3=3 | 2=2 | 1=1 | 0=0",
     "Boolean per platform",
     ""),
    ("Communication (S9)","5","Panel assigns 0–5","1=Very Poor | 2=Poor | 3=Acceptable | 4=Good | 5=Excellent","Panel only"),
    ("Domain Skills (S10)","5","Panel assigns 0–5","Scenario-based testing","Panel only"),
    ("Project Explanation (S11)","3","Panel/Recruiter 0–3","0=Cannot explain | 1=Disjointed | 2=Good | 3=Clear P→D→O",""),
    ("Problem Solving (S14)","3","Panel assigns 0–3","Live problem solving","Panel only"),

    ("── STAGE NORMALISATION ──","","","",""),
    ("Resume Stage Score","76 raw → normalised to 100",
     "normalised = (raw_score / 76) × 100",
     "Auto params total to ~76 raw pts",
     "Shown as /100 for recruiter"),
    ("Recruiter Stage Score","87 raw → normalised to 100",
     "normalised = (raw_score / 87) × 100",
     "After recruiter adds up to 11 more pts",
     "Shown as /100 for panel handoff"),
    ("Panel Final Score","100",
     "No normalisation needed — panel total = 100",
     "Full 100-pt rubric completed",
     "This is the final hiring score"),
]

for r_idx, row in enumerate(scoring_calc, 3):
    if row[0].startswith("──"):
        merge_hdr(ws8, r_idx, 1, 5, f"  {row[0]}", C["hdr_dark"], size=10)
        continue
    bg = (C["exp"] if row[0].startswith(("Companies","Overall","Career","Stability","Awards","Mentor","International","Stakeholder","Career Breaks","Project"))
          else C["edu"] if row[0].startswith(("Institute","Degree","Education","Executive","Patent","LinkedIn","Extra"))
          else C["skill"] if row[0].startswith(("Skill","Certif","Coding","Communication","Domain","Project Exp","Problem"))
          else C["gray_l"])
    for ci, val in enumerate(row, 1):
        data_cell(ws8, r_idx, ci, val, bg=bg, bold=(ci==1), wrap=True, size=9)

set_col_width(ws8, [28, 8, 42, 42, 32])


# ═════════════════════════════════════════════════════════════════════════════
# SAVE
# ═════════════════════════════════════════════════════════════════════════════
out_path = "E:/Dev/resume_intelligence/Resume_Intelligence_DataPoints_Complete.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
