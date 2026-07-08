"""
Resume Intelligence — Data Points x Sources x Derived From x Signals
Parameters taken EXACTLY from Scoring.xlsx (Final Scoring - Candidate sheet).
Signals only added where the data point genuinely reveals something about the person.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── helpers ────────────────────────────────────────────────────────────────
def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, color="000000", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, name="Calibri", italic=italic)
def aln(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
thin = Side(style="thin",   color="CCCCCC")
med  = Side(style="medium", color="888888")
B    = Border(left=thin, right=thin, top=thin, bottom=thin)
BM   = Border(left=med,  right=med,  top=med,  bottom=med)

# global ws reference used by helpers
ws = None

def W(cols):
    for i, w in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def cell(r, c, v, bg="FFFFFF", bold=False, fg="000000", sz=9, ha="left", wrap=True, italic=False):
    cl = ws.cell(row=r, column=c, value=v)
    cl.fill = fill(bg); cl.font = fnt(bold, fg, sz, italic)
    cl.alignment = aln(ha, wrap); cl.border = B
    return cl

def sec_hdr(r, label, bg, cols=14):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    cl = ws.cell(row=r, column=1, value=label)
    cl.fill = fill(bg); cl.font = fnt(True, "FFFFFF", 11)
    cl.alignment = aln("left", False); cl.border = BM

# ── color maps ─────────────────────────────────────────────────────────────
# Source tag colors  (bg, fg)
SRC_CLR = {
    "RESUME":    ("D7F0D7", "1B5E20"),
    "LLM":       ("EDE0F5", "4A148C"),
    "BERT":      ("D4E8FB", "0D47A1"),
    "PYTHON":    ("FFF3CD", "6D4C00"),
    "RECRUITER": ("FFE5CC", "BF360C"),
    "PANEL":     ("FBDCE8", "880E4F"),
}

# Signal category colors
SIG_CLR = {
    "LOYALTY":          "FFF9C4",
    "TRAJECTORY":       "E1F5FE",
    "ETHICS":           "EDE7F6",
    "AMBITION":         "FFF3E0",
    "FLIGHT_RISK":      "FFCDD2",
    "GROWTH_MINDSET":   "E8F5E9",
    "OWNERSHIP":        "FFF8E1",
    "BUSINESS_ORIENT":  "F1F8E9",
    "PRESTIGE":         "FCE4EC",
    "TECHNICAL_DEPTH":  "E3F2FD",
    "PASSION_CRAFT":    "F9FBE7",
    "RESILIENCE":       "F3E5F5",
    "INNOVATION":       "FFFDE7",
    "LEADERSHIP":       "E8EAF6",
    "EXEC_PRESENCE":    "EFEBE9",
    "COLLABORATION":    "E0F7FA",
    "GLOBAL_READY":     "F1F8E9",
    "DOMAIN_DEPTH":     "FBE9E7",
    "DNA_FIT":          "F3E5F5",
}

MOD_CLR = {
    "Experience": "1565C0",
    "Education":  "2E7D32",
    "Skills":     "6A1B9A",
}

# ── DATA definition ────────────────────────────────────────────────────────
# Each row tuple:
# (module, parameter, max_pts, sub_data_point,
#  sources[],         derived_from_text,
#  signal_name,       signal_cat,          positive_reading,  negative_reading,
#  score_impact,      priority)
# signal_name=None means no signal for this sub data point

ROWS = []

# ═══════════════════════════════════════════════════════════════════════════
#  EXPERIENCE
# ═══════════════════════════════════════════════════════════════════════════

# ── Companies worked with ──────────────────────────────────────────────────
ROWS += [
("Experience","Companies worked with",5,
 "Company Name",
 ["RESUME"],
 "NER extraction from resume text (company name per employment block)",
 None,None,None,None,"Low – identifier only","H"),

("Experience","Companies worked with",5,
 "Company Tier (1 to 5)",
 ["RESUME","LLM"],
 "Derived → company_tier_taxonomy.py TIER_MAP lookup by company name; LLM fallback for unknowns",
 "PRESTIGE","PRESTIGE",
 "Tier 1/2 = elite environment; strong process bar, deep technical culture",
 "Tier 4/5 (body-shop IT services / unknown) = likely weak product ownership culture",
 "High – primary pts driver","H"),

("Experience","Companies worked with",5,
 "Company Type (Product / Service / Startup / Consulting / Govt)",
 ["RESUME","LLM"],
 "Derived → LLM classifies company type from name + surrounding resume context",
 "DNA_FIT","DNA_FIT",
 "Product company = ownership, roadmap thinking, outcome accountability",
 "Only service/body-shop background = likely lacks product instinct or ownership",
 "High","H"),

("Experience","Companies worked with",5,
 "Start Date",
 ["RESUME"],
 "Direct extraction – date parser on MM/YYYY or YYYY tokens in employment block",
 None,None,None,None,"Medium – used to compute tenure","H"),

("Experience","Companies worked with",5,
 "End Date (or 'Present')",
 ["RESUME"],
 "Direct extraction – 'Present' / 'Current' mapped to today's date by Python",
 None,None,None,None,"Medium – used to compute tenure","H"),

("Experience","Companies worked with",5,
 "Tenure at Company (months)",
 ["RESUME","PYTHON"],
 "Derived → End Date − Start Date arithmetic (Python); computes months",
 "LOYALTY / FLIGHT_RISK","LOYALTY",
 "24+ months = stable; invested deeply before moving on",
 "Multiple stints < 12 months = serial job-hopper; FLIGHT_RISK pattern",
 "High","H"),

("Experience","Companies worked with",5,
 "Domain / Industry of Company",
 ["LLM"],
 "Derived → LLM maps company name to industry vertical (FinTech / HealthTech / E-commerce etc.)",
 "DOMAIN_DEPTH","DOMAIN_DEPTH",
 "Same domain across multiple companies = deep domain specialist",
 "Domain changes every role = generalist; may not fit niche JDs",
 "Medium","M"),

("Experience","Companies worked with",5,
 "Designation / Title at Company",
 ["RESUME"],
 "Direct extraction – job title text from each employment block",
 "TRAJECTORY","TRAJECTORY",
 "Ascending titles (SDE → Senior → Lead → Manager) = healthy upward path",
 "Flat or downgraded title vs previous company = flag for discussion",
 "High","H"),
]

# ── Overall Experience / Relevant Experience ───────────────────────────────
ROWS += [
("Experience","Overall Experience / Relevant Experience",3,
 "Total Years of Experience",
 ["RESUME","PYTHON"],
 "Derived → sum of all company tenure durations (Python date arithmetic across all blocks)",
 None,None,None,None,"High – gates scoring tier","H"),

("Experience","Overall Experience / Relevant Experience",3,
 "Relevant Years of Experience",
 ["RESUME","LLM"],
 "Derived → LLM filters roles/projects that match JD skill tags; Python sums those tenures",
 None,None,None,None,"High","H"),

("Experience","Overall Experience / Relevant Experience",3,
 "Relevance % (Relevant Exp ÷ Total Exp × 100)",
 ["PYTHON","LLM"],
 "Derived → Python computes ratio after LLM identifies relevant roles",
 "DOMAIN_DEPTH","DOMAIN_DEPTH",
 "80 %+ relevance = strong specialist; aligns to senior JDs",
 "< 40 % relevance = broad generalist; verify domain depth via panel",
 "High","H"),

("Experience","Overall Experience / Relevant Experience",3,
 "Current Role Duration (months)",
 ["RESUME","PYTHON"],
 "Derived → Today − Start Date of most recent company block (Python)",
 "LOYALTY","LOYALTY",
 "12–36 months at current role = stable, not stagnant",
 "< 6 months at current job = possible flight risk or bad hire signal",
 "High","H"),
]

# ── Career Progression ────────────────────────────────────────────────────
ROWS += [
("Experience","Career Progression",3,
 "Seniority Level Sequence across all roles",
 ["RESUME","BERT"],
 "Derived → BERT career_progression classifier (DECLINING/LATERAL/GROWING/FAST_TRACK) on sequence of titles",
 "TRAJECTORY","TRAJECTORY",
 "GROWING / FAST_TRACK = clear ascending path; prioritise candidate",
 "DECLINING = downshift in scope or title; investigate before proceeding",
 "High","H"),

("Experience","Career Progression",3,
 "Title Progression Pattern (IC → Lead → Manager → Director)",
 ["RESUME","LLM"],
 "Derived → LLM maps each title to IC / EM / Director tier and plots progression",
 "AMBITION","AMBITION",
 "IC → Lead → Manager track = ambition + leadership readiness",
 "Senior IC with 10+ yrs and no scope growth = possible performance ceiling",
 "High","H"),

("Experience","Career Progression",3,
 "Promoted within same company? (intra-company title change)",
 ["RESUME","LLM"],
 "Derived → LLM detects same employer name across two entries with a title change",
 "LOYALTY + ETHICS","LOYALTY",
 "Internal promotion = employer trusted and invested in candidate; ethics vote of confidence",
 "Long tenure at one company but never promoted = performance or visibility concern",
 "High","H"),

("Experience","Career Progression",3,
 "Lateral vs Vertical moves ratio",
 ["LLM"],
 "Derived → LLM compares consecutive role seniority levels and labels each move",
 "TRAJECTORY","TRAJECTORY",
 "Mostly vertical moves = deliberate growth path",
 "All lateral = stagnation; deliberate pivots need recruiter context",
 "Medium","M"),
]

# ── Stability ─────────────────────────────────────────────────────────────
ROWS += [
("Experience","Stability",3,
 "Average Tenure per Company (months)",
 ["RESUME","PYTHON"],
 "Derived → Python: mean of all company tenure values",
 "LOYALTY","LOYALTY",
 "24+ months average = stable professional; commits before moving",
 "< 12 months average = serial job-hopper; R9 soft flag",
 "High","H"),

("Experience","Stability",3,
 "Number of companies in last 5 years",
 ["RESUME","PYTHON"],
 "Derived → Python: count distinct employers where any date falls within today − 5 years",
 "FLIGHT_RISK","FLIGHT_RISK",
 "1–2 companies in 5 years = stable; likely to stay",
 "4+ companies in 5 years = FLIGHT_RISK; flag before extending offer",
 "High","H"),

("Experience","Stability",3,
 "Longest single stint (months)",
 ["RESUME","PYTHON"],
 "Derived → Python: max() across all company tenure values",
 "LOYALTY","LOYALTY",
 "Longest stint 36+ months = can go deep, builds roots",
 "Longest ever stint < 18 months = never settled anywhere",
 "Medium","M"),

("Experience","Stability",3,
 "Job-hopping pattern (count of stints < 12 months)",
 ["RESUME","PYTHON"],
 "Derived → Python: count stints where End Date − Start Date < 365 days",
 "FLIGHT_RISK","FLIGHT_RISK",
 "Zero short stints = consistent commitment pattern",
 "3+ short stints = pattern of not staying; investigate reasons",
 "High","H"),

("Experience","Stability",3,
 "Reason for leaving each company (if stated)",
 ["RESUME","RECRUITER"],
 "Extracted from resume if explicitly stated; else Recruiter question during screening call",
 "ETHICS / RESILIENCE","ETHICS",
 "Growth / new challenge reasons = healthy ambition signal",
 "Repeated 'conflict with management' = ethics or culture concern; soft red flag",
 "Medium","H"),
]

# ── Awards & Recognitions ─────────────────────────────────────────────────
ROWS += [
("Experience","Awards & Recognitions",3,
 "Award / Recognition Name",
 ["RESUME","LLM"],
 "Derived → LLM entity extraction of award/recognition mentions from all resume sections",
 None,None,None,None,"Low – identifier only","M"),

("Experience","Awards & Recognitions",3,
 "Award Type (External industry / Internal company / Academic)",
 ["LLM"],
 "Derived → LLM classifies each award as internal company award vs external industry recognition",
 "PRESTIGE","PRESTIGE",
 "External industry award (Forbes, ACM, best paper) = peer-recognised excellence",
 "Only internal 'Employee of the Month' = cultural signal, not market-level validation",
 "Medium","M"),

("Experience","Awards & Recognitions",3,
 "Frequency – total count of awards in career",
 ["LLM"],
 "Derived → LLM counts all award/recognition instances mentioned across resume",
 "GROWTH_MINDSET","GROWTH_MINDSET",
 "Multiple awards across career = consistently high performer",
 "Zero awards in 10+ year career = possible concern, verify with panel",
 "Medium","M"),

("Experience","Awards & Recognitions",3,
 "Recency – years since last award",
 ["RESUME","PYTHON"],
 "Derived → LLM extracts date of most recent award; Python computes years since",
 "TRAJECTORY","TRAJECTORY",
 "Award in last 2 years = currently performing at high level",
 "Last award 5+ years ago with no recent recognition = possible decline",
 "Low","L"),
]

# ── Mentorship / Code Reviews / Interviews ────────────────────────────────
ROWS += [
("Experience","Mentorship / Code Reviews / Interviews",3,
 "Mentored junior engineers? (Y/N + count)",
 ["RESUME","LLM"],
 "Derived → LLM detects verbs: 'mentored', 'coached', 'guided', 'onboarded' in resume bullets",
 "LEADERSHIP","LEADERSHIP",
 "Evidence of mentoring = team multiplier; ready for senior / lead roles",
 "No mentoring signal despite 6+ yrs = purely individual contributor; ceiling concern",
 "High","H"),

("Experience","Mentorship / Code Reviews / Interviews",3,
 "Code review participation (regular reviewer?)",
 ["RESUME","RECRUITER"],
 "Extracted from resume if mentioned; Recruiter asks on screening call if not present",
 "ETHICS","ETHICS",
 "Active code reviewer = holds team to quality standards; owns team output, not just own code",
 "No code review culture = siloed work style or low-process environment",
 "Medium","H"),

("Experience","Mentorship / Code Reviews / Interviews",3,
 "Interview panel participation (conducted interviews?)",
 ["RESUME","RECRUITER"],
 "Extracted from resume if stated; Recruiter confirms volume and seniority of interviews conducted",
 "ETHICS","ETHICS",
 "Conducts interviews = trusted by employer as culture and quality gatekeeper",
 "No interview experience despite seniority = trust or visibility concern",
 "Medium","M"),

("Experience","Mentorship / Code Reviews / Interviews",3,
 "Number of people mentored (direct count)",
 ["RECRUITER"],
 "Recruiter question if not stated in resume: 'How many engineers have you mentored formally?'",
 "LEADERSHIP","LEADERSHIP",
 "Mentored 3+ people = sustained leadership investment",
 "Only one or zero mentees = minimal leadership impact",
 "Medium","M"),
]

# ── International Exposure ────────────────────────────────────────────────
ROWS += [
("Experience","International Exposure",2,
 "Countries / regions worked in",
 ["RESUME","LLM"],
 "Derived → LLM entity extraction of country/city names from work location fields in resume",
 "GLOBAL_READY","GLOBAL_READY",
 "Worked in 2+ countries = adaptable, genuine cross-cultural experience",
 "Only domestic experience for a global team lead role = significant gap",
 "Medium","M"),

("Experience","International Exposure",2,
 "Duration of international assignment (months)",
 ["RESUME","PYTHON"],
 "Derived → Python computes tenure of roles where location is a foreign country",
 "GLOBAL_READY","GLOBAL_READY",
 "12+ months international = substantive cross-cultural working experience",
 "< 3 months = project visit or conference, not substantive immersion",
 "Medium","M"),

("Experience","International Exposure",2,
 "International client-facing work?",
 ["RESUME","LLM"],
 "Derived → LLM detects international client names or cross-border delivery context in resume",
 "COLLABORATION","COLLABORATION",
 "Managed international clients = communication + delivery across cultures",
 "Only back-office / support with no client interaction",
 "Medium","M"),

("Experience","International Exposure",2,
 "Remote cross-border collaboration (time-zone teams)",
 ["RECRUITER"],
 "Recruiter question: 'Have you worked with teams in different time zones or countries?'",
 None,None,None,None,"Low","L"),
]

# ── Stakeholder Management ────────────────────────────────────────────────
ROWS += [
("Experience","Stakeholder Management",2,
 "Highest stakeholder level engaged (C-level / VP / Manager / None)",
 ["RESUME","BERT"],
 "Derived → BERT stakeholder_management classifier (NONE / INTERNAL / CLIENT_FACING / C_LEVEL) on resume text",
 "EXEC_PRESENCE","EXEC_PRESENCE",
 "C-level / VP interaction verified = executive-presence, can handle senior relationships",
 "Only manager-level: not exposed to business strategy conversations",
 "High","H"),

("Experience","Stakeholder Management",2,
 "Business communication evidence in resume",
 ["RESUME","LLM"],
 "Derived → LLM detects phrases: 'presented to CXO', 'aligned with board', 'briefed the CTO'",
 "EXEC_PRESENCE","EXEC_PRESENCE",
 "Clear business-level language in resume = can draft exec decks and drive decisions",
 "All bullets are technical with zero business language = limited exec exposure",
 "High","H"),

("Experience","Stakeholder Management",2,
 "External client-facing experience",
 ["RESUME","LLM"],
 "Derived → LLM identifies external client names or consulting / delivery context",
 "COLLABORATION","COLLABORATION",
 "Direct client management = delivery accountability beyond internal teams",
 "No external client experience = internal-only; may struggle in client-facing roles",
 "Medium","M"),

("Experience","Stakeholder Management",2,
 "Cross-functional collaboration (Product / Biz / Engg / Finance)",
 ["RESUME","LLM"],
 "Derived → LLM detects cross-team mentions ('worked with PM', 'aligned with finance', 'partnered with ops')",
 "COLLABORATION","COLLABORATION",
 "Cross-functional = breaks silos, business-aligned thinking",
 "Purely within engineering team = may lack business context for senior roles",
 "Medium","M"),
]

# ── Career Breaks ─────────────────────────────────────────────────────────
ROWS += [
("Experience","Career Breaks",2,
 "Break duration (months per gap)",
 ["RESUME","PYTHON"],
 "Derived → Python detects gap between consecutive employment End Date and next Start Date > 90 days",
 None,None,None,None,"Medium – gating check","H"),

("Experience","Career Breaks",2,
 "Number of breaks in career",
 ["RESUME","PYTHON"],
 "Derived → Python count of date gaps > 3 months between any two consecutive employment blocks",
 None,None,None,None,"Medium","H"),

("Experience","Career Breaks",2,
 "Reason for break (stated or recruiter-confirmed)",
 ["RESUME","RECRUITER"],
 "Extracted from resume if stated; Recruiter question if gap > 3 months and no explanation found",
 "RESILIENCE","RESILIENCE",
 "Valid reason (health / family / upskilling) with clear re-entry = resilient",
 "Unexplained 12+ month break = concern; needs recruiter follow-up before proceeding",
 "High","H"),

("Experience","Career Breaks",2,
 "Activity during break (courses / freelance / projects)",
 ["RESUME","LLM"],
 "Derived → LLM checks if resume mentions courses, freelance projects, or personal builds during gap period",
 "GROWTH_MINDSET","GROWTH_MINDSET",
 "Productive break with visible output = growth mindset even in adversity",
 "Completely inactive break with no explanation = passive response to adversity",
 "Medium","M"),
]

# ── Project 1 ─────────────────────────────────────────────────────────────
ROWS += [
("Experience","Project 1 - Latest Project",8,
 "Project Duration (months)",
 ["RESUME","PYTHON"],
 "Direct extraction → date range from project header; Python computes months",
 None,None,None,None,"Medium","H"),

("Experience","Project 1 - Latest Project",8,
 "About the Project (business problem being solved)",
 ["RESUME","LLM"],
 "Derived → LLM extracts the business problem statement from project description bullets",
 "DOMAIN_DEPTH","DOMAIN_DEPTH",
 "Clear real-world business problem framing = domain awareness and contextual thinking",
 "Vague or purely technical description = candidate may not understand their own project",
 "High","H"),

("Experience","Project 1 - Latest Project",8,
 "Role Played (individual contribution vs team lead)",
 ["RESUME","LLM"],
 "Derived → LLM extracts candidate's specific action verbs (built, led, owned, designed, supported)",
 "OWNERSHIP","OWNERSHIP",
 "'Led', 'owned', 'designed' verbs = clear personal accountability and delivery",
 "Only 'assisted', 'contributed', 'supported' = peripheral involvement, not ownership",
 "High","H"),

("Experience","Project 1 - Latest Project",8,
 "Skills Used (list of technologies in project context)",
 ["RESUME","LLM"],
 "Derived → LLM + NER extracts skill names mentioned within the project description (not just skill section)",
 None,None,None,None,"High – feeds skill section","H"),

("Experience","Project 1 - Latest Project",8,
 "Skill Depth in Project (AWARENESS to ARCHITECT_LEVEL)",
 ["RESUME","BERT"],
 "Derived → BERT skill_depth classifier on skill-usage sentences in project block → AWARENESS / FOUNDATIONAL / HANDS_ON / ADVANCED / ARCHITECT_LEVEL",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "ADVANCED / ARCHITECT_LEVEL = deep hands-on mastery demonstrated in project context",
 "AWARENESS / FOUNDATIONAL = listed in project but no substantive usage evidence",
 "High","H"),

("Experience","Project 1 - Latest Project",8,
 "Domain Depth (industry knowledge shown in project)",
 ["RESUME","LLM"],
 "Derived → LLM judges domain-specific vocabulary, concepts, and problem framing in project description",
 "DOMAIN_DEPTH","DOMAIN_DEPTH",
 "Domain-fluent description = genuine domain expertise, not just copied jargon",
 "Generic / tech-only description with zero domain context",
 "High","H"),

("Experience","Project 1 - Latest Project",8,
 "Quantified Business Impact (metrics / numbers / outcomes)",
 ["RESUME","LLM"],
 "Derived → LLM scans for numbers, %, $ values, latency reductions, revenue impact in bullets",
 "BUSINESS_ORIENT","BUSINESS_ORIENT",
 "Quantified outcomes = thinks in business results, not just code shipped",
 "All bullets are technical activities with zero business outcome = pure executor",
 "High","H"),

("Experience","Project 1 - Latest Project",8,
 "DevOps / Advanced Engineering (CI/CD, cloud, scale)",
 ["RESUME","LLM"],
 "Derived → LLM detects deployment, CI/CD, Docker, K8s, cloud infra mentions in project",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "Production engineering evidence = full-stack mindset, delivers beyond notebooks",
 "No production / deployment mention = research or POC only; may not ship to prod",
 "Medium","M"),

("Experience","Project 1 - Latest Project",8,
 "Team Size and Ownership Level",
 ["RESUME","RECRUITER"],
 "Extracted if stated in resume; Recruiter question if not mentioned",
 "OWNERSHIP","OWNERSHIP",
 "Led team of 3+ or sole owner = high accountability and delivery track",
 "Part of 50-person team with undefined personal role = weak signal",
 "High","H"),

("Experience","Project 1 - Latest Project",8,
 "Project Type (Research / Product Feature / Platform / POC)",
 ["RESUME","BERT","LLM"],
 "Derived → BERT project_type classifier on project description; LLM fallback for novel types",
 "DNA_FIT","DNA_FIT",
 "Project type matches JD DNA (product project for product role) = natural fit",
 "All POC / research projects for a production engineering role = mismatch",
 "High","H"),
]

# ── Project 2 ─────────────────────────────────────────────────────────────
ROWS += [
("Experience","Project 2 - 2nd Latest Project",6,
 "Project Duration (months)",
 ["RESUME","PYTHON"],
 "Direct extraction → date range from project 2 header; Python computes months",
 None,None,None,None,"Medium","H"),

("Experience","Project 2 - 2nd Latest Project",6,
 "About the Project (business problem)",
 ["RESUME","LLM"],
 "Derived → LLM extracts business problem from 2nd project description bullets",
 "DOMAIN_DEPTH","DOMAIN_DEPTH",
 "Clear business context in 2nd project = pattern of domain understanding, not a one-off",
 "Two consecutive vague projects = may be unable to articulate own work",
 "High","H"),

("Experience","Project 2 - 2nd Latest Project",6,
 "Role Played",
 ["RESUME","LLM"],
 "Derived → LLM extracts ownership verbs from 2nd project description",
 "OWNERSHIP","OWNERSHIP",
 "Lead / owner role in both projects = consistent ownership pattern",
 "Peripheral in both projects = follower profile; needs structure to deliver",
 "High","H"),

("Experience","Project 2 - 2nd Latest Project",6,
 "Skills Used (in project context)",
 ["RESUME","LLM"],
 "Derived → LLM + NER extracts skill names in 2nd project block only",
 None,None,None,None,"High – feeds skill section","H"),

("Experience","Project 2 - 2nd Latest Project",6,
 "Skill Depth in Project",
 ["RESUME","BERT"],
 "Derived → BERT skill_depth on skill sentences in 2nd project block",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "Consistently deep across both projects = reliable expertise, not one-time event",
 "High in project 1, low in project 2 = recency or breadth gap",
 "High","H"),

("Experience","Project 2 - 2nd Latest Project",6,
 "Domain Depth",
 ["RESUME","LLM"],
 "Derived → LLM domain vocabulary scoring on 2nd project description",
 "DOMAIN_DEPTH","DOMAIN_DEPTH",
 "Same or adjacent domain in project 2 = confirmed specialist pattern",
 "Different domain in both projects = domain switcher; generalist",
 "Medium","M"),

("Experience","Project 2 - 2nd Latest Project",6,
 "Quantified Business Impact",
 ["RESUME","LLM"],
 "Derived → LLM scans for metrics / numbers in 2nd project bullets",
 "BUSINESS_ORIENT","BUSINESS_ORIENT",
 "Metrics present in both projects = consistent outcome-focus habit",
 "Both projects lack metrics = activity-only thinker, not results-oriented",
 "High","H"),

("Experience","Project 2 - 2nd Latest Project",6,
 "Team Size and Ownership Level",
 ["RESUME","RECRUITER"],
 "Extracted from resume; Recruiter fills if missing during screening",
 "OWNERSHIP","OWNERSHIP",
 "Owned both projects = natural owner; ready for lead roles",
 "No ownership in either project = needs significant mentoring to lead",
 "Medium","M"),

("Experience","Project 2 - 2nd Latest Project",6,
 "Project Type",
 ["RESUME","BERT","LLM"],
 "Derived → BERT project_type on 2nd project block; LLM fallback",
 "DNA_FIT","DNA_FIT",
 "Consistent project type across both = confirmed DNA alignment",
 "Scattered project types = unfocused background; unclear DNA",
 "Medium","M"),
]

# ═══════════════════════════════════════════════════════════════════════════
#  EDUCATION
# ═══════════════════════════════════════════════════════════════════════════

# ── Institutes – Tier, GPA, Stream ────────────────────────────────────────
ROWS += [
("Education","Institutes - Tier, GPA, Stream",5,
 "Institute Name",
 ["RESUME"],
 "Direct extraction → NER on education section of resume",
 None,None,None,None,"Low – used for lookup","H"),

("Education","Institutes - Tier, GPA, Stream",5,
 "Institute Tier (Tier-1 / Tier-2 / Tier-3 / Private / International)",
 ["RESUME","PYTHON"],
 "Derived → education_engine.py: INSTITUTE_DICTIONARY lookup on institute name → taxonomy.py tier mapping",
 "PRESTIGE","PRESTIGE",
 "IIT / IIM / ISI / Top-10 intl = elite academic pedigree; proven intellectual rigor",
 "Tier-3 / unknown private college = lower prior; cross-validate via panel performance",
 "High","H"),

("Education","Institutes - Tier, GPA, Stream",5,
 "GPA / Percentage / CGPA",
 ["RESUME"],
 "Direct extraction → regex on patterns like '8.5/10', '85%', 'CGPA 3.8' in education block",
 None,None,None,None,"Medium – score gating","H"),

("Education","Institutes - Tier, GPA, Stream",5,
 "Stream / Branch (CS / ECE / Statistics / MBA etc.)",
 ["RESUME","PYTHON"],
 "Derived → taxonomy.py COURSE_DICTIONARY lookup → maps to IT / ANALYTICS / MANAGEMENT / DOMAIN categories",
 "DOMAIN_DEPTH","DOMAIN_DEPTH",
 "CS / Stats / Analytics degree = strong theoretical foundation for data/ML roles",
 "Non-technical degree with no subsequent tech upskilling = foundational gap flag",
 "High","H"),
]

# ── Highest Education ─────────────────────────────────────────────────────
ROWS += [
("Education","Highest Education (Bachelors / Masters / PhD) and Stream",2,
 "Degree Level (UG / PG / PhD)",
 ["RESUME"],
 "Direct extraction → keyword detection: B.Tech, M.Tech, MBA, PhD, MS, BE, ME etc.",
 None,None,None,None,"High – gates bonus pts","H"),

("Education","Highest Education (Bachelors / Masters / PhD) and Stream",2,
 "Degree Field / Specialisation",
 ["RESUME","LLM"],
 "Derived → LLM maps degree name to tech-relevant / management / unrelated category",
 "DOMAIN_DEPTH","DOMAIN_DEPTH",
 "Relevant specialisation = aligned foundational knowledge for role",
 "Unrelated specialisation (arts, law) without any tech upskilling = foundational gap",
 "High","H"),

("Education","Highest Education (Bachelors / Masters / PhD) and Stream",2,
 "Degree Relevance to Target Role",
 ["RESUME","LLM"],
 "Derived → LLM judges overlap between degree field and JD required skills",
 None,None,None,None,"Medium","M"),
]

# ── Education Gaps ────────────────────────────────────────────────────────
ROWS += [
("Education","Education Gaps",1,
 "Gap duration between degrees / institutions (months)",
 ["RESUME","PYTHON"],
 "Derived → Python computes difference between end year of one degree and start year of next institution",
 None,None,None,None,"Low","M"),

("Education","Education Gaps",1,
 "Reason for education gap",
 ["RECRUITER"],
 "Recruiter question triggered when Python detects gap > 12 months in education history",
 "RESILIENCE","RESILIENCE",
 "Valid reason with productive activity during gap = shows character and self-direction",
 "Unexplained 2+ year education gap = concern; recruiter must resolve before proceeding",
 "Medium","M"),
]

# ── Education to Job Relevance ────────────────────────────────────────────
ROWS += [
("Education","Education to Job Relevance",2,
 "Field-of-study vs JD skill match",
 ["RESUME","LLM"],
 "Derived → LLM compares degree taxonomy category against JD mandatory and preferred skills",
 "DOMAIN_DEPTH","DOMAIN_DEPTH",
 "Strong match = academic foundation directly supports job requirements",
 "Mismatch = self-taught path; verify via certifications, projects, panel",
 "High","H"),

("Education","Education to Job Relevance",2,
 "Academic thesis / capstone project relevance to role",
 ["RESUME","LLM"],
 "Derived → LLM scans thesis / capstone / academic project titles for skill overlap with JD",
 None,None,None,None,"Medium","M"),
]

# ── Executive Education / Distant Learning ────────────────────────────────
ROWS += [
("Education","Executive Education / Distant Learning",1,
 "Course Name",
 ["RESUME"],
 "Direct extraction → text from certifications or courses section (outside formal degree blocks)",
 None,None,None,None,"Low – identifier","M"),

("Education","Executive Education / Distant Learning",1,
 "Provider / Institution (IIM / Coursera / Udemy / Stanford Online etc.)",
 ["RESUME","LLM"],
 "Derived → LLM classifies provider tier (IIM Executive / Stanford = premium vs Udemy = self-study)",
 "GROWTH_MINDSET","GROWTH_MINDSET",
 "IIM / Wharton / Stanford Executive = invests in premium, structured upskilling",
 "Only Udemy / YouTube self-study with no structured programs = less committed to formal growth",
 "Medium","M"),

("Education","Executive Education / Distant Learning",1,
 "Course relevance to current / target role",
 ["RESUME","LLM"],
 "Derived → LLM cross-checks course topic against JD skill taxonomy",
 "GROWTH_MINDSET","GROWTH_MINDSET",
 "Relevant course in last 2 years = proactively closing skill gaps",
 "Random unrelated courses = scatter-shot learning, no clear upskilling intent",
 "Medium","M"),

("Education","Executive Education / Distant Learning",1,
 "Recency of last course (years ago)",
 ["RESUME","PYTHON"],
 "Derived → Python computes years since most recent course / executive education date in resume",
 "GROWTH_MINDSET","GROWTH_MINDSET",
 "Course completed in last 1–2 years = actively learning alongside work",
 "Last course 5+ years ago = may have stopped investing in own skills",
 "Medium","M"),
]

# ── Patents / Publications ────────────────────────────────────────────────
ROWS += [
("Education","Patents / Publications",2,
 "Patent / Publication Title",
 ["RESUME"],
 "Direct extraction → LLM entity extraction from patents or publications section",
 None,None,None,None,"Low – identifier","M"),

("Education","Patents / Publications",2,
 "Status (Filed / Granted / Published / Under Review)",
 ["RESUME"],
 "Direct extraction → keyword detection: 'granted', 'filed', 'published', arXiv link etc.",
 "INNOVATION","INNOVATION",
 "Granted patent or published peer-reviewed paper = idea taken to formal completion",
 "Only 'filed' with no grants or publications = aspirational, not yet substantive",
 "Medium","M"),

("Education","Patents / Publications",2,
 "Domain relevance to role",
 ["RESUME","LLM"],
 "Derived → LLM maps patent or paper domain to JD skill taxonomy",
 "INNOVATION","INNOVATION",
 "Patent directly in the skill domain = thought leader and practitioner in this area",
 "Patent in unrelated area = innovation signal but not role-relevant",
 "Medium","M"),

("Education","Patents / Publications",2,
 "Sole author vs co-authored",
 ["RESUME"],
 "Direct extraction → author list parsing: solo author vs multiple contributors",
 "COLLABORATION","COLLABORATION",
 "Co-authored with diverse teams = builds on others' thinking, collaborative innovator",
 "Always sole author despite team environment = may prefer to work in isolation",
 "Low","L"),
]

# ── LinkedIn / Professional Social Media ──────────────────────────────────
ROWS += [
("Education","LinkedIn / Professional Social Media Activeness",1,
 "Profile exists? (Y / N)",
 ["RECRUITER"],
 "Recruiter checks for LinkedIn URL in resume header; manual search if missing",
 None,None,None,None,"Medium – gate","H"),

("Education","LinkedIn / Professional Social Media Activeness",1,
 "Activity level band (Not present / Less active / More active / Regular)",
 ["RECRUITER"],
 "Recruiter observation during screening; maps to 4-band scoring: Not present=0, Less=0.25, More=0.5, Regular=1.0",
 "GROWTH_MINDSET","GROWTH_MINDSET",
 "Regular poster / updater = engaged in industry, professional presence, thought visibility",
 "Not on LinkedIn or zero activity in 2+ years = low professional branding for a senior role",
 "Medium","M"),

("Education","LinkedIn / Professional Social Media Activeness",1,
 "Content published (articles / posts on technical topics)",
 ["RECRUITER"],
 "Recruiter reads recent posts or articles during screening call research",
 "INNOVATION","INNOVATION",
 "Technical articles or insight posts = visible expert; thought leadership signal",
 "No content or only reshares = passive presence; not contributing ideas",
 "Low","L"),
]

# ── Extra Curricular Activities ───────────────────────────────────────────
ROWS += [
("Education","Extra Curricular Activities",1,
 "Activity type (Sports / Community / Tech club / Hackathon)",
 ["RESUME","RECRUITER"],
 "Direct extraction from resume extra-curricular section; Recruiter follows up if missing",
 None,None,None,None,"Low","L"),

("Education","Extra Curricular Activities",1,
 "Leadership role in activity (Captain / President / Lead Organiser)",
 ["RESUME","LLM"],
 "Derived → LLM detects leadership title within activity description",
 "LEADERSHIP","LEADERSHIP",
 "Led a club or team = early leadership instinct; consistent pattern with career",
 "Participant only across all activities = limited early leadership signal",
 "Low","L"),

("Education","Extra Curricular Activities",1,
 "Tech relevance (Hackathon / OSS community / Research group)",
 ["RESUME","LLM"],
 "Derived → LLM judges if activity type is tech-adjacent or domain-relevant",
 "PASSION_CRAFT","PASSION_CRAFT",
 "Tech hackathons / open source / research groups = passion that extends beyond work hours",
 "No tech-adjacent activities = 9-to-5 mindset; lower signal for senior tech roles",
 "Low","L"),
]

# ── Coding Platforms / Community Contributions ────────────────────────────
ROWS += [
("Education","Coding Platforms / Community Contributions",3,
 "GitHub profile link",
 ["RESUME"],
 "Direct extraction → URL pattern matching for github.com in resume header or body",
 None,None,None,None,"Low – enables further check","H"),

("Education","Coding Platforms / Community Contributions",3,
 "GitHub activity (repos / stars / contribution graph)",
 ["RESUME","LLM"],
 "Derived → LLM assesses GitHub profile or candidate states contribution count in resume",
 "PASSION_CRAFT","PASSION_CRAFT",
 "Active public repos + consistent contribution graph = builds beyond work hours",
 "Empty or fully private GitHub = low community and craft signal",
 "High","H"),

("Education","Coding Platforms / Community Contributions",3,
 "LeetCode / HackerRank – rank or problems solved",
 ["RESUME"],
 "Direct extraction → mention of platform rating, streak, or problem count in resume",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "LeetCode 200+ Hard / HackerRank Gold = strong DSA depth; disciplined practice",
 "Not on any competitive platform = may lack algorithmic thinking habit",
 "Medium","M"),

("Education","Coding Platforms / Community Contributions",3,
 "Open source contributions (PRs / issues to external repos)",
 ["RESUME","LLM"],
 "Derived → GitHub URL + LLM checks if contributions are to external / popular repos (not just personal)",
 "PASSION_CRAFT","PASSION_CRAFT",
 "Merged PRs to popular OSS projects = high craft standards; peer-reviewed code",
 "No open source contribution for senior engineers = missed community involvement",
 "High","H"),

("Education","Coding Platforms / Community Contributions",3,
 "Kaggle rank / competition results",
 ["RESUME"],
 "Direct extraction → kaggle.com URL or competition ranking / medal mention in resume",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "Kaggle Master / Grand Master = world-class ML competition performance",
 "Not on Kaggle (fine for non-ML roles; flag for DS / ML JDs)",
 "Medium","M"),
]

# ═══════════════════════════════════════════════════════════════════════════
#  SKILLS
# ═══════════════════════════════════════════════════════════════════════════

# ── Skill List – Years – Timeline ─────────────────────────────────────────
ROWS += [
("Skills","Skill List - Years of Experience - Timeline",6,
 "Skill Name",
 ["RESUME"],
 "Derived → NER + regex skill extractor against internal skill taxonomy dictionary",
 None,None,None,None,"High – all skill scoring depends on this","H"),

("Skills","Skill List - Years of Experience - Timeline",6,
 "Years of experience per skill",
 ["RESUME","PYTHON"],
 "Derived → Python aggregates tenure of all roles and projects that mention this skill",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "8+ years on primary JD skill = deep practitioner; high confidence score",
 "Primary mandatory JD skill with < 1 yr experience = hard mismatch",
 "High","H"),

("Skills","Skill List - Years of Experience - Timeline",6,
 "Skill timeline (first used year → last used year)",
 ["RESUME","PYTHON"],
 "Derived → Python finds earliest and latest role dates where this skill is mentioned",
 "GROWTH_MINDSET","GROWTH_MINDSET",
 "Skill still used in current role = active and current; no recency penalty",
 "Skill last used 4+ years ago = stale; risk for fast-moving tech (LLMs, cloud)",
 "High","H"),

("Skills","Skill List - Years of Experience - Timeline",6,
 "Mandatory skill present? (Y / N per JD)",
 ["RESUME","LLM"],
 "Derived → LLM cross-checks extracted skill list against JD mandatory skills list",
 None,None,None,None,"High – hard gate for scoring","H"),

("Skills","Skill List - Years of Experience - Timeline",6,
 "Good-to-have skill present? (Y / N per JD)",
 ["RESUME","LLM"],
 "Derived → LLM cross-checks extracted skills against JD nice-to-have skills list",
 None,None,None,None,"Medium – bonus pts","M"),
]

# ── Certifications ────────────────────────────────────────────────────────
ROWS += [
("Skills","Certifications - Validity and Type",3,
 "Certification Name",
 ["RESUME"],
 "Direct extraction → text from certifications section of resume",
 None,None,None,None,"Low – identifier","H"),

("Skills","Certifications - Validity and Type",3,
 "Issuing body (AWS / Google / Microsoft / PMI etc.)",
 ["RESUME","LLM"],
 "Derived → LLM maps certification name to known issuing body and market tier",
 "PRESTIGE","PRESTIGE",
 "AWS Professional / GCP Pro / Azure Expert = platform-certified; high credibility",
 "Unknown or self-issued certification = low market signal",
 "Medium","M"),

("Skills","Certifications - Validity and Type",3,
 "Validity / expiry status",
 ["RESUME","PYTHON"],
 "Derived → Python computes years since certification date; flags if > 3 years (most certs expire)",
 "GROWTH_MINDSET","GROWTH_MINDSET",
 "Recently renewed or current cert = actively maintaining skill currency",
 "Expired cert 5+ years old without renewal = may have lapsed on skill",
 "Medium","M"),

("Skills","Certifications - Validity and Type",3,
 "Skill relevance to JD",
 ["RESUME","LLM"],
 "Derived → LLM judges if certification topic maps to JD required skills",
 None,None,None,None,"High – score driver","H"),

("Skills","Certifications - Validity and Type",3,
 "Certification level (Associate / Professional / Expert / Specialty)",
 ["RESUME","LLM"],
 "Derived → LLM extracts level keyword from certification full name",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "Professional / Expert level = validated advanced capability",
 "Only Associate-level certs for a senior role = foundational, not advanced",
 "Medium","M"),
]

# ── Skill Depth ───────────────────────────────────────────────────────────
ROWS += [
("Skills","Skill Depth",8,
 "BERT depth label per skill (AWARENESS to ARCHITECT_LEVEL)",
 ["BERT"],
 "Derived → BERT skill_depth classifier on skill-usage sentences → AWARENESS / FOUNDATIONAL / HANDS_ON / ADVANCED / ARCHITECT_LEVEL",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "ADVANCED / ARCHITECT_LEVEL = can design systems, mentor others, solve novel problems",
 "AWARENESS / FOUNDATIONAL on mandatory JD skill = hard rejection signal",
 "High","H"),

("Skills","Skill Depth",8,
 "Evidence level per skill (NONE to EXPERT)",
 ["BERT","LLM"],
 "Derived → mapped from depth label → NONE / MENTION / WEAK / APPLIED / DEEP / EXPERT scale in rubric_engine.py",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "DEEP / EXPERT = multiple project uses + quantified outcomes",
 "MENTION / WEAK = listed in skills section only; zero project evidence",
 "High","H"),

("Skills","Skill Depth",8,
 "Panel depth Q&A (conceptual + applied questions)",
 ["PANEL"],
 "Panel asks conceptual and applied questions per skill; LLM scores response quality against rubric",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "Correct, confident, deep answers under pressure = verified depth",
 "Shallow or incorrect answers = claimed depth exceeds actual depth",
 "High","H"),

("Skills","Skill Depth",8,
 "Architecture-level thinking shown (system design, trade-offs)",
 ["PANEL"],
 "Panel observes: can candidate design a system, pick trade-offs, reason at scale?",
 "INNOVATION","INNOVATION",
 "System-design thinking with trade-off awareness = ARCHITECT_LEVEL confirmed",
 "Knows usage syntax but cannot design or reason about systems = HANDS_ON ceiling",
 "High","H"),
]

# ── Skill Recency ─────────────────────────────────────────────────────────
ROWS += [
("Skills","Skill Recency",6,
 "Last used date per skill",
 ["RESUME","PYTHON"],
 "Derived → Python finds latest role / project end date where this skill appears in text",
 "GROWTH_MINDSET","GROWTH_MINDSET",
 "Still used at current or last role = active and current; zero recency penalty",
 "Last used 4+ years ago = stale; high risk for fast-moving tech domains",
 "High","H"),

("Skills","Skill Recency",6,
 "Skill present in current role description?",
 ["RESUME","LLM"],
 "Derived → LLM checks if current role bullets or project descriptions mention this skill",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "Skill active in current role = no recency score deduction",
 "Not used in current role = recency penalty applied; panel must verify",
 "High","H"),

("Skills","Skill Recency",6,
 "Skills active in last 2 years (count)",
 ["RESUME","PYTHON"],
 "Derived → Python: count of skills where last-used date > today − 730 days",
 None,None,None,None,"High – score gate","H"),

("Skills","Skill Recency",6,
 "Panel recency assessment (recent scenario questions)",
 ["PANEL"],
 "Panel asks: 'Tell me about the last time you used X in production. What has changed in X since then?'",
 "GROWTH_MINDSET","GROWTH_MINDSET",
 "Cites very recent, current-version usage = truly active practitioner",
 "Describes usage from 3+ years ago despite listing it as current = stale claim",
 "High","H"),
]

# ── Skills Learning Acumen ────────────────────────────────────────────────
ROWS += [
("Skills","Skills Learning Acumen",3,
 "New skills added per year (learning velocity)",
 ["RESUME","LLM"],
 "Derived → LLM counts distinct new skills appearing per chronological role block across career",
 "GROWTH_MINDSET","GROWTH_MINDSET",
 "2+ new skills per year = high learning velocity; adapts to new tech demands",
 "Same skill set for 5+ years with zero new additions = stagnant learner",
 "High","H"),

("Skills","Skills Learning Acumen",3,
 "Self-taught vs formal learning mix",
 ["RESUME","RECRUITER"],
 "Extracted from resume (bootcamp / course mentions); Recruiter asks 'How do you stay updated?'",
 "GROWTH_MINDSET","GROWTH_MINDSET",
 "Mix of formal + self-directed learning = structured and intrinsically motivated",
 "No evidence of any learning method = passive; relies only on employer to upskill",
 "Medium","M"),

("Skills","Skills Learning Acumen",3,
 "Breadth vs depth balance (T-shaped profile check)",
 ["LLM","BERT"],
 "Derived → BERT dna_fit label + LLM skill cluster analysis: specialist vs generalist ratio",
 "INNOVATION","INNOVATION",
 "T-shaped: deep in 1-2 primary skills + broad across adjacent areas = ideal for most senior roles",
 "Extremely narrow OR extremely scattered skill set without coherent theme",
 "Medium","M"),
]

# ── Communication and Presentation Skills ────────────────────────────────
ROWS += [
("Skills","Communication and Presentation Skills",5,
 "Clarity of 10-minute project walkthrough",
 ["PANEL"],
 "Panel scores candidate's 10-min project walkthrough: structure, clarity, problem-to-outcome arc",
 "EXEC_PRESENCE","EXEC_PRESENCE",
 "Clear story (problem → approach → outcome → learnings) = executive-level communication",
 "Rambling, technical jargon dump with no business context = communication risk for senior roles",
 "High","H"),

("Skills","Communication and Presentation Skills",5,
 "Structured thinking in Q&A answers",
 ["PANEL"],
 "Panel observes: does candidate answer with structure (MECE / STAR) or ramble uncontrolled?",
 "EXEC_PRESENCE","EXEC_PRESENCE",
 "Top-down structured answers = consulting / product-level business thinking",
 "Unstructured, unfocused answers = communication gap; risky for client-facing or lead roles",
 "High","H"),

("Skills","Communication and Presentation Skills",5,
 "Confidence and articulation under pressure",
 ["PANEL"],
 "Panel qualitative observation during full interview session",
 "EXEC_PRESENCE","EXEC_PRESENCE",
 "Confident, articulate, listens well, handles pushback = C-level meeting ready",
 "Hesitant, overly defensive, cannot simplify complex ideas = executive readiness gap",
 "High","H"),

("Skills","Communication and Presentation Skills",5,
 "Resume writing quality (action verbs / quantification / structure)",
 ["RESUME","LLM"],
 "Derived → LLM rates resume on clarity, action verb use, quantified impact, and visual structure",
 "EXEC_PRESENCE","EXEC_PRESENCE",
 "Well-structured resume with clear impact statements = professional written communication",
 "Poorly written, passive-voice, no quantification = may struggle with written biz communication",
 "Medium","M"),
]

# ── Domain Skills ─────────────────────────────────────────────────────────
ROWS += [
("Skills","Domain Skills",5,
 "Industry domain knowledge (panel Q&A)",
 ["PANEL"],
 "Panel asks domain-specific questions (e.g., 'How would a fraud detection model fail in FinTech?')",
 "DOMAIN_DEPTH","DOMAIN_DEPTH",
 "Accurate domain vocabulary + practical judgment in context = genuine domain expert",
 "Technical-only answers with no domain context = code implementer, not domain expert",
 "High","H"),

("Skills","Domain Skills",5,
 "Domain vocabulary fluency (natural use of domain terms)",
 ["PANEL"],
 "Panel observes if candidate uses domain-specific terms naturally without prompting",
 "DOMAIN_DEPTH","DOMAIN_DEPTH",
 "Speaks fluently in domain language = insider knowledge; worked deeply in this space",
 "Uses only tech terms with no domain language = surface exposure only",
 "High","H"),

("Skills","Domain Skills",5,
 "Cross-domain application (can transfer patterns to adjacent domains)",
 ["PANEL"],
 "Panel probes: 'How would your approach in X domain apply if the industry changed to Y?'",
 "INNOVATION","INNOVATION",
 "Can abstract and transfer patterns = creative, portable thinker",
 "Cannot think beyond own specific domain = rigid specialist; limited in new contexts",
 "Medium","M"),
]

# ── Project Explanation Skills ────────────────────────────────────────────
ROWS += [
("Skills","Project Explanation Skills",3,
 "Problem → Solution → Impact narrative quality",
 ["PANEL"],
 "Panel scores on whether candidate covers all 3 elements in 10-min walkthrough",
 "OWNERSHIP","OWNERSHIP",
 "Crisp P→S→I narrative = owns the project; understands business value of own work",
 "Describes only implementation steps with no problem framing or impact = execution-only mindset",
 "High","H"),

("Skills","Project Explanation Skills",3,
 "Technical depth in explanation (follow-up questions)",
 ["PANEL"],
 "Panel follow-up: 'Why did you choose X over Y?' / 'What would you do differently?'",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "Deep reasoning on design choices = genuine builder; actually wrote and owns this work",
 "Cannot answer follow-ups = may have exaggerated or copied the project",
 "High","H"),

("Skills","Project Explanation Skills",3,
 "Own role vs team clarity",
 ["PANEL"],
 "Panel probes: 'What specifically did YOU do vs what the team did?'",
 "OWNERSHIP","OWNERSHIP",
 "Clear personal contribution differentiated from team = honest, accountable professional",
 "Vague and cannot separate own work from team = inflating contribution; integrity concern",
 "High","H"),
]

# ── Problem Solving Skills ────────────────────────────────────────────────
ROWS += [
("Skills","Problem Solving Skills",3,
 "Approach to novel problems (structured decomposition)",
 ["PANEL"],
 "Panel presents new problem statement; observes if candidate breaks it down systematically",
 "INNOVATION","INNOVATION",
 "Structured decomposition of an unknown problem = strong analytical problem solver",
 "Jumps to solution without understanding the problem = reactive, not analytical",
 "High","H"),

("Skills","Problem Solving Skills",3,
 "Trade-off awareness (considers multiple solution options)",
 ["PANEL"],
 "Panel asks: 'What are the pros / cons of your approach vs alternatives you considered?'",
 "EXEC_PRESENCE","EXEC_PRESENCE",
 "Articulates clear trade-offs with context = senior-level engineering judgment",
 "Single-solution thinking with no trade-off awareness = junior mindset in senior candidate",
 "High","H"),

("Skills","Problem Solving Skills",3,
 "Live coding / case performance",
 ["PANEL"],
 "Panel runs structured coding problem or case from panel question database",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "Correct, clean, optimised solution with edge cases = strong practitioner",
 "Cannot solve moderate problem under time pressure = execution gap despite claims",
 "High","H"),
]

# ── Coding Skills ─────────────────────────────────────────────────────────
ROWS += [
("Skills","Coding Skills",None,
 "Data structures & algorithms knowledge",
 ["PANEL"],
 "Panel coding questions covering arrays, trees, graphs, DP etc.",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "Correct answers with time / space complexity analysis = strong CS fundamentals",
 "Cannot solve standard DS / Algo questions = foundational CS gap",
 "High","H"),

("Skills","Coding Skills",None,
 "Code quality and cleanliness",
 ["PANEL"],
 "Panel observes: variable naming, modularity, edge case handling in live coding session",
 "ETHICS","ETHICS",
 "Clean, readable, well-structured code = professional engineering standards and craft pride",
 "Messy, undocumented code with no edge cases = may ship poor quality to production",
 "High","H"),

("Skills","Coding Skills",None,
 "Problem-solving approach (talks through, tests assumptions, refines)",
 ["PANEL"],
 "Panel observes: does candidate verbalise approach, question assumptions, iterate?",
 "INNOVATION","INNOVATION",
 "Iterates, tests, refines with clear reasoning = mature collaborative engineering approach",
 "Writes code silently without explaining or testing = black-box style; hard to pair with",
 "Medium","M"),
]

# ── Conceptual Skills ─────────────────────────────────────────────────────
ROWS += [
("Skills","Conceptual Skills",None,
 "ML / AI theoretical understanding depth",
 ["PANEL"],
 "Panel asks conceptual questions: 'Explain backpropagation' / 'When would you choose LSTM vs Transformer?'",
 "TECHNICAL_DEPTH","TECHNICAL_DEPTH",
 "Accurate, deep theoretical answers = true ML practitioner; not just an API caller",
 "Knows how to call libraries but no theoretical understanding = surface-level ML",
 "High","H"),

("Skills","Conceptual Skills",None,
 "First-principles thinking (can derive, not just recall)",
 ["PANEL"],
 "Panel asks 'why' questions: 'Why does dropout prevent overfitting? Prove it intuitively.'",
 "INNOVATION","INNOVATION",
 "Can derive from first principles = deep, portable understanding of concepts",
 "Only memorised answers without underlying understanding = surface knowledge risk",
 "High","H"),

("Skills","Conceptual Skills",None,
 "Context and limitation awareness ('when NOT to use X')",
 ["PANEL"],
 "Panel asks: 'When would you NOT use Random Forest?' / 'What is the failure mode of this architecture?'",
 "EXEC_PRESENCE","EXEC_PRESENCE",
 "Knows limitations and context = mature practitioner; avoids over-engineering",
 "One-size-fits-all thinking with no awareness of tool limitations = limited practical wisdom",
 "High","H"),
]

# ═══════════════════════════════════════════════════════════════════════════
#  BUILD SHEET 1
# ═══════════════════════════════════════════════════════════════════════════
ws_main = wb.active
ws_main.title = "Data Points & Signals"
ws = ws_main

# Columns:
# 1:Module  2:Parameter  3:Max  4:Sub Data Point
# 5:Source(s)  6:Derived From
# 7:Signal?  8:Signal Name  9:Signal Category
# 10:Positive Reading  11:Negative / Red Flag
# 12:Score Impact  13:Priority
W([13, 30, 6, 32, 20, 42, 7, 22, 16, 38, 38, 20, 6])

# Header row
HDRS = ["Module","Parameter","Max\nPts","Sub Data Point",
        "Source(s)","Derived From",
        "Signal?","Signal Name","Signal Category",
        "Positive Signal Reading","Negative Signal / Red Flag",
        "Score Impact","Pri"]
HDR_BG = "1A237E"
for ci, h in enumerate(HDRS, 1):
    cl = ws.cell(row=1, column=ci, value=h)
    cl.fill = fill(HDR_BG); cl.font = fnt(True,"FFFFFF",9)
    cl.alignment = aln("center", True); cl.border = BM
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

r = 2
prev_mod = None
prev_param = None

for (module, param, max_pts, sub_dp,
     sources, derived_from,
     signal_name, signal_cat, positive, negative,
     score_impact, priority) in ROWS:

    # Section banner when module changes
    if module != prev_mod:
        sec_hdr(r, f"   {module.upper()}", MOD_CLR[module], cols=13)
        ws.row_dimensions[r].height = 18
        r += 1
        prev_mod = module
        prev_param = None

    # Param group header
    if param != prev_param:
        for ci in range(1, 14):
            cl = ws.cell(row=r, column=ci, value="")
            cl.fill = fill("EEEEEE"); cl.border = B
        ws.cell(row=r, column=2, value=param).fill = fill("EEEEEE")
        ws.cell(row=r, column=2).font = fnt(True,"333333",8)
        ws.cell(row=r, column=2).alignment = aln("left",False)
        ws.cell(row=r, column=2).border = B
        if max_pts:
            ws.cell(row=r, column=3, value=f"{max_pts}").fill = fill("EEEEEE")
            ws.cell(row=r, column=3).font = fnt(True,"555555",8)
            ws.cell(row=r, column=3).alignment = aln("center",False)
            ws.cell(row=r, column=3).border = B
        ws.row_dimensions[r].height = 14
        r += 1
        prev_param = param

    # Module cell
    mod_fg = MOD_CLR[module]
    cell(r,1, module, "F8F9FA", bold=True, fg=mod_fg, sz=8)

    # Parameter (small italic)
    cell(r,2, param, "FFFFFF", italic=True, fg="888888", sz=7)

    # Max pts
    pts_txt = str(max_pts) if max_pts else ""
    cell(r,3, pts_txt, "FFFFFF", ha="center", fg="555555", sz=8)

    # Sub data point – bold
    cell(r,4, sub_dp, "FFFFFF", bold=True, sz=9)

    # Source(s) – colour from primary
    primary = sources[0]
    s_bg, s_fg = SRC_CLR.get(primary, ("FFFFFF","000000"))
    src_txt = " + ".join(sources)
    cl = ws.cell(row=r, column=5, value=src_txt)
    cl.fill = fill(s_bg); cl.font = fnt(True, s_fg, 8)
    cl.alignment = aln("center", False); cl.border = B

    # Derived From – yellow tint, italic
    cell(r,6, derived_from, "FFFDE7", italic=True, fg="5D4037", sz=8)

    # Signal?
    has_signal = signal_name is not None
    s_val = "YES" if has_signal else ""
    s_bg2 = "C8E6C9" if has_signal else "F5F5F5"
    s_fg2 = "1B5E20" if has_signal else "BBBBBB"
    cl = ws.cell(row=r, column=7, value=s_val)
    cl.fill = fill(s_bg2); cl.font = fnt(True, s_fg2, 8)
    cl.alignment = aln("center", False); cl.border = B

    # Signal Name
    cell(r,8, signal_name or "", "FFFFFF", bold=True, fg="37474F", sz=8)

    # Signal Category – coloured
    sig_bg = SIG_CLR.get(signal_cat, "FFFFFF") if has_signal else "FFFFFF"
    cl = ws.cell(row=r, column=9, value=signal_cat or "")
    cl.fill = fill(sig_bg); cl.font = fnt(True,"333333",8)
    cl.alignment = aln("center", False); cl.border = B

    # Positive reading
    cell(r,10, positive or "", "F9FBF9" if has_signal else "FFFFFF",
         fg="2E7D32" if has_signal else "AAAAAA", sz=8)

    # Negative reading
    cell(r,11, negative or "", "FFF8F8" if has_signal else "FFFFFF",
         fg="B71C1C" if has_signal else "AAAAAA", sz=8)

    # Score impact
    imp_map = {"High":"FFEBEE","Medium":"FFF3E0","Low":"F1F8E9"}
    imp_bg = imp_map.get(score_impact.split("–")[0].strip() if score_impact else "", "FFFFFF")
    cell(r,12, score_impact, imp_bg, sz=8)

    # Priority
    pri_map = {"H":("FFCDD2","B71C1C"),"M":("FFF9C4","E65100"),"L":("F1F8E9","2E7D32")}
    p_bg, p_fg = pri_map.get(priority, ("FFFFFF","000000"))
    cl = ws.cell(row=r, column=13, value=priority)
    cl.fill = fill(p_bg); cl.font = fnt(True, p_fg, 9)
    cl.alignment = aln("center", False); cl.border = B

    ws.row_dimensions[r].height = 44
    r += 1

# ═══════════════════════════════════════════════════════════════════════════
#  SHEET 2: SIGNAL LEGEND
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Signal Legend")
ws = ws2
W([18, 42, 40, 50, 6])

LEG_HDRS = ["Signal Category","What it tells you about the person",
            "Key data points that produce it",
            "Positive hire action  |  Red flag action", "Colour"]
for ci, h in enumerate(LEG_HDRS, 1):
    cl = ws.cell(row=1, column=ci, value=h)
    cl.fill = fill("1A237E"); cl.font = fnt(True,"FFFFFF",10)
    cl.alignment = aln("center",True); cl.border = BM
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"

LEGEND = [
("LOYALTY",
 "How committed is this person to an employer? Do they go deep before moving?",
 "Tenure per company (Python), Promoted within same company (LLM), Intra-company title growth",
 "POSITIVE: Long tenures + internal promotions → place in senior, long-commitment role\n"
 "RED FLAG: Avg tenure < 12 months OR 3+ short stints → flight risk; likely to leave in < 1 year"),

("TRAJECTORY",
 "Is this person on a deliberate upward path — scope, title, complexity all growing?",
 "BERT career_progression label, LLM title tier mapping, Lateral vs vertical move analysis",
 "POSITIVE: FAST_TRACK / GROWING label → prioritise; trajectory may outpace role\n"
 "RED FLAG: DECLINING label or all-lateral career → stagnation; investigate before proceeding"),

("ETHICS",
 "Can this person be trusted — do they hold themselves and peers to high standards?",
 "Code review participation (Recruiter), Interview panel trust (Recruiter), Reason for leaving (LLM), Code quality (Panel)",
 "POSITIVE: Active code reviewer + interview panelist → trusted by employer; strong internal trust vote\n"
 "RED FLAG: Repeated 'conflict with management' exits OR messy code in panel → ethics / culture concern"),

("AMBITION",
 "Does this person want to grow into bigger, more impactful roles?",
 "LLM title progression (IC→EM→Director), Award frequency (LLM), BERT career_progression",
 "POSITIVE: Clear manager / director track + multiple awards → place in growth role with clear path\n"
 "RED FLAG: 10+ years with no scope expansion → comfort-zone player; may resist stretch assignments"),

("FLIGHT_RISK",
 "How likely is this person to leave within 12–18 months of joining?",
 "Python: companies in last 5 yrs, short stints count (<12 mo); Recruiter: reason for leaving",
 "POSITIVE: Stable long-tenure history → low risk; likely to stay and deliver\n"
 "RED FLAG: 4+ companies in 5 years OR 3+ stints <12 mo → flag before extending offer; escalate to hiring manager"),

("GROWTH_MINDSET",
 "Does this person actively invest in their own learning — not just what the employer provides?",
 "LLM: new skills per year, course recency; Recruiter: LinkedIn activity, learning method; Python: course date",
 "POSITIVE: 2+ new skills/yr + recent relevant course → high adaptability for fast-moving tech roles\n"
 "RED FLAG: Same skill set 5+ years + zero courses + no community involvement → may not keep pace"),

("OWNERSHIP",
 "Does this person take full accountability — or do they wait for instructions?",
 "LLM: action verbs in project bullets ('led','owned','designed' vs 'assisted'); Recruiter: team size; Panel: role clarity",
 "POSITIVE: Owner / lead verbs across both projects + clear personal contribution → give high-accountability roles\n"
 "RED FLAG: Only peripheral verbs in both projects → needs guided structure; not ready for independent delivery"),

("BUSINESS_ORIENT",
 "Does this person think in business outcomes, not just tasks completed?",
 "LLM: quantified impact in projects, stakeholder level; Panel: P→S→I narrative; Resume quality (LLM)",
 "POSITIVE: Metrics-rich resume + business narrative in panel → can communicate ROI to business stakeholders\n"
 "RED FLAG: All technical bullets with zero business outcomes + no metrics → pure executor; not a business thinker"),

("PRESTIGE",
 "What quality signal does the company / college / certification tier give as a prior?",
 "Company Tier (Python + LLM), Institute Tier (education_engine.py), Cert issuing body (LLM), Award type",
 "POSITIVE: Tier-1 company + IIT/IIM + AWS Professional → high prior probability of quality; bar already tested\n"
 "RED FLAG: All Tier-4/5 companies + Tier-3 college + no certs → validate thoroughly via panel; no quality prior"),

("TECHNICAL_DEPTH",
 "How deep is the actual technical skill — not claimed but verified through evidence?",
 "BERT skill_depth classifier, Evidence level (rubric_engine.py), Panel depth Q&A, Live coding session",
 "POSITIVE: ARCHITECT_LEVEL on primary skills confirmed by BERT + Panel → can independently design and lead\n"
 "RED FLAG: AWARENESS / FOUNDATIONAL on a mandatory JD skill → hard rejection signal; do not proceed"),

("PASSION_CRAFT",
 "Does this person love the craft enough to go beyond what is required?",
 "Resume: GitHub link, OSS contributions, LeetCode; LLM: GitHub activity assessment; Resume: Kaggle rank",
 "POSITIVE: Active OSS contributions + community presence + side projects → self-motivated; will stay sharp\n"
 "RED FLAG: No GitHub, no community, no learning outside work → 9-to-5 only; may stagnate in senior roles"),

("RESILIENCE",
 "Has this person faced adversity and responded constructively?",
 "Recruiter: break reason + activity during break; LLM: short stint context; Resume: education gap reason",
 "POSITIVE: Clear break reason + productive re-entry with evidence → bounce-back capacity demonstrated\n"
 "RED FLAG: Unexplained multi-month gap + no explanation offered to recruiter → investigate before offer"),

("INNOVATION",
 "Does this person create new ideas, approaches, or inventions — not just implement existing ones?",
 "Resume: patent status (granted), publications; Panel: architecture thinking, system design; problem-solving approach",
 "POSITIVE: Granted patents OR novel system design in panel → creative builder; generates new value\n"
 "RED FLAG: Only ever implements others' designs; never proposes new approaches = execution-only ceiling"),

("LEADERSHIP",
 "Can this person multiply others — mentor, guide, build team capability?",
 "LLM: mentoring verbs in resume; Recruiter: interview panel participation, mentee count; ECA leadership roles",
 "POSITIVE: Mentored 3+ people + conducts interviews → team multiplier; ready for lead or manager track\n"
 "RED FLAG: 8+ years with zero mentoring or leadership signal → IC ceiling; will not scale a team"),

("EXEC_PRESENCE",
 "Can this person operate at the business / executive level with authority and clarity?",
 "BERT: stakeholder level; LLM: business language in resume; Panel: 10-min walkthrough, structured Q&A",
 "POSITIVE: C-level interaction evidence + structured panel communication → senior / exec role ready today\n"
 "RED FLAG: Never engaged above manager level + poor panel communication → needs development before senior placement"),

("COLLABORATION",
 "Does this person work well with others across functions, organisations, and cultures?",
 "LLM: cross-functional mentions, external client work; Recruiter: time-zone team experience; Panel: framing of 'we' vs 'I'",
 "POSITIVE: Cross-functional + international + client experience → high EQ; brings others along\n"
 "RED FLAG: Always worked solo, no cross-team mentions, no client work → lone wolf risk in team-centric role"),

("GLOBAL_READY",
 "Is this person equipped and comfortable in global, cross-cultural roles?",
 "LLM: countries worked, intl assignment duration; Recruiter: remote cross-border collab; Resume: intl client names",
 "POSITIVE: 12+ months international + multiple country experience → genuinely global-ready\n"
 "RED FLAG: Domestic-only career for a global team lead role → significant exposure gap; flag for hiring manager"),

("DOMAIN_DEPTH",
 "How deeply does this person understand the industry domain — not just the technology?",
 "LLM: company industry, project domain vocab, field-of-study; Panel: domain Q&A, vocabulary fluency",
 "POSITIVE: Same domain across career + domain-fluent in panel → go-to domain expert; irreplaceable in niche roles\n"
 "RED FLAG: Domain-hopping every role + no domain vocabulary in panel → generalist; not suitable for specialist JDs"),

("DNA_FIT",
 "Does this person's background match the cultural DNA of the hiring org (Product / Consulting / Platform)?",
 "LLM: company type (Product/Service/Startup); BERT: dna_fit label (CONSULTING/PRODUCT/PLATFORM_INFRA/DOMAIN_SPECIALIST); BERT: project_type",
 "POSITIVE: BERT dna_fit = PRODUCT for a product-first org → natural culture fit; onboarding will be smooth\n"
 "RED FLAG: Only services / body-shop background for a product org → culture mismatch risk; need deep panel validation"),
]

for i, (cat, meaning, key_dps, actions) in enumerate(LEGEND, 2):
    sig_bg = SIG_CLR.get(cat, "FFFFFF")

    cl = ws.cell(row=i, column=1, value=cat)
    cl.fill = fill(sig_bg); cl.font = fnt(True,"212121",9)
    cl.alignment = aln("center",False); cl.border = B

    cl = ws.cell(row=i, column=2, value=meaning)
    cl.fill = fill("FFFFFF"); cl.font = fnt(size=9)
    cl.alignment = aln("left",True); cl.border = B

    cl = ws.cell(row=i, column=3, value=key_dps)
    cl.fill = fill("F8F9FA"); cl.font = fnt(size=8, italic=True)
    cl.alignment = aln("left",True); cl.border = B

    cl = ws.cell(row=i, column=4, value=actions)
    cl.fill = fill("FFFFFF"); cl.font = fnt(size=8)
    cl.alignment = aln("left",True); cl.border = B

    cl = ws.cell(row=i, column=5, value="")
    cl.fill = fill(sig_bg); cl.border = B

    ws.row_dimensions[i].height = 60

# ── SHEET 3: SOURCE LEGEND ─────────────────────────────────────────────────
ws3 = wb.create_sheet("Source Legend")
ws = ws3
W([16, 50, 50])

SRC_HDRS = ["Source Tag","What it means","How it feeds the scoring pipeline"]
for ci, h in enumerate(SRC_HDRS, 1):
    cl = ws.cell(row=1, column=ci, value=h)
    cl.fill = fill("37474F"); cl.font = fnt(True,"FFFFFF",10)
    cl.alignment = aln("center",True); cl.border = BM
ws.row_dimensions[1].height = 22

SOURCE_LEGEND = [
("RESUME",
 "Raw text extracted directly from the candidate's uploaded PDF / DOCX resume. No inference — fact as stated.",
 "Feeds NER, date parser, skill extractor, regex patterns. Highest factual reliability. If missing, escalate to RECRUITER."),

("PYTHON",
 "A value computed by Python code from other extracted data points (arithmetic, aggregation, date math).",
 "Examples: End Date − Start Date → Tenure; mean(tenures) → Avg Tenure; count(stints < 12mo) → Job-hop count."),

("LLM",
 "A value inferred by the Large Language Model (LLM judge) reading the resume text and applying reasoning.",
 "Examples: Company type, domain vocabulary score, ownership verb detection, business impact extraction. Reliability = Medium-High; confidence-blended with BERT where applicable."),

("BERT",
 "A value produced by a fine-tuned BERT classifier trained on labelled resume data.",
 "Tasks: skill_depth, role_family, dna_fit, career_progression, stakeholder_management, mentorship_signal, project_type. Outputs a label + confidence score; blended with LLM at low confidence."),

("RECRUITER",
 "A value that must be filled in by the recruiter during the screening call — not derivable from resume alone.",
 "Triggers when Python detects a missing data point (e.g., break reason, mentee count, LinkedIn activity, reasons for leaving). Recruiter fills a structured form. Feeds Score(Recruiter) column."),

("PANEL",
 "A value observed or assessed by the interview panel during the structured panel interview.",
 "Inputs: panel feedback form scored by LLM + rubric. Covers: communication, domain depth, coding, conceptual skills, project walkthrough quality. Feeds Score(Panel) column."),
]

for i, (tag, meaning, pipeline) in enumerate(SOURCE_LEGEND, 2):
    src_bg, src_fg = SRC_CLR.get(tag, ("FFFFFF","000000"))

    cl = ws.cell(row=i, column=1, value=tag)
    cl.fill = fill(src_bg); cl.font = fnt(True, src_fg, 10)
    cl.alignment = aln("center",False); cl.border = B

    cl = ws.cell(row=i, column=2, value=meaning)
    cl.fill = fill("FFFFFF"); cl.font = fnt(size=9)
    cl.alignment = aln("left",True); cl.border = B

    cl = ws.cell(row=i, column=3, value=pipeline)
    cl.fill = fill("F8F9FA"); cl.font = fnt(size=9, italic=True)
    cl.alignment = aln("left",True); cl.border = B

    ws.row_dimensions[i].height = 55

out = "E:/Dev/resume_intelligence/Resume_DataPoints_Signals.xlsx"
wb.save(out)
print(f"Saved: {out}")
sig_count = sum(1 for row in ROWS if row[6] is not None)
print(f"Total sub data points: {len(ROWS)}")
print(f"Sub data points WITH signals: {sig_count}")
print(f"Sub data points WITHOUT signals (raw data only): {len(ROWS) - sig_count}")
